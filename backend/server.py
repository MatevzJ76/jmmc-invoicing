from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Form, status, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from argon2 import PasswordHasher
from argon2.exceptions import VerifyMismatchError
import jwt
from collections import defaultdict
import openpyxl
import xlrd
from io import BytesIO
import litellm
litellm.drop_params = True  # ignore unsupported params silently

class UserMessage:
    """Compatibility shim replacing emergentintegrations.llm.chat.UserMessage"""
    def __init__(self, text: str):
        self.text = text

class LlmChat:
    """Compatibility shim replacing emergentintegrations.llm.chat.LlmChat"""
    def __init__(self, api_key: str, session_id: str = "", system_message: str = ""):
        self.api_key = api_key
        self.session_id = session_id
        self.system_message = system_message
        self._provider = "anthropic"
        self._model = "claude-3-5-sonnet-20241022"

    def with_model(self, provider: str, model: str) -> "LlmChat":
        self._provider = provider
        self._model = model
        return self

    # Map only Emergent-specific / non-standard model names to valid models.
    # Standard OpenAI names (gpt-4o, gpt-4-turbo, etc.) are NOT remapped so they
    # are sent directly to OpenAI when the user selects the OpenAI provider.
    _MODEL_MAP = {
        "gpt-5-nano":  "claude-3-haiku-20240307",
        "gpt-5-mini":  "claude-3-5-haiku-20241022",
        "gpt-5":       "claude-3-5-sonnet-20241022",
        "gpt-4-mini":  "claude-3-5-haiku-20241022",  # not a real OpenAI name
    }

    async def send_message(self, message: UserMessage) -> str:
        original_model = self._model
        model = self._MODEL_MAP.get(original_model, original_model)
        provider = self._provider
        # Only force anthropic provider when the model was remapped from an
        # Emergent-specific name — do NOT override if the user explicitly chose OpenAI.
        if model != original_model and model.startswith("claude"):
            provider = "anthropic"
        # Map provider + model to LiteLLM model string
        if provider == "anthropic":
            litellm_model = f"anthropic/{model}"
        elif provider == "google":
            litellm_model = f"gemini/{model}"
        else:
            litellm_model = model  # openai models use name directly

        msgs = []
        if self.system_message:
            msgs.append({"role": "system", "content": self.system_message})
        msgs.append({"role": "user", "content": message.text})

        response = await litellm.acompletion(
            model=litellm_model,
            messages=msgs,
            api_key=self.api_key,
        )
        return response.choices[0].message.content or ""

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Password hasher
ph = PasswordHasher()

# JWT config
SECRET_KEY = os.environ.get("JWT_SECRET", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 15
REFRESH_TOKEN_EXPIRE_DAYS = 7

# eRačuni config
ERACUNI_MODE = os.environ.get("ERACUNI_MODE", "stub")
ERACUNI_BASE_URL = os.environ.get("ERACUNI_BASE_URL", "")
ERACUNI_API_KEY = os.environ.get("ERACUNI_API_KEY", "")

# LLM config
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

@app.on_event("startup")
async def ensure_indexes():
    """Create MongoDB indexes for performance-critical queries."""
    await db.importBatches.create_index("id",         background=True)
    await db.invoices.create_index("batchId",         background=True)
    await db.invoices.create_index("id",              background=True)
    await db.timeEntries.create_index("batchId",      background=True)
    await db.timeEntries.create_index("id",           background=True)
    await db.customers.create_index("id",             background=True)
    await db.projects.create_index("id",              background=True)
    await db.invoiceLines.create_index("timeEntryId", background=True)
    await db.invoiceLines.create_index("invoiceId",   background=True)
security = HTTPBearer()

# Rate limiting store (simple in-memory) - 10 attempts per 15 minutes
login_attempts = defaultdict(list)

# ============ MODELS ============
class User(BaseModel):
    email: str
    role: str  # ADMIN or USER
    mustReset: bool = False
    status: str = "active"  # active or archived
    username: Optional[str] = None

class LoginRequest(BaseModel):
    email: str
    password: str

class ChangePasswordRequest(BaseModel):
    currentPassword: str
    newPassword: str

class CreateUserRequest(BaseModel):
    email: EmailStr
    username: str
    password: str
    role: str  # ADMIN or USER

class UpdateUserRoleRequest(BaseModel):
    role: str  # ADMIN or USER

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    user: User

class ImportMetadata(BaseModel):
    invoiceDate: str
    periodFrom: str
    periodTo: str
    dueDate: str

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    customerId: str
    customerName: str
    number: Optional[str] = None
    invoiceDate: str
    periodFrom: str
    periodTo: str
    dueDate: str
    status: str
    total: float
    externalNumber: Optional[str] = None

class InvoiceLine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    invoiceId: str
    description: str
    quantity: float
    unitPrice: float
    amount: float
    taxCode: Optional[str] = None

class InvoiceUpdate(BaseModel):
    number: Optional[str] = None
    invoiceDate: Optional[str] = None
    dueDate: Optional[str] = None
    periodFrom: Optional[str] = None
    periodTo: Optional[str] = None
    lines: List[Dict[str, Any]]

class AIRequest(BaseModel):
    text: str
    feature: str  # grammar, fraud, gdpr

class AISettings(BaseModel):
    aiProvider: str = "custom"
    customApiKey: Optional[str] = None
    customModel: str = "claude-3-5-sonnet-20241022"

    # AI Prompts with individual model selection
    grammarPrompt: str = "Correct any grammar errors, spelling mistakes, and improve the clarity of this time entry description. Return ONLY the corrected text without any explanations, comments, or additional formatting."
    grammarModel: str = "claude-3-haiku-20240307"       # Fast, low-cost for simple edits

    fraudPrompt: str = "Analyze this time entry for fraud indicators: 1) Suspicious hours (e.g., 20 hours claimed for a simple 1-hour task), 2) Vague or generic descriptions (e.g., \"work done\", \"various tasks\"), 3) Unusual patterns or inconsistencies. If suspicious, explain the concern. If normal, respond with \"No issues detected\"."
    fraudModel: str = "claude-3-5-haiku-20241022"       # Balanced for fraud detection

    gdprPrompt: str = "Check this time entry description for GDPR compliance. Identify any personal data such as: employee full names (should be initials only), personal email addresses, phone numbers, or other identifying information. If found, return the text with sensitive data replaced by initials or [MASKED]. If compliant, return the original text unchanged."
    gdprModel: str = "claude-3-5-haiku-20241022"        # Balanced for compliance checks

    verificationPrompt: str = "Perform a general verification check on this time entry. Look for: 1) Data quality issues, 2) Missing information, 3) Formatting problems, 4) Business logic violations, 5) Any other anomalies. Provide specific feedback on what needs attention, or respond with \"Entry looks good\" if no issues found."
    verificationModel: str = "claude-3-5-sonnet-20241022"  # Deep reasoning for complex checks

    dtmPrompt: str = "Do the magic on this data. Enhance, optimize, and improve it in the best way possible."
    dtmModel: str = "claude-3-5-sonnet-20241022"        # Deep reasoning for magic enhancements
    
    eracuniEndpoint: Optional[str] = "https://e-racuni.com/WebServicesSI/API"
    eracuniUsername: Optional[str] = None
    eracuniSecretKey: Optional[str] = None
    eracuniToken: Optional[str] = None
    testPrompt: Optional[str] = None  # For testing AI connection quality

class Article(BaseModel):
    code: str
    description: str
    unitMeasure: str
    priceWithoutVAT: float
    vatPercentage: float
    tariffCode: str = ""  # For future mapping to tariff codes

class Employee(BaseModel):
    employee_name: str
    cost: Optional[float] = None  # Can be null/empty
    archived: bool = False
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

class Tariff(BaseModel):
    code: str  # e.g., "001 - Računovodstvo"
    description: str  # Full description
    value: Optional[float] = 0.0  # EUR value
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

# ============ AUTH HELPERS ============
def create_token(data: dict, expires_delta: timedelta):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + expires_delta
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user_doc = await db.users.find_one({"email": email}, {"_id": 0, "passwordHash": 0})
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")
        
        return User(**user_doc)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

def check_rate_limit(email: str) -> bool:
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(minutes=15)
    login_attempts[email] = [t for t in login_attempts[email] if t > cutoff]
    return len(login_attempts[email]) < 10  # Increased from 5 to 10

# ============ AUTH ENDPOINTS ============
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(request: LoginRequest):
    if not check_rate_limit(request.email):
        raise HTTPException(status_code=429, detail="Too many login attempts. Please try again in 15 minutes.")
    
    user_doc = await db.users.find_one({"email": request.email})
    if not user_doc:
        login_attempts[request.email].append(datetime.now(timezone.utc))  # Only count failed attempts
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is archived
    if user_doc.get("status", "active") == "archived":
        login_attempts[request.email].append(datetime.now(timezone.utc))
        raise HTTPException(status_code=401, detail="Account is archived. Please contact administrator.")
    
    try:
        ph.verify(user_doc["passwordHash"], request.password)
    except VerifyMismatchError:
        login_attempts[request.email].append(datetime.now(timezone.utc))  # Only count failed attempts
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Successful login - clear attempts for this email
    if request.email in login_attempts:
        login_attempts[request.email] = []
    
    access_token = create_token({"sub": request.email}, timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    refresh_token = create_token({"sub": request.email}, timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS))
    
    user = User(
        email=user_doc["email"], 
        role=user_doc["role"], 
        mustReset=user_doc.get("mustReset", False),
        status=user_doc.get("status", "active"),
        username=user_doc.get("username")
    )
    return TokenResponse(access_token=access_token, refresh_token=refresh_token, user=user)

@api_router.post("/auth/refresh", response_model=TokenResponse)
async def refresh(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        
        user_doc = await db.users.find_one({"email": email})
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Check if user is archived
        if user_doc.get("status", "active") == "archived":
            raise HTTPException(status_code=401, detail="Account is archived")
        
        access_token = create_token({"sub": email}, timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
        refresh_token = create_token({"sub": email}, timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS))
        
        user = User(
            email=user_doc["email"], 
            role=user_doc["role"], 
            mustReset=user_doc.get("mustReset", False),
            status=user_doc.get("status", "active"),
            username=user_doc.get("username")
        )
        return TokenResponse(access_token=access_token, refresh_token=refresh_token, user=user)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

@api_router.post("/auth/change-password")
async def change_password(request: ChangePasswordRequest, current_user: User = Depends(get_current_user)):
    user_doc = await db.users.find_one({"email": current_user.email})
    
    try:
        ph.verify(user_doc["passwordHash"], request.currentPassword)
    except VerifyMismatchError:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    new_hash = ph.hash(request.newPassword)
    await db.users.update_one(
        {"email": current_user.email},
        {"$set": {"passwordHash": new_hash, "mustReset": False}}
    )
    
    return {"message": "Password changed successfully"}

@api_router.post("/auth/clear-rate-limit")
async def clear_rate_limit(email: str, current_user: User = Depends(get_current_user)):
    """Clear rate limiting for a specific email (admin debugging)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin role required")
    
    if email in login_attempts:
        login_attempts[email] = []
    
    return {"message": f"Rate limit cleared for {email}"}

# ============ USER MANAGEMENT ENDPOINTS (ADMIN) ============
@api_router.get("/user/profile")
async def get_user_profile(current_user: User = Depends(get_current_user)):
    """Get current user's profile"""
    user_doc = await db.users.find_one({"email": current_user.email})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "email": user_doc["email"],
        "username": user_doc.get("username"),
        "role": user_doc["role"],
        "status": user_doc.get("status", "active"),
        "createdAt": user_doc.get("createdAt")
    }

@api_router.get("/admin/users")
async def list_users(current_user: User = Depends(get_current_user)):
    """List all users (admin only)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}).to_list(length=None)
    
    return [{
        "id": user.get("id"),
        "email": user["email"],
        "username": user.get("username"),
        "role": user["role"],
        "status": user.get("status", "active"),
        "createdAt": user.get("createdAt"),
        "mustReset": user.get("mustReset", False)
    } for user in users]

@api_router.post("/admin/users")
async def create_user(request: CreateUserRequest, current_user: User = Depends(get_current_user)):
    """Create a new user (admin only)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate password strength
    if len(request.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters long")
    if not any(c.isupper() for c in request.password):
        raise HTTPException(status_code=400, detail="Password must contain at least one uppercase letter")
    if not any(c.islower() for c in request.password):
        raise HTTPException(status_code=400, detail="Password must contain at least one lowercase letter")
    if not any(c.isdigit() for c in request.password):
        raise HTTPException(status_code=400, detail="Password must contain at least one number")
    if not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in request.password):
        raise HTTPException(status_code=400, detail="Password must contain at least one special character")
    
    # Check if user already exists
    existing = await db.users.find_one({"email": request.email})
    if existing:
        raise HTTPException(status_code=400, detail="User with this email already exists")
    
    # Validate role
    if request.role not in ["ADMIN", "USER"]:
        raise HTTPException(status_code=400, detail="Role must be either ADMIN or USER")
    
    # Create user
    user_id = str(uuid.uuid4())
    password_hash = ph.hash(request.password)
    
    new_user = {
        "id": user_id,
        "email": request.email,
        "username": request.username,
        "role": request.role,
        "status": "active",
        "passwordHash": password_hash,
        "mustReset": False,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    
    return {
        "id": user_id,
        "email": request.email,
        "username": request.username,
        "role": request.role,
        "status": "active",
        "message": "User created successfully"
    }

@api_router.put("/admin/users/{user_id}/archive")
async def archive_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Archive a user (admin only)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find user by ID
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent archiving self
    if user_doc["email"] == current_user.email:
        raise HTTPException(status_code=400, detail="Cannot archive your own account")
    
    # Archive user
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"status": "archived", "archivedAt": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "User archived successfully"}

@api_router.put("/admin/users/{user_id}/role")
async def update_user_role(user_id: str, request: UpdateUserRoleRequest, current_user: User = Depends(get_current_user)):
    """Update user role (admin only)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate role
    if request.role not in ["ADMIN", "USER"]:
        raise HTTPException(status_code=400, detail="Role must be either ADMIN or USER")
    
    # Find user by ID
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent changing own role
    if user_doc["email"] == current_user.email:
        raise HTTPException(status_code=400, detail="Cannot change your own role")
    
    # Update role
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": request.role}}
    )
    
    return {"message": f"User role updated to {request.role}"}

@api_router.post("/admin/cleanup-duplicate-customers")
async def cleanup_duplicate_customers(
    dry_run: bool = True,
    current_user: User = Depends(get_current_user)
):
    """
    Delete customer rows that have NO assigned company (JMMC HP / JMMC Finance)
    AND have no historicalInvoices. Pass dry_run=false to actually delete.
    """
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")

    # Get all valid company IDs (JMMC HP and JMMC Finance)
    valid_companies = await db.companies.find(
        {"name": {"$in": ["JMMC HP d.o.o.", "JMMC Finance d.o.o."]}}
    ).to_list(None)
    valid_company_ids = {c["id"] for c in valid_companies}
    logger.info(f"Valid company IDs: {valid_company_ids}")

    # Find customers to delete:
    #   - companyId not in valid set (missing, null, or unknown company)
    #   - historicalInvoices is empty or missing
    all_customers = await db.customers.find({}).to_list(None)

    to_delete = []
    to_keep = []
    for c in all_customers:
        has_valid_company = c.get("companyId") in valid_company_ids
        has_history = bool(c.get("historicalInvoices"))
        if not has_valid_company and not has_history:
            to_delete.append({"id": c["id"], "name": c["name"], "companyId": c.get("companyId")})
        else:
            to_keep.append(c["name"])

    if not dry_run and to_delete:
        # Delete one by one using the custom 'id' field — most reliable approach
        deleted = 0
        for c in to_delete:
            result = await db.customers.delete_one({"id": c["id"]})
            deleted += result.deleted_count
        logger.info(f"Deleted {deleted} / {len(to_delete)} customers")
    else:
        deleted = 0

    return {
        "dry_run": dry_run,
        "would_delete_count": len(to_delete),
        "actually_deleted": deleted,
        "to_delete": to_delete,
        "kept_count": len(to_keep),
    }

# ============ IMPORT ENDPOINTS ============
def load_excel_file(contents: bytes, filename: str):
    """Load Excel file (supports both .xls and .xlsx formats)"""
    if filename.endswith('.xls'):
        # Old Excel format (.xls) - use xlrd
        workbook = xlrd.open_workbook(file_contents=contents)
        sheet = workbook.sheet_by_index(0)
        
        # Convert xlrd sheet to a format similar to openpyxl for compatibility
        class XLSSheet:
            def __init__(self, xlrd_sheet):
                self.xlrd_sheet = xlrd_sheet
                
            def iter_rows(self, min_row=1, values_only=False):
                for row_idx in range(min_row - 1, self.xlrd_sheet.nrows):
                    if values_only:
                        yield self.xlrd_sheet.row_values(row_idx)
                    else:
                        yield [self.xlrd_sheet.cell(row_idx, col_idx) for col_idx in range(self.xlrd_sheet.ncols)]
            
            def __getitem__(self, row_num):
                row_idx = row_num - 1
                return [self.xlrd_sheet.cell(row_idx, col_idx) for col_idx in range(self.xlrd_sheet.ncols)]
        
        return XLSSheet(sheet), workbook
    else:
        # New Excel format (.xlsx) - use openpyxl
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        return sheet, workbook

@api_router.post("/imports")
async def import_xlsx(
    file: UploadFile = File(...),
    title: str = Form(...),
    invoiceDate: str = Form(...),
    periodFrom: str = Form(...),
    periodTo: str = Form(...),
    dueDate: str = Form(...),
    saveAsProgress: str = Form(None),
    current_user: User = Depends(get_current_user)
):
    try:
        contents = await file.read()
        sheet, wb = load_excel_file(contents, file.filename)
        
        # Validate headers - strip None values and check
        expected_headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št. računa"]
        alternative_headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št.računa"]
        
        # Get headers - handle both xlrd and openpyxl formats
        if file.filename.endswith('.xls'):
            raw_headers = sheet.xlrd_sheet.row(0)  # xlrd: get first row (row 0)
            headers = [cell.value for cell in raw_headers if cell.value is not None and cell.value != '#']
        else:
            raw_headers = [cell.value for cell in sheet[1]]
            headers = [h for h in raw_headers if h is not None and h != '#']
        
        if headers != expected_headers and headers != alternative_headers:
            raise HTTPException(status_code=400, detail=f"Invalid Excel headers. Expected: {expected_headers}, Got: {headers}")
        
        # Create batch with appropriate status
        batch_status = "in progress" if saveAsProgress == "true" else "imported"
        
        batch_id = str(uuid.uuid4())
        batch_doc = {
            "id": batch_id,
            "title": title,
            "filename": file.filename,
            "periodFrom": periodFrom,
            "periodTo": periodTo,
            "invoiceDate": invoiceDate,
            "dueDate": dueDate,
            "status": batch_status,
            "createdBy": current_user.email,
            "createdAt": datetime.now(timezone.utc).isoformat()
        }
        await db.importBatches.insert_one(batch_doc)
        
        # Fetch tariff codes to calculate hourly rates
        tariff_codes = await db.tariffs.find({}, {"_id": 0}).to_list(1000)
        tariff_rates = {t["code"]: t.get("value", 0) for t in tariff_codes}
        
        # --- PASS 1: parse raw rows, collect unique customer/project names ---
        raw_rows = []
        current_project = "General"
        unique_customer_names = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None or cell == '' for cell in row):
                continue
            if row[0] and (str(row[0]).endswith('.') or row[0] == '#'):
                row_data = row[1:11]
            else:
                row_data = row[0:10]
            projekt_val = row_data[0]
            stranka_val = row_data[1]
            datum_val   = row_data[2]
            if not datum_val:
                continue
            if projekt_val and str(projekt_val).strip():
                current_project = str(projekt_val).strip()
            customer_name = str(stranka_val).strip() if stranka_val and str(stranka_val).strip() else None
            if customer_name:
                unique_customer_names.add(customer_name)
            raw_rows.append({
                "projekt": current_project,
                "customer_name": customer_name,
                "datum": datum_val,
                "tariff": row_data[3],
                "employee": row_data[4],
                "notes": row_data[5],
                "hours_val": row_data[6],
            })

        # --- PASS 2: batch-resolve customers (2 queries total) ---
        existing_customers = await db.customers.find(
            {"name": {"$in": list(unique_customer_names)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(None)
        customer_map = {c["name"]: c["id"] for c in existing_customers}

        new_customers_created = []
        new_customer_docs = []
        for name in unique_customer_names:
            if name not in customer_map:
                cid = str(uuid.uuid4())
                customer_map[name] = cid
                doc = {"id": cid, "name": name, "status": "new", "unitPrice": 0, "historicalInvoices": []}
                new_customer_docs.append(doc)
                new_customers_created.append({"id": cid, "name": name})
        if new_customer_docs:
            await db.customers.insert_many(new_customer_docs)

        # --- PASS 3: batch-resolve projects (2 queries total) ---
        # Build set of (project_name, customer_id) pairs needed
        needed_projects = set()
        for r in raw_rows:
            cid = customer_map.get(r["customer_name"]) if r["customer_name"] else None
            needed_projects.add((r["projekt"], cid))

        project_names = list({p for p, _ in needed_projects})
        existing_projects = await db.projects.find(
            {"name": {"$in": project_names}}, {"_id": 0, "id": 1, "name": 1, "customerId": 1}
        ).to_list(None)
        project_map = {(p["name"], p.get("customerId")): p["id"] for p in existing_projects}

        new_project_docs = []
        for proj_name, cid in needed_projects:
            if (proj_name, cid) not in project_map:
                pid = str(uuid.uuid4())
                project_map[(proj_name, cid)] = pid
                new_project_docs.append({"id": pid, "name": proj_name, "customerId": cid})
        if new_project_docs:
            await db.projects.insert_many(new_project_docs)

        # --- PASS 4: build entries using resolved maps (no DB queries) ---
        entries = []
        for r in raw_rows:
            customer_name = r["customer_name"]
            customer_id   = customer_map.get(customer_name) if customer_name else None
            project_id    = project_map.get((r["projekt"], customer_id))

            try:
                hours = round(float(str(r["hours_val"]).replace(',', '.')) if r["hours_val"] else 0.0, 2)
            except:
                hours = 0.0

            tariff_code   = str(r["tariff"]) if r["tariff"] else "N/A"
            hourly_rate   = tariff_rates.get(tariff_code, 0)
            calculated_value = round(hours * hourly_rate, 2)
            datum_val     = r["datum"]

            entry = {
                "id": str(uuid.uuid4()),
                "batchId": batch_id,
                "projectId": project_id,
                "customerId": customer_id,
                "employeeName": r["employee"] or "Unknown",
                "date": datum_val.isoformat() if hasattr(datum_val, 'isoformat') else str(datum_val),
                "hours": hours,
                "tariff": tariff_code,
                "hourlyRate": hourly_rate,
                "notes": str(r["notes"]) if r["notes"] else "",
                "value": calculated_value,
                "aiCorrectionApplied": False,
                "manuallyEdited": False,
                "originalNotes": None,
                "originalHours": None,
                "originalCustomerId": None,
                "originalTariff": None,
                "status": "uninvoiced",
                "entrySource": "imported"
            }
            entries.append(entry)
        
        if entries:
            await db.timeEntries.insert_many(entries)
        
        # Calculate summary statistics
        total_hours = sum(e["hours"] for e in entries)
        total_value = sum(e["value"] for e in entries)
        unique_employees = len(set(e["employeeName"] for e in entries))
        unique_customers = len(set(e["customerId"] for e in entries if e["customerId"]))
        
        return {
            "batchId": batch_id,
            "rowCount": len(entries),
            "newCustomers": new_customers_created,
            "summary": {
                "totalRows": len(entries),
                "totalHours": round(total_hours, 2),
                "totalValue": round(total_value, 2),
                "uniqueEmployees": unique_employees,
                "uniqueCustomers": unique_customers
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/imports/from-verification")
async def import_from_verification(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Import filtered rows from Import Verification page.
    Accepts JSON data with batch metadata and filtered rows.
    """
    try:
        # Extract data from request
        title = request.get('title')
        invoiceDate = request.get('invoiceDate')
        periodFrom = request.get('periodFrom')
        periodTo = request.get('periodTo')
        dueDate = request.get('dueDate')
        rows = request.get('rows', [])
        filename = request.get('filename', 'import.xlsx')
        
        if not all([title, invoiceDate, periodFrom, periodTo, dueDate]):
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        # Create batch
        batch_id = str(uuid.uuid4())
        batch_doc = {
            "id": batch_id,
            "title": title,
            "filename": filename,
            "periodFrom": periodFrom,
            "periodTo": periodTo,
            "invoiceDate": invoiceDate,
            "dueDate": dueDate,
            "status": "imported",
            "createdBy": current_user.email,
            "createdAt": datetime.now(timezone.utc).isoformat()
        }
        await db.importBatches.insert_one(batch_doc)
        
        # Create time entries from provided rows
        entries = []
        
        for row in rows:
            # Extract row data
            customer_name = row.get('customer', 'General')
            project_name = row.get('project', 'General')
            employee = row.get('employee', 'Unknown')
            date_str = row.get('date', '')
            tariff = row.get('tariff', 'N/A')
            notes = row.get('comments', '')
            hours = float(row.get('hours', 0))
            value = float(row.get('value', 0))
            
            # Find or create customer
            customer = await db.customers.find_one({"name": customer_name})
            if not customer:
                customer_id = str(uuid.uuid4())
                await db.customers.insert_one({"id": customer_id, "name": customer_name})
            else:
                customer_id = customer["id"]
            
            # Find or create project
            project = await db.projects.find_one({"name": project_name, "customerId": customer_id})
            if not project:
                project_id = str(uuid.uuid4())
                await db.projects.insert_one({"id": project_id, "name": project_name, "customerId": customer_id})
            else:
                project_id = project["id"]
            
            # Create time entry
            entry = {
                "id": str(uuid.uuid4()),
                "batchId": batch_id,
                "projectId": project_id,
                "customerId": customer_id,
                "employeeName": employee,
                "date": date_str,
                "hours": hours,
                "tariff": tariff,
                "notes": notes,
                "value": value,
                "aiCorrectionApplied": row.get('aiCorrectionApplied', False),
                "manuallyEdited": row.get('manuallyEdited', False),
                "originalNotes": row.get('originalNotes'),
                "originalHours": row.get('originalHours'),
                "originalCustomerId": row.get('originalCustomerId')
            }
            entries.append(entry)
        
        # Insert all entries
        if entries:
            await db.timeEntries.insert_many(entries)
        
        return {"batchId": batch_id, "rowCount": len(entries)}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============ INVOICE ENDPOINTS ============
@api_router.get("/batches")
async def list_batches(current_user: User = Depends(get_current_user)):
    """Get all import batches with invoice counts and totals — O(3) queries regardless of batch count"""
    batches = await db.importBatches.find({}, {"_id": 0}).to_list(1000)

    if batches:
        batch_ids = [b.get("id") for b in batches if b.get("id")]

        # Single aggregation: invoice count per batch
        invoice_counts_agg = await db.invoices.aggregate([
            {"$match": {"batchId": {"$in": batch_ids}}},
            {"$group": {"_id": "$batchId", "count": {"$sum": 1}}}
        ]).to_list(None)
        invoice_count_map = {item["_id"]: item["count"] for item in invoice_counts_agg}

        # Single aggregation: total value per batch from time entries
        totals_agg = await db.timeEntries.aggregate([
            {"$match": {"batchId": {"$in": batch_ids}}},
            {"$group": {"_id": "$batchId", "total": {"$sum": "$value"}}}
        ]).to_list(None)
        total_amount_map = {item["_id"]: round(item["total"] or 0, 2) for item in totals_agg}

        for batch in batches:
            bid = batch.get("id")
            batch["invoiceCount"] = invoice_count_map.get(bid, 0)
            batch["totalAmount"] = total_amount_map.get(bid, 0.0)

    return batches

@api_router.get("/batches/{batch_id}")
async def get_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """Get batch details"""
    batch = await db.importBatches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    return batch

@api_router.get("/batches/{batch_id}/invoices")
async def get_batch_invoices(batch_id: str, current_user: User = Depends(get_current_user)):
    """Get all invoices for a specific batch"""
    invoices = await db.invoices.find({"batchId": batch_id}, {"_id": 0}).to_list(1000)
    return invoices

@api_router.get("/batches/{batch_id}/time-entries")
async def get_batch_time_entries(batch_id: str, current_user: User = Depends(get_current_user)):
    """Get all time entries for a batch (for resuming in-progress batches)"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get all time entries for this batch
    entries = await db.timeEntries.find({"batchId": batch_id}, {"_id": 0}).to_list(10000)

    if entries:
        # --- Batch all lookups: O(5) queries regardless of entry count ---

        # Collect unique IDs
        customer_ids = list({e["customerId"] for e in entries if e.get("customerId")} |
                            {e["originalCustomerId"] for e in entries if e.get("originalCustomerId")})
        project_ids  = list({e["projectId"] for e in entries if e.get("projectId")})
        invoiced_ids = [e["id"] for e in entries if e.get("status") == "invoiced" and e.get("id")]

        # Fetch customers, projects in parallel
        customers_cursor = db.customers.find({"id": {"$in": customer_ids}}, {"_id": 0, "id": 1, "name": 1})
        projects_cursor  = db.projects.find({"id": {"$in": project_ids}},   {"_id": 0, "id": 1, "name": 1})

        customer_map = {c["id"]: c.get("name", "") for c in await customers_cursor.to_list(None)}
        project_map  = {p["id"]: p.get("name", "") for p in await projects_cursor.to_list(None)}

        # Fetch invoice lines for invoiced entries (1 query)
        invoice_lines = await db.invoiceLines.find(
            {"timeEntryId": {"$in": invoiced_ids}},
            {"_id": 0, "timeEntryId": 1, "invoiceId": 1}
        ).to_list(None)
        entry_to_invoice_id = {il["timeEntryId"]: il["invoiceId"] for il in invoice_lines}

        # Fetch invoices (1 query)
        invoice_ids = list(set(entry_to_invoice_id.values()))
        invoices = await db.invoices.find(
            {"id": {"$in": invoice_ids}},
            {"_id": 0, "id": 1, "status": 1, "invoiceNumber": 1}
        ).to_list(None)
        invoice_map = {inv["id"]: inv for inv in invoices}

        # Enrich entries using maps — zero additional queries
        for entry in entries:
            entry["customerName"]         = customer_map.get(entry.get("customerId"), "")
            entry["originalCustomerName"] = customer_map.get(entry.get("originalCustomerId"), "") if entry.get("originalCustomerId") else ""
            entry["projectName"]          = project_map.get(entry.get("projectId"), "")

            if entry.get("status") == "invoiced":
                inv_id = entry_to_invoice_id.get(entry.get("id"))
                if inv_id:
                    inv = invoice_map.get(inv_id)
                    if inv:
                        entry["invoiceStatus"] = inv.get("status", "")
                        entry["invoiceNumber"] = inv.get("invoiceNumber", "")
                        entry["invoiceId"]     = inv.get("id", "")

    return entries

@api_router.put("/batches/{batch_id}/time-entries")
async def update_batch_time_entries(batch_id: str, updates: List[dict], current_user: User = Depends(get_current_user)):
    """Update time entries with corrected data from verification page"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get all current time entries for mapping
    all_entries = await db.timeEntries.find({"batchId": batch_id}, {"_id": 0}).to_list(10000)
    
    updated_count = 0
    for update_data in updates:
        # Find matching entry by index or unique characteristics
        index = update_data.get('index', 0)
        if index < len(all_entries):
            entry = all_entries[index]
            
            # Update fields if provided
            update_fields = {}
            
            # If this is the first AI correction, save original values
            if 'aiCorrectionApplied' in update_data and update_data['aiCorrectionApplied'] and not entry.get('aiCorrectionApplied'):
                # Save original values before overwriting
                if 'originalNotes' not in entry or entry.get('originalNotes') is None:
                    update_fields['originalNotes'] = entry.get('notes', '')
                if 'originalHours' not in entry or entry.get('originalHours') is None:
                    update_fields['originalHours'] = entry.get('hours', 0)
                if 'originalCustomerId' not in entry or entry.get('originalCustomerId') is None:
                    update_fields['originalCustomerId'] = entry.get('customerId', '')
            
            if 'comments' in update_data:
                # Save original notes if this is the first edit
                if 'originalNotes' not in entry or entry.get('originalNotes') is None:
                    update_fields['originalNotes'] = entry.get('notes', '')
                update_fields['notes'] = update_data['comments']
            if 'hours' in update_data:
                # Save original hours if this is the first edit
                if 'originalHours' not in entry or entry.get('originalHours') is None:
                    update_fields['originalHours'] = entry.get('hours', 0)
                update_fields['hours'] = float(update_data['hours'])
                
                # Recalculate value when hours change: value = hours × hourlyRate
                new_hours = float(update_data['hours'])
                current_hourly_rate = entry.get('hourlyRate', 0)
                update_fields['value'] = round(new_hours * current_hourly_rate, 2)
                
            if 'customerId' in update_data:
                # If this is the first customer change, save original
                if 'originalCustomerId' not in entry or entry.get('originalCustomerId') is None:
                    update_fields['originalCustomerId'] = entry.get('customerId', '')
                update_fields['customerId'] = update_data['customerId']
                
            if 'tariff' in update_data:
                # If this is the first tariff change, save original
                if 'originalTariff' not in entry or entry.get('originalTariff') is None:
                    update_fields['originalTariff'] = entry.get('tariff', '')
                update_fields['tariff'] = update_data['tariff']
                
                # When tariff changes, update hourlyRate from tariff codes
                tariff_code = update_data['tariff']
                tariff_doc = await db.tariffs.find_one({"code": tariff_code})
                if tariff_doc:
                    new_hourly_rate = tariff_doc.get('value', 0)
                    update_fields['hourlyRate'] = new_hourly_rate
                    
                    # Recalculate value: value = hours × new hourlyRate
                    current_hours = entry.get('hours', 0)
                    update_fields['value'] = round(current_hours * new_hourly_rate, 2)
                    
            if 'hourlyRate' in update_data:
                # Allow manual hourlyRate updates
                new_hourly_rate = float(update_data['hourlyRate'])
                update_fields['hourlyRate'] = new_hourly_rate
                
                # Recalculate value when hourlyRate changes manually: value = hours × hourlyRate
                current_hours = entry.get('hours', 0)
                update_fields['value'] = round(current_hours * new_hourly_rate, 2)
                
            if 'aiCorrectionApplied' in update_data:
                update_fields['aiCorrectionApplied'] = bool(update_data['aiCorrectionApplied'])
            if 'manuallyEdited' in update_data:
                update_fields['manuallyEdited'] = bool(update_data['manuallyEdited'])
            if 'status' in update_data:
                # Allow changing status: uninvoiced, invoiced, internal, free, forfait, ready
                allowed_statuses = ['uninvoiced', 'invoiced', 'internal', 'free', 'forfait', 'ready']
                if update_data['status'] in allowed_statuses:
                    update_fields['status'] = update_data['status']
            
            if update_fields:
                await db.timeEntries.update_one(
                    {"id": entry["id"]},
                    {"$set": update_fields}
                )
                updated_count += 1
    
    return {"message": f"Updated {updated_count} time entries", "updated_count": updated_count}

@api_router.post("/batches/{batch_id}/manual-entry")
async def add_manual_entry(
    batch_id: str,
    entry_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Add a manual time entry or forfait batch entry to a batch"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    entry_source = entry_data.get("entrySource", "manual")
    
    # Get customer data
    customer_name = ""
    customer = None
    if entry_data.get("customerId"):
        customer = await db.customers.find_one({"id": entry_data["customerId"]})
        if customer:
            customer_name = customer.get("name", "")
    
    # Calculate value based on entry source
    hours = float(entry_data.get("hours", 0))
    
    if entry_source == "forfait_batch" and customer:
        # For forfait batch entries, use customer's fixedForfaitValue
        hourly_rate = customer.get("fixedForfaitValue", 0)
        calculated_value = customer.get("fixedForfaitValue", 0)  # Use fixed forfait value directly
    else:
        # For manual entries, calculate from tariff
        hourly_rate = 0
        if entry_data.get("tariff"):
            tariff = await db.tariffs.find_one({"code": entry_data["tariff"]})
            if tariff:
                hourly_rate = tariff.get("value", 0)
        calculated_value = round(hours * hourly_rate, 2)
    
    # Create manual entry
    entry_source = entry_data.get("entrySource", "manual")  # Support "manual" or "forfait_batch"
    
    # For forfait batch entries, use "Forfait Batch" as project name
    project_name = "Forfait Batch" if entry_source == "forfait_batch" else "Manual Entry"
    
    # Find or create project
    project = await db.projects.find_one({"name": project_name, "customerId": entry_data.get("customerId")})
    if not project:
        project_id = str(uuid.uuid4())
        await db.projects.insert_one({
            "id": project_id,
            "name": project_name,
            "customerId": entry_data.get("customerId")
        })
    else:
        project_id = project["id"]
    
    # Create entry with forfait correlation support
    manual_entry = {
        "id": str(uuid.uuid4()),
        "batchId": batch_id,
        "projectId": project_id,
        "projectName": project_name,
        "customerId": entry_data.get("customerId"),
        "customerName": customer_name,
        "employeeName": entry_data.get("employeeName", ""),
        "date": entry_data.get("date"),
        "hours": hours,
        "tariff": entry_data.get("tariff", ""),
        "hourlyRate": hourly_rate,
        "notes": entry_data.get("notes", ""),
        "value": calculated_value,
        "aiCorrectionApplied": False,
        "manuallyEdited": False,
        "originalNotes": None,
        "originalHours": None,
        "originalCustomerId": None,
        "originalTariff": None,
        "status": entry_data.get("status", "uninvoiced"),
        "entrySource": entry_source,  # "manual" or "forfait_batch"
        "forfaitBatchParentId": entry_data.get("forfaitBatchParentId"),  # For future: link to forfait parent
        "forfaitBatchSubRows": entry_data.get("forfaitBatchSubRows", [])  # For future: list of sub-row IDs
    }
    
    await db.timeEntries.insert_one(manual_entry)
    
    return {"message": f"{project_name} added successfully", "entryId": manual_entry["id"]}

@api_router.get("/batches/{batch_id}/verification")
async def get_batch_verification(batch_id: str, current_user: User = Depends(get_current_user)):
    """Get verification data for specific clients and no-client entries - only for composed batches"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Only return verification data if batch has been composed (invoices created)
    # For "in progress" batches, return empty arrays
    if batch.get("status") == "in progress":
        return {
            "jmmcHP": [],
            "jmmcFinance": [],
            "noClient": [],
            "extra": []
        }
    
    # Get all time entries for this batch
    all_entries = await db.timeEntries.find({"batchId": batch_id}, {"_id": 0}).to_list(10000)
    
    # Get customer names and project names for each entry
    jmmc_hp_entries = []
    jmmc_finance_entries = []
    no_client_entries = []
    extra_entries = []
    
    for entry in all_entries:
        # Get customer name
        customer = await db.customers.find_one({"id": entry["customerId"]})
        customer_name = customer["name"] if customer else ""
        
        # Get tariff value
        tariff_value = entry.get("tariff", "")
        
        # Categorize entries
        if "JMMC HP d.o.o." in customer_name:
            jmmc_hp_entries.append(entry)
        elif "JMMC Finance d.o.o." in customer_name:
            jmmc_finance_entries.append(entry)
        elif (not customer_name or customer_name.strip() == "" or customer_name == "General") and tariff_value == "999 - EXTRA":
            # EXTRA category: no client and tariff is "999 - EXTRA"
            extra_entries.append(entry)
        elif not customer_name or customer_name.strip() == "" or customer_name == "General":
            # No client category (excluding EXTRA entries)
            no_client_entries.append(entry)
    
    return {
        "jmmcHP": jmmc_hp_entries,
        "jmmcFinance": jmmc_finance_entries,
        "noClient": no_client_entries,
        "extra": extra_entries

    }

@api_router.post("/batches/{batch_id}/verify-entries")
async def verify_batch_entries(batch_id: str, current_user: User = Depends(get_current_user)):
    """Run AI verification on time entry descriptions using batch processing"""
    import json
    import asyncio
    
    # Get user's AI settings
    user_settings = await db.aiSettings.find_one({"userId": current_user.email})
    
    if not user_settings:
        user_settings = AISettings().model_dump()
    
    # Determine API key and model
    ai_provider = user_settings.get("aiProvider", "custom")
    if user_settings.get("customApiKey") and ai_provider in ("custom", "openai", "anthropic"):
        api_key = user_settings["customApiKey"]
        model = user_settings.get("customModel", "claude-3-5-sonnet-20241022")
    elif EMERGENT_LLM_KEY:
        api_key = EMERGENT_LLM_KEY
        model = "claude-3-5-sonnet-20241022"
    else:
        return {"results": {}, "message": "AI not configured. Please add an API key in Settings."}

    # Get entries specifically from verification categories (JMMC HP, JMMC Finance, No Client)
    # This ensures we verify the entries the user actually sees in the UI
    all_entries = await db.timeEntries.find({"batchId": batch_id}, {"_id": 0}).to_list(10000)
    
    if not all_entries:
        return {"results": {}, "message": "No entries found"}
    
    # Categorize to match what user sees in verification tile
    verification_entries = []
    for entry in all_entries:
        customer = await db.customers.find_one({"id": entry["customerId"]})
        customer_name = customer["name"] if customer else ""
        
        # Only include entries from the three verification categories
        if ("JMMC HP d.o.o." in customer_name or 
            "JMMC Finance d.o.o." in customer_name or 
            not customer_name or customer_name.strip() == "" or customer_name == "General"):
            verification_entries.append(entry)
    
    #  Limit to prevent long processing
    MAX_ENTRIES = 50
    if len(verification_entries) > MAX_ENTRIES:
        entries_to_check = verification_entries[:MAX_ENTRIES]
        logger.info(f"Limited to {MAX_ENTRIES} of {len(verification_entries)} verification entries")
    else:
        entries_to_check = verification_entries
        logger.info(f"Verifying all {len(verification_entries)} verification entries")
    
    if not entries_to_check:
        return {"results": {}, "message": "No entries with descriptions to check"}
    
    verification_prompt = user_settings.get("verificationPrompt", 
        "Analyze this work description for suspicious patterns. If no description is provided or it's empty/vague, flag it as suspicious. If suspicious, respond with JSON: {\"flagged\": true, \"reason\": \"brief explanation\"}. If normal: {\"flagged\": false, \"reason\": \"\"}")
    
    try:
        # Use batch processing - combine multiple entries into one prompt
        batch_size = 10  # Process 10 entries per API call
        results = {}
        total_batches = (len(entries_to_check) + batch_size - 1) // batch_size
        
        logger.info(f"Starting verification of {len(entries_to_check)} entries in {total_batches} batches")
        
        for i in range(0, len(entries_to_check), batch_size):
            batch = entries_to_check[i:i + batch_size]
            
            # Create a batch prompt
            batch_text = "Analyze the following work entries and return a JSON array with results for each entry. Format: [{\"entry_id\": \"id1\", \"flagged\": true/false, \"reason\": \"explanation\"}]\n\n"
            batch_text += "Entries to analyze:\n"
            
            for idx, entry in enumerate(batch):
                description = entry.get('notes', '') or '(No description provided)'
                hours_value = entry.get('hours', 0)
                # Round hours to 2 decimals for cleaner AI output
                hours_rounded = round(float(hours_value), 2) if hours_value else 0
                batch_text += f"\n{idx + 1}. Entry ID: {entry.get('id')}\n"
                batch_text += f"   Description: {description}\n"
                batch_text += f"   Employee: {entry.get('employeeName', 'N/A')}\n"
                batch_text += f"   Hours: {hours_rounded}\n"
                batch_text += f"   Date: {entry.get('date', 'N/A')}\n"
            
            batch_text += f"\n{verification_prompt}"
            
            # Make single API call for batch with timeout
            try:
                # Determine provider based on model
                if "claude" in model.lower():
                    provider = "anthropic"
                elif "gemini" in model.lower():
                    provider = "google"
                else:
                    provider = "openai"
                
                chat = LlmChat(
                    api_key=api_key,
                    session_id=f"verification-{current_user.email}-{i}",
                    system_message="You are an AI assistant for invoice verification. Always respond with valid JSON array."
                ).with_model(provider, model)
                
                message = UserMessage(text=batch_text)
                
                # Add timeout to prevent hanging
                response = await asyncio.wait_for(
                    chat.send_message(message),
                    timeout=30.0  # 30 second timeout per batch
                )
                
                # Parse batch response
                try:
                    clean_response = response.strip()
                    if clean_response.startswith("```"):
                        clean_response = clean_response.split("```")[1]
                        if clean_response.startswith("json"):
                            clean_response = clean_response[4:]
                    clean_response = clean_response.strip()
                    
                    batch_results = json.loads(clean_response)
                    
                    # Process batch results
                    for result in batch_results:
                        if isinstance(result, dict) and result.get("flagged", False):
                            entry_id = result.get("entry_id")
                            if entry_id:
                                results[entry_id] = {
                                    "flagged": True,
                                    "reason": result.get("reason", "Suspicious activity detected")
                                }
                except (json.JSONDecodeError, TypeError) as e:
                    # If batch fails, continue to next batch
                    logger.warning(f"Failed to parse batch {i}: {str(e)}")
                    continue
                    
            except asyncio.TimeoutError:
                logger.warning(f"Batch {i} timed out after 30 seconds")
                continue
        
        logger.info(f"Verification complete: {len(results)} flagged out of {len(entries_to_check)}")
        return {"results": results, "total_checked": len(entries_to_check), "flagged_count": len(results)}
        
    except Exception as e:
        logger.error(f"AI verification error: {str(e)}")
        return {"results": {}, "message": f"AI verification error: {str(e)}"}


@api_router.post("/batches/{batch_id}/run-ai-prompts")
async def run_ai_prompts_on_entries(
    batch_id: str,
    entry_ids: List[str] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """
    Run all AI prompts (Grammar, Fraud, GDPR, Verification) consecutively on selected time entries.
    Returns suggestions for each entry that the user can review and apply.
    """
    import asyncio
    
    # Get user's AI settings
    user_settings = await db.aiSettings.find_one({"userId": current_user.email})
    
    if not user_settings:
        user_settings = AISettings().model_dump()
    
    # Determine API key (not model - each prompt has its own model)
    ai_provider = user_settings.get("aiProvider", "custom")
    if user_settings.get("customApiKey") and ai_provider in ("custom", "openai", "anthropic"):
        api_key = user_settings["customApiKey"]
    elif EMERGENT_LLM_KEY:
        api_key = EMERGENT_LLM_KEY
    else:
        raise HTTPException(status_code=400, detail="AI not configured. Please add an API key in Settings.")
    
    # Get the prompts and their specific models from settings
    grammar_prompt = user_settings.get("grammarPrompt", "")
    grammar_model = user_settings.get("grammarModel", "gpt-5-nano")
    
    fraud_prompt = user_settings.get("fraudPrompt", "")
    fraud_model = user_settings.get("fraudModel", "gpt-5-mini")
    
    gdpr_prompt = user_settings.get("gdprPrompt", "")
    gdpr_model = user_settings.get("gdprModel", "gpt-5-mini")
    
    verification_prompt = user_settings.get("verificationPrompt", "")
    verification_model = user_settings.get("verificationModel", "gpt-5")
    
    # Fetch the time entries
    entries = await db.timeEntries.find(
        {"batchId": batch_id, "id": {"$in": entry_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    if not entries:
        raise HTTPException(status_code=404, detail="No entries found")
    
    # Helper function to determine provider from model name
    def get_provider(model_name: str) -> str:
        if "claude" in model_name.lower():
            return "anthropic"
        elif "gemini" in model_name.lower():
            return "google"
        else:
            return "openai"
    
    results = []
    
    # Process each entry
    for entry in entries:
        entry_id = entry.get("id")
        description = entry.get("notes", "")
        hours = entry.get("hours", 0)
        employee = entry.get("employeeName", "")
        date = entry.get("date", "")
        
        # Prepare context for AI
        entry_context = f"""
Time Entry Details:
- Description: {description}
- Employee: {employee}
- Hours: {hours}
- Date: {date}
"""
        
        entry_result = {
            "entryId": entry_id,
            "originalDescription": description,
            "suggestions": {}
        }
        
        try:
            # Run all prompts consecutively with their specific models
            
            # 1. Grammar Correction
            if grammar_prompt:
                try:
                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"grammar-{current_user.email}-{entry_id}",
                        system_message="You are an expert grammar and spelling corrector."
                    ).with_model(get_provider(grammar_model), grammar_model)
                    
                    prompt_text = f"{grammar_prompt}\n\n{entry_context}"
                    message = UserMessage(text=prompt_text)
                    response = await asyncio.wait_for(chat.send_message(message), timeout=20.0)
                    
                    entry_result["suggestions"]["grammar"] = {
                        "type": "grammar",
                        "suggestion": response.strip(),
                        "applied": False
                    }
                except asyncio.TimeoutError:
                    entry_result["suggestions"]["grammar"] = {
                        "type": "grammar",
                        "error": "Request timed out"
                    }
                except Exception as e:
                    entry_result["suggestions"]["grammar"] = {
                        "type": "grammar",
                        "error": str(e)
                    }
            
            # 2. Fraud Detection
            if fraud_prompt:
                try:
                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"fraud-{current_user.email}-{entry_id}",
                        system_message="You are a fraud detection expert for time entries and invoicing."
                    ).with_model(get_provider(fraud_model), fraud_model)
                    
                    prompt_text = f"{fraud_prompt}\n\n{entry_context}"
                    message = UserMessage(text=prompt_text)
                    response = await asyncio.wait_for(chat.send_message(message), timeout=20.0)
                    
                    entry_result["suggestions"]["fraud"] = {
                        "type": "fraud",
                        "suggestion": response.strip(),
                        "applied": False
                    }
                except asyncio.TimeoutError:
                    entry_result["suggestions"]["fraud"] = {
                        "type": "fraud",
                        "error": "Request timed out"
                    }
                except Exception as e:
                    entry_result["suggestions"]["fraud"] = {
                        "type": "fraud",
                        "error": str(e)
                    }
            
            # 3. GDPR Data Masking
            if gdpr_prompt:
                try:
                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"gdpr-{current_user.email}-{entry_id}",
                        system_message="You are a GDPR compliance expert focused on data privacy."
                    ).with_model(get_provider(gdpr_model), gdpr_model)
                    
                    prompt_text = f"{gdpr_prompt}\n\n{entry_context}"
                    message = UserMessage(text=prompt_text)
                    response = await asyncio.wait_for(chat.send_message(message), timeout=20.0)
                    
                    entry_result["suggestions"]["gdpr"] = {
                        "type": "gdpr",
                        "suggestion": response.strip(),
                        "applied": False
                    }
                except asyncio.TimeoutError:
                    entry_result["suggestions"]["gdpr"] = {
                        "type": "gdpr",
                        "error": "Request timed out"
                    }
                except Exception as e:
                    entry_result["suggestions"]["gdpr"] = {
                        "type": "gdpr",
                        "error": str(e)
                    }
            
            # 4. Invoice Verification (General)
            if verification_prompt:
                try:
                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"verification-{current_user.email}-{entry_id}",
                        system_message="You are a general verification expert for time entry data quality."
                    ).with_model(get_provider(verification_model), verification_model)
                    
                    prompt_text = f"{verification_prompt}\n\n{entry_context}"
                    message = UserMessage(text=prompt_text)
                    response = await asyncio.wait_for(chat.send_message(message), timeout=20.0)
                    
                    entry_result["suggestions"]["verification"] = {
                        "type": "verification",
                        "suggestion": response.strip(),
                        "applied": False
                    }
                except asyncio.TimeoutError:
                    entry_result["suggestions"]["verification"] = {
                        "type": "verification",
                        "error": "Request timed out"
                    }
                except Exception as e:
                    entry_result["suggestions"]["verification"] = {
                        "type": "verification",
                        "error": str(e)
                    }
            
            results.append(entry_result)
            
        except Exception as e:
            logger.error(f"Error processing entry {entry_id}: {str(e)}")
            entry_result["error"] = str(e)
            results.append(entry_result)
    
    return {
        "success": True,
        "results": results,
        "total_entries": len(entries),
        "message": f"AI prompts executed on {len(entries)} entries"
    }

@api_router.post("/imports/verify-preview")
async def verify_import_preview(rows: List[dict], current_user: User = Depends(get_current_user)):
    """Run AI verification on import preview rows before creating batch"""
    import json
    
    # Get user's AI settings
    user_settings = await db.aiSettings.find_one({"userId": current_user.email})
    
    if not user_settings:
        user_settings = AISettings().model_dump()
    
    # Determine API key and model
    ai_provider = user_settings.get("aiProvider", "custom")
    if user_settings.get("customApiKey") and ai_provider in ("custom", "openai", "anthropic"):
        api_key = user_settings["customApiKey"]
        model = user_settings.get("customModel", "claude-3-5-sonnet-20241022")
    elif EMERGENT_LLM_KEY:
        api_key = EMERGENT_LLM_KEY
        model = "claude-3-5-sonnet-20241022"
    else:
        return {"results": {}, "message": "AI not configured. Please add an API key in Settings."}

    # Process all rows (frontend handles chunking for progress)
    entries_to_check = rows
    
    logger.info(f"Verifying {len(entries_to_check)} import rows")
    
    # Get all prompts from settings
    grammar_prompt = user_settings.get("grammarPrompt", "")
    fraud_prompt = user_settings.get("fraudPrompt", "")
    gdpr_prompt = user_settings.get("gdprPrompt", "")
    verification_prompt = user_settings.get("verificationPrompt", "")
    
    # Combine all prompts into verification criteria
    all_criteria = []
    if grammar_prompt:
        all_criteria.append(f"GRAMMAR CHECK: {grammar_prompt}")
    if fraud_prompt:
        all_criteria.append(f"FRAUD DETECTION: {fraud_prompt}")
    if gdpr_prompt:
        all_criteria.append(f"GDPR COMPLIANCE: {gdpr_prompt}")
    if verification_prompt:
        all_criteria.append(f"GENERAL VERIFICATION: {verification_prompt}")
    
    combined_criteria = "\n\n".join(all_criteria)
    
    logger.info(f"Using {len(all_criteria)} verification prompts")
    
    try:
        # Process in batches
        batch_size = 10
        results = {}
        
        for i in range(0, len(entries_to_check), batch_size):
            batch = entries_to_check[i:i + batch_size]
            
            # Create batch prompt with ALL verification criteria
            batch_text = f"""Analyze work entries. Check ALL criteria. Flag if ANY violated.

RESPONSE FORMAT (MANDATORY - follow exactly):
[
  {{
    "entry_index": 0,
    "flagged": true,
    "reason": "Grammar: missing capitalization",
    "suggestions": {{
      "description": "Priprava najemnin za obdobje",
      "hours": null
    }}
  }}
]

VERIFICATION CRITERIA (Flag if violated):

{combined_criteria}

RULES:
1. Grammar/capitalization wrong → flagged=true, reason="Grammar issue", suggestions.description=corrected_text
2. Vague description → flagged=true, reason="Vague description"  
3. Wrong customer in text → flagged=true, reason="Wrong customer"
4. ALWAYS use exact field names: entry_index, flagged, reason, suggestions
5. ALWAYS return array [...] even for 1 entry

Entries:
"""
            
            for idx, row in enumerate(batch):
                global_idx = i + idx
                description = row.get('comments', '') or '(No description)'
                hours = row.get('hours', 0)
                employee = row.get('employee', '')
                customer = row.get('customer', '')
                
                batch_text += f"\nEntry {global_idx}:\n"
                batch_text += f"  Employee: {employee}\n"
                batch_text += f"  Customer: {customer}\n"
                batch_text += f"  Description: {description}\n"
                batch_text += f"  Hours: {hours}\n"
            
            # Call AI
            logger.info(f"Calling AI with model: {model}, batch size: {len(batch)}")
            
            chat = LlmChat(
                api_key=api_key,
                session_id=f"import-verify-{current_user.email}-{i}",
                system_message="You are an AI assistant that analyzes work time entries for anomalies and provides corrections. Always respond with valid JSON array format."
            )
            
            # Set model if using custom provider
            if user_settings.get("aiProvider") == "custom":
                chat = chat.with_model("openai", model)
            else:
                chat = chat.with_model("openai", "gpt-5")
            
            ai_response = await chat.send_message(UserMessage(text=batch_text))
            response_text = ai_response.strip()
            
            logger.info(f"Batch {i//batch_size + 1} AI response (first 500 chars): {response_text[:500]}")
            
            # Parse JSON response
            try:
                # Extract JSON from markdown code blocks if present
                if '```json' in response_text:
                    response_text = response_text.split('```json')[1].split('```')[0].strip()
                elif '```' in response_text:
                    response_text = response_text.split('```')[1].split('```')[0].strip()
                
                logger.info(f"Parsing JSON response: {response_text[:300]}")
                
                # Try to parse as JSON
                parsed = json.loads(response_text)
                
                # Handle both array and single object responses
                if isinstance(parsed, dict):
                    # Single object response - convert to array
                    batch_results = [parsed]
                    logger.info(f"Converted single object to array")
                elif isinstance(parsed, list):
                    batch_results = parsed
                else:
                    logger.error(f"Unexpected JSON type: {type(parsed)}")
                    continue
                
                logger.info(f"Successfully parsed {len(batch_results)} results from batch")
                
                # Store results by entry index
                for result in batch_results:
                    entry_idx = result.get('entry_index', 0)
                    if result.get('flagged'):
                        results[str(entry_idx)] = {
                            "reason": result.get('reason', 'Flagged by AI'),
                            "suggestions": result.get('suggestions', {})
                        }
                        logger.info(f"Entry {entry_idx} flagged: {result.get('reason', 'No reason')}")
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse AI response as JSON: {e}")
                logger.error(f"Raw response: {response_text[:500]}")
                continue
        
        logger.info(f"Total flagged entries: {len(results)} out of {len(entries_to_check)}")
        
        return {
            "results": results,
            "total_checked": len(entries_to_check),
            "total_flagged": len(results)
        }
        
    except Exception as e:
        logger.error(f"Import preview AI verification error: {str(e)}")
        return {"results": {}, "message": f"AI verification error: {str(e)}"}

@api_router.put("/batches/{batch_id}")
async def update_batch(batch_id: str, update_data: dict, current_user: User = Depends(get_current_user)):
    """Update batch details"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Update allowed fields
    allowed_fields = ["title", "invoiceDate", "periodFrom", "periodTo", "dueDate", "status", "rowsPerPage", "filterPreferences"]
    update_fields = {k: v for k, v in update_data.items() if k in allowed_fields}
    
    if update_fields:
        await db.importBatches.update_one(
            {"id": batch_id},
            {"$set": update_fields}
        )
        
        # Update all invoices in this batch with new dates if changed
        if any(k in update_fields for k in ["invoiceDate", "periodFrom", "periodTo", "dueDate"]):
            invoice_updates = {}
            if "invoiceDate" in update_fields:
                invoice_updates["invoiceDate"] = update_fields["invoiceDate"]
            if "periodFrom" in update_fields:
                invoice_updates["periodFrom"] = update_fields["periodFrom"]
            if "periodTo" in update_fields:
                invoice_updates["periodTo"] = update_fields["periodTo"]
            if "dueDate" in update_fields:
                invoice_updates["dueDate"] = update_fields["dueDate"]
            
            if invoice_updates:
                await db.invoices.update_many(
                    {"batchId": batch_id},
                    {"$set": invoice_updates}
                )
    
    return {"message": "Batch updated successfully"}

@api_router.post("/batches/{batch_id}/archive")
async def archive_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """Archive a batch"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    await db.importBatches.update_one(
        {"id": batch_id},
        {"$set": {"status": "archived"}}
    )
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "archive_batch",
        "entity": "Batch",
        "entityId": batch_id,
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Batch archived successfully"}


@api_router.post("/batches/{batch_id}/unarchive")
async def unarchive_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """Unarchive a batch - restore to 'composed' or 'imported' status based on invoice count"""
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    if batch.get("status") != "archived":
        raise HTTPException(status_code=400, detail="Batch is not archived")
    
    # Determine new status based on whether batch has invoices
    invoice_count = await db.invoices.count_documents({"batchId": batch_id})
    new_status = "composed" if invoice_count > 0 else "imported"
    
    await db.importBatches.update_one(
        {"id": batch_id},
        {"$set": {"status": new_status}}
    )
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "unarchive_batch",
        "entity": "Batch",
        "entityId": batch_id,
        "metadata": {"newStatus": new_status},
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Batch unarchived successfully", "newStatus": new_status}


@api_router.delete("/batches/{batch_id}")
async def delete_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """
    Delete a batch and all its time entries.
    Only allowed if the batch has 0 invoices prepared.
    """
    # Check if batch exists
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Check if batch has any invoices (count fresh from database)
    invoice_count = await db.invoices.count_documents({"batchId": batch_id})
    
    if invoice_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete batch with {invoice_count} invoice(s). Please delete all invoices first or use Archive instead."
        )
    
    # Count time entries before deletion (for confirmation message)
    time_entry_count = await db.timeEntries.count_documents({"batchId": batch_id})
    
    # Delete all time entries for this batch (Import Verification rows - but only if no invoices)
    delete_entries_result = await db.timeEntries.delete_many({"batchId": batch_id})
    
    # Delete the batch itself
    delete_batch_result = await db.importBatches.delete_one({"id": batch_id})
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "delete_batch",
        "entity": "Batch",
        "entityId": batch_id,
        "metadata": {
            "batchTitle": batch.get("title", "Unknown"),
            "timeEntriesDeleted": delete_entries_result.deleted_count,
            "status": batch.get("status", "Unknown")
        },
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": "Batch deleted successfully",
        "batchTitle": batch.get("title", "Unknown"),
        "timeEntriesDeleted": time_entry_count,
        "invoicesDeleted": 0
    }


@api_router.delete("/admin/batches/{batch_id}/force")
async def force_delete_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """
    ADMIN ONLY: Force delete a batch including all invoices and time entries.
    WARNING: This is destructive and deletes all related data.
    """
    # Check admin role
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if batch exists
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Count what will be deleted
    time_entry_count = await db.timeEntries.count_documents({"batchId": batch_id})
    invoice_count = await db.invoices.count_documents({"batchId": batch_id})
    
    # Delete all invoices for this batch
    await db.invoices.delete_many({"batchId": batch_id})
    
    # Delete all invoice lines for this batch's invoices
    await db.invoiceLines.delete_many({"batchId": batch_id})
    
    # Delete all time entries for this batch
    await db.timeEntries.delete_many({"batchId": batch_id})
    
    # Delete the batch itself
    await db.importBatches.delete_one({"id": batch_id})
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "force_delete_batch",
        "entity": "Batch",
        "entityId": batch_id,
        "metadata": {
            "batchTitle": batch.get("title", "Unknown"),
            "timeEntriesDeleted": time_entry_count,
            "invoicesDeleted": invoice_count,
            "status": batch.get("status", "Unknown"),
            "warning": "FORCE DELETE - All related data removed"
        },
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": "Batch force deleted successfully",
        "batchTitle": batch.get("title", "Unknown"),
        "timeEntriesDeleted": time_entry_count,
        "invoicesDeleted": invoice_count
    }

    invoice_count = await db.invoices.count_documents({"batchId": batch_id})
    
    if invoice_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete batch with {invoice_count} invoice(s). Please delete all invoices first or use Archive instead."
        )
    
    # Count time entries before deletion (for confirmation message)
    time_entry_count = await db.timeEntries.count_documents({"batchId": batch_id})
    
    # Delete all time entries for this batch
    delete_entries_result = await db.timeEntries.delete_many({"batchId": batch_id})
    
    # Delete the batch itself
    delete_batch_result = await db.importBatches.delete_one({"id": batch_id})
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "delete_batch",
        "entity": "Batch",
        "entityId": batch_id,
        "metadata": {
            "batchTitle": batch.get("title", "Unknown"),
            "timeEntriesDeleted": delete_entries_result.deleted_count,
            "status": batch.get("status", "Unknown")
        },
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": "Batch deleted successfully",
        "batchTitle": batch.get("title", "Unknown"),
        "timeEntriesDeleted": time_entry_count,
        "invoicesDeleted": 0
    }


# ============ COMPANIES ============
@api_router.get("/companies")
async def get_all_companies(current_user: User = Depends(get_current_user)):
    """Get all companies - ensure only JMMC HP d.o.o. and JMMC Finance d.o.o. exist"""
    
    # Define the two official companies
    official_companies = [
        {"name": "JMMC HP d.o.o."},
        {"name": "JMMC Finance d.o.o."}
    ]
    
    # Check if official companies exist
    existing_companies = await db.companies.find({}, {"_id": 0}).to_list(1000)
    
    # Get or create each official company
    final_companies = []
    for official in official_companies:
        company = await db.companies.find_one({"name": official["name"]})
        if not company:
            # Create the official company
            company_id = str(uuid.uuid4())
            company = {
                "id": company_id,
                "name": official["name"]
            }
            await db.companies.insert_one(company)
            logger.info(f"Created official company: {official['name']}")
        final_companies.append({"id": company["id"], "name": company["name"]})
    
    # Sort by name
    final_companies.sort(key=lambda x: x.get("name", "").lower())
    return final_companies

# ============ CUSTOMERS ============
@api_router.post("/customers")
async def create_customer(
    customer_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Create a new customer manually"""
    # Generate unique ID
    customer_id = str(uuid.uuid4())
    
    # Build customer object
    new_customer = {
        "id": customer_id,
        "name": customer_data.get("name", "").strip(),
        "unitPrice": float(customer_data.get("unitPrice", 0)),
        "historicalInvoices": [],
        "status": "active"  # Default status for new customers
    }
    
    # Add company if provided
    if customer_data.get("companyId"):
        new_customer["companyId"] = customer_data["companyId"]
    
    # Check if customer with same name already exists
    existing = await db.customers.find_one({"name": new_customer["name"]})
    if existing:
        raise HTTPException(status_code=400, detail="Customer with this name already exists")
    
    # Insert into database
    await db.customers.insert_one(new_customer)
    
    return {"message": "Customer created successfully", "customerId": customer_id}

@api_router.get("/customers")
async def get_all_customers(company_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get all customers from database with statistics from historical data"""
    # Build query filter
    query = {}
    if company_id:
        query["companyId"] = company_id
    
    customers = await db.customers.find(query, {"_id": 0}).to_list(10000)

    # Batch-fetch all companies needed (1 query instead of N)
    company_ids = list({c["companyId"] for c in customers if c.get("companyId")})
    if company_ids:
        companies = await db.companies.find({"id": {"$in": company_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        company_name_map = {c["id"]: c.get("name", "") for c in companies}
    else:
        company_name_map = {}

    for customer in customers:
        historical_invoices = customer.get("historicalInvoices", [])
        total_amount  = sum(inv.get("amount", 0) for inv in historical_invoices)
        invoice_count = len(historical_invoices)

        customer["invoiceCount"]   = invoice_count
        customer["totalInvoiced"]  = total_amount
        customer["averageInvoice"] = total_amount / invoice_count if invoice_count > 0 else 0
        customer["unitPrice"]      = customer.get("unitPrice", 0)
        customer["companyName"]    = company_name_map.get(customer.get("companyId"), "")

        if "status" not in customer or customer["status"] is None:
            customer["status"] = "active"

    customers.sort(key=lambda x: x.get("name", "").lower())
    return customers

@api_router.get("/customers/{customer_id}")
async def get_customer_detail(customer_id: str, current_user: User = Depends(get_current_user)):
    """Get detailed customer information with historical invoices from uploaded data"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get all historical invoices from uploaded data (sorted by date descending)
    historical_invoices = customer.get("historicalInvoices", [])
    
    # Filter to only show entries with individualRows (new format) OR manual entries
    # This removes old plain entries without expandable details
    historical_invoices = [
        inv for inv in historical_invoices 
        if inv.get("individualRows") or inv.get("source") == "manual"
    ]
    
    # Sort by date descending (show all, no limit)
    if historical_invoices:
        # Add unique IDs to each historical invoice for deletion
        for idx, inv in enumerate(historical_invoices):
            if "id" not in inv:
                inv["id"] = f"hist_{idx}_{inv.get('date', '')}_{inv.get('amount', 0)}"
        
        historical_invoices.sort(key=lambda x: x.get("date", ""), reverse=True)
        # No limit - show all periods
    
    # Calculate statistics from all historical data (including old format)
    all_historical = customer.get("historicalInvoices", [])
    total_amount = sum(inv.get("amount", 0) for inv in all_historical)
    invoice_count = len(all_historical)
    
    customer["lastInvoices"] = historical_invoices
    customer["invoiceCount"] = invoice_count
    customer["totalInvoiced"] = total_amount
    customer["averageInvoice"] = total_amount / invoice_count if invoice_count > 0 else 0
    customer["unitPrice"] = customer.get("unitPrice", 0)
    customer["historicalInvoices"] = customer.get("historicalInvoices", [])
    
    # Add company name if customer has companyId
    if customer.get("companyId"):
        company = await db.companies.find_one({"id": customer["companyId"]}, {"_id": 0})
        customer["companyName"] = company.get("name", "") if company else ""
    else:
        customer["companyName"] = ""
    
    return customer

@api_router.delete("/customers/{customer_id}/historical/{invoice_index}")
async def delete_historical_invoice(
    customer_id: str,
    invoice_index: int,
    current_user: User = Depends(get_current_user)
):
    """Delete a historical invoice entry"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    historical_invoices = customer.get("historicalInvoices", [])
    
    if 0 <= invoice_index < len(historical_invoices):
        historical_invoices.pop(invoice_index)
        
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": {"historicalInvoices": historical_invoices}}
        )
        
        return {"message": "Historical invoice deleted successfully"}
    else:
        raise HTTPException(status_code=404, detail="Historical invoice not found")

@api_router.put("/customers/{customer_id}")
async def update_customer(
    customer_id: str, 
    update_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update customer information (unit price, company)"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Only allow updating specific fields
    allowed_fields = ["unitPrice", "companyId", "invoicingType", "invoicingPeriod", "offersAddress", "invoicingStartDate", "invoicingEndDate", "addressServiceUnitPrice", "fixedForfaitValue", "status"]
    update_fields = {k: v for k, v in update_data.items() if k in allowed_fields}
    
    if update_fields:
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": update_fields}
        )
    
    return {"message": "Customer updated successfully"}

@api_router.post("/customers/{customer_id}/add-manual-entry")
async def add_manual_historical_entry(
    customer_id: str,
    entry_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Manually add a historical invoice entry"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Create manual entry
    manual_entry = {
        "date": entry_data.get("date"),
        "description": entry_data.get("description", ""),
        "amount": float(entry_data.get("amount", 0)),
        "source": "manual",
        "individualRows": []  # Manual entries don't have sub-rows
    }
    
    # Add to historical invoices
    existing_history = customer.get("historicalInvoices", [])
    existing_history.append(manual_entry)
    
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"historicalInvoices": existing_history}}
    )
    
    return {"message": "Manual entry added successfully"}

@api_router.post("/customers/refresh-invoicing-settings")
async def refresh_invoicing_settings(
    customer_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Refresh invoicing settings for customers based on their historical invoice data
    
    Analyzes Article 000001 entries from the latest period to auto-populate:
    - invoicingType (fixed-forfait, by-hours, or hybrid)
    - fixedForfaitValue
    - unitPrice (hourly rate)
    
    If customer_id is provided, updates only that customer.
    Otherwise, updates ALL customers with historical data.
    """
    import re
    
    # Determine which customers to update
    if customer_id:
        customers = await db.customers.find({"id": customer_id}, {"_id": 0}).to_list(1)
        if not customers:
            raise HTTPException(status_code=404, detail="Customer not found")
    else:
        # Get all customers with historical invoices
        customers = await db.customers.find(
            {"historicalInvoices": {"$exists": True, "$ne": []}},
            {"_id": 0}
        ).to_list(10000)
    
    updated_count = 0
    skipped_count = 0
    results = []
    
    for customer in customers:
        customer_name = customer.get("name", "Unknown")
        customer_id_val = customer.get("id")
        historical_invoices = customer.get("historicalInvoices", [])
        
        if not historical_invoices:
            skipped_count += 1
            continue
        
        # Get the latest (most recent) entry
        # Sort by date to ensure we get the latest
        sorted_entries = sorted(
            historical_invoices,
            key=lambda x: x.get("date", ""),
            reverse=True
        )
        
        if not sorted_entries:
            skipped_count += 1
            continue
        
        latest_entry = sorted_entries[0]
        individual_rows = latest_entry.get("individualRows", [])
        
        # Find all Article 000001 entries in the latest period
        article_000001_rows = []
        for row in individual_rows:
            article_code = row.get("articleCode", "").strip()
            if article_code == "000001":
                article_000001_rows.append(row)
        
        if not article_000001_rows:
            skipped_count += 1
            results.append({
                "customerId": customer_id_val,
                "customerName": customer_name,
                "status": "skipped",
                "reason": "No Article 000001 found in latest period"
            })
            continue
        
        # Analyze and determine invoicing type
        invoicing_type = None
        fixed_forfait_value = None
        hourly_rate = None
        
        if len(article_000001_rows) == 1:
            # Single Article 000001 entry
            single_row = article_000001_rows[0]
            detailed_desc = single_row.get("detailedDescription", "").strip()
            unit_price = single_row.get("unitPrice")
            
            # Check if detailed description contains work list (dates, tasks, values)
            has_work_list = False
            if detailed_desc:
                # Look for patterns like "2024-10-17" or "17.10.24" followed by tasks
                date_patterns = [
                    r'\d{4}-\d{2}-\d{2}',  # 2024-10-17
                    r'\d{2}\.\d{2}\.\d{2,4}',  # 17.10.24 or 17.10.2024
                ]
                for pattern in date_patterns:
                    if re.search(pattern, detailed_desc):
                        has_work_list = True
                        break
            
            if has_work_list and unit_price is not None:
                # Case C: By Hours Spent
                invoicing_type = "by-hours"
                hourly_rate = unit_price
            elif unit_price is not None:
                # Case A: Fixed Forfait (empty or simple description)
                invoicing_type = "fixed-forfait"
                fixed_forfait_value = unit_price
        
        elif len(article_000001_rows) >= 2:
            # Case B: Hybrid - Multiple Article 000001 entries
            invoicing_type = "hybrid"
            # First row = Fixed Forfait
            if article_000001_rows[0].get("unitPrice") is not None:
                fixed_forfait_value = article_000001_rows[0]["unitPrice"]
            # Second row = Hourly Rate (usually "dodatna dela")
            if article_000001_rows[1].get("unitPrice") is not None:
                hourly_rate = article_000001_rows[1]["unitPrice"]
        
        # Update customer if we detected settings
        if invoicing_type:
            update_data = {"invoicingType": invoicing_type}
            if fixed_forfait_value is not None:
                update_data["fixedForfaitValue"] = fixed_forfait_value
            if hourly_rate is not None:
                update_data["unitPrice"] = hourly_rate
            
            await db.customers.update_one(
                {"id": customer_id_val},
                {"$set": update_data}
            )
            
            updated_count += 1
            results.append({
                "customerId": customer_id_val,
                "customerName": customer_name,
                "status": "updated",
                "invoicingType": invoicing_type,
                "fixedForfaitValue": fixed_forfait_value,
                "hourlyRate": hourly_rate
            })
            
            logger.info(f"Updated {customer_name}: type={invoicing_type}, forfait={fixed_forfait_value}, hourly={hourly_rate}")
        else:
            skipped_count += 1
            results.append({
                "customerId": customer_id_val,
                "customerName": customer_name,
                "status": "skipped",
                "reason": "Could not determine invoicing type"
            })
    
    return {
        "message": f"Refreshed invoicing settings for {updated_count} customer(s)",
        "updated": updated_count,
        "skipped": skipped_count,
        "total": len(customers),
        "details": results
    }

@api_router.post("/customers/upload-history")
async def upload_customer_history(
    file: UploadFile = File(...),
    customer_ids: str = Form(None),  # Comma-separated customer IDs, or "all"
    current_user: User = Depends(get_current_user)
):
    """Upload historical invoice data from XLSX/XLS file"""
    import uuid
    from datetime import datetime
    
    try:
        contents = await file.read()
        sheet, wb = load_excel_file(contents, file.filename)
        
        # Check for company name in metadata format (Row 3: "Podjetje: JMMC HP d.o.o.")
        company_name_from_metadata = None
        max_row = sheet.xlrd_sheet.nrows if file.filename.endswith('.xls') else sheet.max_row
        for row_num in range(1, min(10, max_row + 1)):
            if file.filename.endswith('.xls'):
                row_values = sheet.xlrd_sheet.row_values(row_num - 1) if row_num <= sheet.xlrd_sheet.nrows else []
            else:
                row_values = [cell.value for cell in sheet[row_num]]
            # Look for "Podjetje:" in first column
            if row_values and row_values[0] and 'podjetje:' in str(row_values[0]).lower():
                # Company name is usually in the next column
                if len(row_values) > 1 and row_values[1]:
                    company_name_from_metadata = str(row_values[1]).strip()
                    logger.info(f"Found company in metadata: {company_name_from_metadata}")
                break
        
        # Find the header row (look for row containing multiple key headers)
        header_row_num = 1
        headers = None
        for row_num in range(1, min(20, max_row + 1)):  # Check first 20 rows
            if file.filename.endswith('.xls'):
                row_values = sheet.xlrd_sheet.row_values(row_num - 1) if row_num <= sheet.xlrd_sheet.nrows else []
            else:
                row_values = [cell.value for cell in sheet[row_num]]
            # Count how many key columns are present
            key_columns_found = 0
            if any(cell and 'poz' in str(cell).lower() and len(str(cell).strip()) < 10 for cell in row_values):
                key_columns_found += 1
            if any(cell and 'kupec' in str(cell).lower() and 'kupec:' not in str(cell).lower() for cell in row_values):
                key_columns_found += 1
            if any(cell and 'dat.dok' in str(cell).lower() for cell in row_values):
                key_columns_found += 1
            if any(cell and 'znesek eur' in str(cell).lower() for cell in row_values):
                key_columns_found += 1
            
            # If we found 3+ key columns, this is likely the header row
            if key_columns_found >= 3:
                headers = row_values
                header_row_num = row_num
                logger.info(f"Found header row at line {row_num} with {key_columns_found} key columns")
                break
        
        if not headers:
            raise HTTPException(status_code=400, detail="Could not find header row. Please ensure XLSX has proper headers.")
        
        logger.info(f"Headers found at row {header_row_num}: {headers}")
        
        # Map column names to indices
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if 'kupec' in header_lower:
                    col_map['customer'] = idx
                elif 'dat.dok' in header_lower or 'datum dokumenta' in header_lower:
                    col_map['date'] = idx
                elif 'šifra art' in header_lower or 'sifra art' in header_lower:
                    col_map['article_code'] = idx
                elif 'naziv artikla' in header_lower:
                    col_map['description'] = idx
                elif 'opis artikla' in header_lower:
                    col_map['alt_description'] = idx
                elif 'znesek eur' in header_lower:
                    col_map['amount'] = idx
                elif 'podjetje' in header_lower:
                    col_map['company'] = idx
                elif 'kol.' in header_lower or 'količina' in header_lower:
                    col_map['quantity'] = idx
                elif header_lower == 'em' or 'enota mere' in header_lower:
                    col_map['unit'] = idx
                elif 'cena' in header_lower and 'brez' not in header_lower:
                    col_map['unit_price'] = idx
        
        logger.info(f"Column mapping: {col_map}")
        logger.info(f"Found alt_description (Opis artikla) at index: {col_map.get('alt_description', 'NOT FOUND')}")
        
        # Extract data by month
        monthly_data = {}  # {customer_name: {month_key: {date, total_amount, descriptions[], company_name}}}
        
        for row in sheet.iter_rows(min_row=header_row_num + 1, values_only=True):
            if not row or all(cell is None or cell == '' for cell in row):
                continue
            
            # Extract values
            customer_name = row[col_map.get('customer')] if 'customer' in col_map else None
            date_val = row[col_map.get('date')] if 'date' in col_map else None
            article_code = row[col_map.get('article_code')] if 'article_code' in col_map else ""
            description = row[col_map.get('description')] if 'description' in col_map else ""
            alt_description = row[col_map.get('alt_description')] if 'alt_description' in col_map else ""
            # Keep both separate - do NOT merge
            
            # Debug logging for first few rows
            if len(monthly_data) == 0 and customer_name:
                logger.info(f"Sample row - Description: '{description}', Alt Description: '{alt_description}'")
            
            amount_val = row[col_map.get('amount')] if 'amount' in col_map else None
            company_name = row[col_map.get('company')] if 'company' in col_map else None
            
            # Extract additional columns
            quantity_val = row[col_map.get('quantity')] if 'quantity' in col_map else None
            unit_val = row[col_map.get('unit')] if 'unit' in col_map else None
            unit_price_val = row[col_map.get('unit_price')] if 'unit_price' in col_map else None
            
            if not customer_name or not date_val:
                continue
            
            customer_name = str(customer_name).strip()
            
            # Clean company name if present
            if company_name:
                company_name = str(company_name).strip()
            
            # Parse date
            date_str = ""
            if hasattr(date_val, 'isoformat'):
                date_obj = date_val
                date_str = date_val.isoformat()
            elif isinstance(date_val, (int, float)):
                # xlrd returns .xls dates as float serial numbers (e.g. 46090.0 = 2026-03-09)
                try:
                    date_obj = xlrd.xldate_as_datetime(date_val, 0)
                    date_str = date_obj.strftime('%Y-%m-%d')
                except:
                    continue
            elif isinstance(date_val, str):
                try:
                    date_obj = datetime.strptime(date_val, '%Y-%m-%d')
                    date_str = date_val
                except:
                    try:
                        date_obj = datetime.strptime(date_val, '%d.%m.%Y')
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except:
                        continue
            else:
                continue
            
            # Get month key (YYYY-MM)
            month_key = f"{date_obj.year}-{str(date_obj.month).zfill(2)}"
            
            # Parse amount
            try:
                if isinstance(amount_val, (int, float)):
                    amount = float(amount_val)
                elif amount_val:
                    amount = float(str(amount_val).replace(',', '.'))
                else:
                    amount = 0.0
            except:
                amount = 0.0
            
            # Parse quantity
            try:
                if isinstance(quantity_val, (int, float)):
                    quantity = float(quantity_val)
                elif quantity_val:
                    quantity = float(str(quantity_val).replace(',', '.'))
                else:
                    quantity = None
            except:
                quantity = None
            
            # Parse unit price
            try:
                if isinstance(unit_price_val, (int, float)):
                    unit_price = float(unit_price_val)
                elif unit_price_val:
                    unit_price = float(str(unit_price_val).replace(',', '.'))
                else:
                    unit_price = None
            except:
                unit_price = None
            
            # Initialize customer data
            if customer_name not in monthly_data:
                monthly_data[customer_name] = {}
            
            # Initialize month data
            if month_key not in monthly_data[customer_name]:
                # Use company from column, or fall back to metadata company
                effective_company_name = company_name if company_name else company_name_from_metadata
                monthly_data[customer_name][month_key] = {
                    'date': date_str,
                    'month': month_key,
                    'total_amount': 0.0,
                    'descriptions': [],
                    'individual_rows': [],  # Store individual row details
                    'company_name': effective_company_name  # Store company name (from column or metadata)
                }
            
            # Accumulate data for the month
            monthly_data[customer_name][month_key]['total_amount'] += amount
            if description and str(description).strip():
                monthly_data[customer_name][month_key]['descriptions'].append(str(description).strip())
            
            # Store individual row with all details
            individual_row = {
                'date': date_str,
                'articleCode': str(article_code).strip() if article_code else "",
                'description': str(description).strip() if description else "",
                'detailedDescription': str(alt_description).strip() if alt_description else "",
                'amount': amount
            }
            
            # Add optional fields if available
            if quantity is not None:
                individual_row['quantity'] = quantity
            if unit_val:
                individual_row['unit'] = str(unit_val).strip()
            if unit_price is not None:
                individual_row['unitPrice'] = unit_price
            
            monthly_data[customer_name][month_key]['individual_rows'].append(individual_row)
        
        logger.info(f"Extracted data for {len(monthly_data)} customers")
        
        # Convert to historical invoice entries
        historical_entries_by_customer = {}
        for customer_name, months in monthly_data.items():
            entries = []
            for month_key, month_data in months.items():
                # Create a single entry per month with combined description
                unique_descriptions = list(set(month_data['descriptions']))
                combined_description = "; ".join(unique_descriptions[:5])  # Limit to 5 unique descriptions
                if len(unique_descriptions) > 5:
                    combined_description += f" (+{len(unique_descriptions) - 5} more)"
                
                entries.append({
                    "date": month_data['date'],
                    "month": month_key,
                    "description": combined_description,
                    "amount": round(month_data['total_amount'], 2),
                    "individualRows": month_data['individual_rows']  # Include individual row details
                })
            
            historical_entries_by_customer[customer_name] = entries
            logger.info(f"Customer '{customer_name}': {len(entries)} monthly entries")
        
        # Process customers: create if not exists, update historical data
        updated_count = 0
        created_count = 0
        total_entries = 0
        
        # Get all customer names from the uploaded data
        all_customer_names = list(historical_entries_by_customer.keys())
        
        # If filtering by specific customer_id (single customer upload from detail page)
        if customer_ids and customer_ids != "all":
            target_ids = [cid.strip() for cid in customer_ids.split(",")]
            
            # For single customer upload, apply ALL data from file to that customer
            if len(target_ids) == 1:
                target_customer = await db.customers.find_one({"id": target_ids[0]})
                if target_customer:
                    logger.info(f"Single customer upload mode: applying all data to {target_customer['name']}")
                    
                    # Extract company name from uploaded data
                    company_name_from_data = None
                    for customer_months in monthly_data.values():
                        for month_data in customer_months.values():
                            if month_data.get('company_name'):
                                company_name_from_data = month_data['company_name']
                                break
                        if company_name_from_data:
                            break
                    
                    # Normalize company name to official names
                    if company_name_from_data:
                        company_lower = company_name_from_data.lower()
                        if 'hp' in company_lower:
                            company_name_from_data = "JMMC HP d.o.o."
                        elif 'finance' in company_lower or 'financa' in company_lower:
                            company_name_from_data = "JMMC Finance d.o.o."
                        else:
                            # Unknown company - set to None to skip assignment
                            logger.warning(f"Unknown company '{company_name_from_data}' - skipping assignment")
                            company_name_from_data = None
                    
                    # Update company if present in data
                    if company_name_from_data:
                        # Find or create company
                        company = await db.companies.find_one({"name": company_name_from_data})
                        if not company:
                            company_id = str(uuid.uuid4())
                            company = {
                                "id": company_id,
                                "name": company_name_from_data
                            }
                            await db.companies.insert_one(company)
                            logger.info(f"Created new company: {company_name_from_data}")
                        else:
                            company_id = company["id"]
                        
                        # Update customer's company
                        await db.customers.update_one(
                            {"id": target_customer["id"]},
                            {"$set": {"companyId": company_id}}
                        )
                        logger.info(f"Updated customer '{target_customer['name']}' with company: {company_name_from_data}")
                    
                    # Combine all entries from all customers in the file
                    all_new_entries = []
                    for entries in historical_entries_by_customer.values():
                        all_new_entries.extend(entries)
                    
                    # Mark as imported
                    for entry in all_new_entries:
                        entry["source"] = "imported"
                    
                    # Extract unit price from last "Računovodstvo" entry
                    unit_price_to_set = None
                    for entry in reversed(all_new_entries):  # Start from most recent
                        individual_rows = entry.get("individualRows", [])
                        for row in reversed(individual_rows):  # Most recent rows first
                            description = row.get("description", "").lower()
                            if "računovodstvo" in description or "racunovodstvo" in description:
                                if row.get("unitPrice") is not None:
                                    unit_price_to_set = row["unitPrice"]
                                    logger.info(f"Found Računovodstvo unit price: €{unit_price_to_set}")
                                    break
                        if unit_price_to_set is not None:
                            break
                    
                    # Extract address service unit price from last "Najem sedeža" entry (Article 000002)
                    address_service_price = None
                    for entry in reversed(all_new_entries):  # Start from most recent
                        individual_rows = entry.get("individualRows", [])
                        for row in reversed(individual_rows):  # Most recent rows first
                            article_code = row.get("articleCode", "").strip()
                            description = row.get("description", "").lower()
                            # Check for article code 000002 OR description containing "najem sedeža"
                            if (article_code == "000002" or 
                                "najem sedeža" in description or 
                                "najem sedeza" in description or
                                "sede legale" in description):
                                if row.get("unitPrice") is not None:
                                    address_service_price = row["unitPrice"]
                                    logger.info(f"Found Address Service (Najem sedeža) unit price: €{address_service_price}")
                                    break
                        if address_service_price is not None:
                            break
                    
                    # AUTO-POPULATE INVOICING TYPE AND PRICING based on Article 000001 from latest period
                    invoicing_type = None
                    fixed_forfait_value = None
                    hourly_rate = None
                    
                    # Get the latest (most recent) entry
                    if all_new_entries:
                        latest_entry = all_new_entries[-1]  # Most recent entry
                        individual_rows = latest_entry.get("individualRows", [])
                        
                        # Find all Article 000001 entries in the latest period
                        article_000001_rows = []
                        for row in individual_rows:
                            article_code = row.get("articleCode", "").strip()
                            if article_code == "000001":
                                article_000001_rows.append(row)
                        
                        logger.info(f"Found {len(article_000001_rows)} Article 000001 entries in latest period")
                        
                        if len(article_000001_rows) == 1:
                            # Single Article 000001 entry
                            single_row = article_000001_rows[0]
                            detailed_desc = single_row.get("detailedDescription", "").strip()
                            unit_price = single_row.get("unitPrice")
                            
                            # Check if detailed description contains work list (dates, tasks, values)
                            has_work_list = False
                            if detailed_desc:
                                # Look for patterns like "2024-10-17" or "17.10.24" followed by tasks
                                import re
                                date_patterns = [
                                    r'\d{4}-\d{2}-\d{2}',  # 2024-10-17
                                    r'\d{2}\.\d{2}\.\d{2,4}',  # 17.10.24 or 17.10.2024
                                ]
                                for pattern in date_patterns:
                                    if re.search(pattern, detailed_desc):
                                        has_work_list = True
                                        break
                            
                            if has_work_list and unit_price is not None:
                                # Case C: By Hours Spent
                                invoicing_type = "by-hours"
                                hourly_rate = unit_price
                                logger.info(f"Auto-detected: By Hours Spent (hourly rate: €{hourly_rate})")
                            elif unit_price is not None:
                                # Case A: Fixed Forfait (empty or simple description)
                                invoicing_type = "fixed-forfait"
                                fixed_forfait_value = unit_price
                                logger.info(f"Auto-detected: Fixed Forfait (value: €{fixed_forfait_value})")
                        
                        elif len(article_000001_rows) >= 2:
                            # Case B: Hybrid - Multiple Article 000001 entries
                            invoicing_type = "hybrid"
                            # First row = Fixed Forfait
                            if article_000001_rows[0].get("unitPrice") is not None:
                                fixed_forfait_value = article_000001_rows[0]["unitPrice"]
                            # Second row = Hourly Rate (usually "dodatna dela")
                            if article_000001_rows[1].get("unitPrice") is not None:
                                hourly_rate = article_000001_rows[1]["unitPrice"]
                            logger.info(f"Auto-detected: Hybrid (forfait: €{fixed_forfait_value}, hourly: €{hourly_rate})")
                    
                    # Merge by month: overwrite only periods present in the new file,
                    # preserve all existing entries (manual AND imported) for other periods
                    existing_history = target_customer.get("historicalInvoices", [])
                    new_month_keys = {entry["month"] for entry in all_new_entries}
                    preserved_entries = [e for e in existing_history if e.get("month") not in new_month_keys]
                    merged_history = preserved_entries + all_new_entries
                    logger.info(f"Merge: {len(existing_history)} existing → kept {len(preserved_entries)} (outside new periods) + {len(all_new_entries)} new = {len(merged_history)} total")
                    
                    # Prepare update data
                    update_data = {"historicalInvoices": merged_history}
                    if unit_price_to_set is not None:
                        update_data["unitPrice"] = unit_price_to_set
                        logger.info(f"Updating customer unit price to: €{unit_price_to_set}")
                    if address_service_price is not None:
                        update_data["addressServiceUnitPrice"] = address_service_price
                        logger.info(f"Updating address service unit price to: €{address_service_price}")
                    
                    # Add auto-detected invoicing settings
                    if invoicing_type is not None:
                        update_data["invoicingType"] = invoicing_type
                        logger.info(f"Setting invoicing type to: {invoicing_type}")
                    if fixed_forfait_value is not None:
                        update_data["fixedForfaitValue"] = fixed_forfait_value
                        logger.info(f"Setting fixed forfait value to: €{fixed_forfait_value}")
                    if hourly_rate is not None:
                        update_data["unitPrice"] = hourly_rate  # Use unitPrice for hourly rate
                        logger.info(f"Setting hourly rate to: €{hourly_rate}")
                    
                    # Update in database
                    await db.customers.update_one(
                        {"id": target_customer["id"]},
                        {"$set": update_data}
                    )
                    updated_count = 1
                    total_entries = len(all_new_entries)
                    
                    return {
                        "message": f"Historical data uploaded and grouped by month",
                        "customersUpdated": updated_count,
                        "customersCreated": 0,
                        "monthlyEntriesCreated": total_entries
                    }
        
        # For "all" mode or multiple customers, match by name
        for customer_name in all_customer_names:
            # Get company name from first month's data (they should all be the same)
            company_name_from_data = None
            first_month = list(historical_entries_by_customer[customer_name])[0] if historical_entries_by_customer[customer_name] else None
            if first_month:
                months = monthly_data.get(customer_name, {})
                for month_data in months.values():
                    if month_data.get('company_name'):
                        company_name_from_data = month_data['company_name']
                        break
            
            # Normalize company name to official names
            if company_name_from_data:
                company_lower = company_name_from_data.lower()
                if 'hp' in company_lower:
                    company_name_from_data = "JMMC HP d.o.o."
                elif 'finance' in company_lower or 'financa' in company_lower:
                    company_name_from_data = "JMMC Finance d.o.o."
                else:
                    # Unknown company - set to None to skip assignment
                    logger.warning(f"Unknown company '{company_name_from_data}' - skipping assignment")
                    company_name_from_data = None
            
            # Find or create company if company name exists
            company_id = None
            if company_name_from_data:
                # Check if company exists
                company = await db.companies.find_one({"name": company_name_from_data})
                if not company:
                    # Create company
                    company_id = str(uuid.uuid4())
                    company = {
                        "id": company_id,
                        "name": company_name_from_data
                    }
                    await db.companies.insert_one(company)
                    logger.info(f"Created new company: {company_name_from_data}")
                else:
                    company_id = company["id"]
                    logger.info(f"Found existing company: {company_name_from_data} (ID: {company_id})")
            
            # Check if customer exists
            customer = await db.customers.find_one({"name": customer_name})
            
            logger.info(f"Processing customer '{customer_name}' - Found in DB: {customer is not None}")
            
            # If filtering by specific customer_ids, skip if not in the list
            if customer_ids and customer_ids != "all":
                target_ids = [cid.strip() for cid in customer_ids.split(",")]
                logger.info(f"Filtering by customer IDs: {target_ids}")
                
                if customer and customer["id"] not in target_ids:
                    logger.info(f"Skipping {customer_name} - ID {customer['id']} not in target list")
                    continue
                elif not customer:
                    # Customer doesn't exist, but we're filtering by ID
                    # Check if we should create it anyway when uploading to "all"
                    logger.info(f"Customer '{customer_name}' not found in DB and filtering by IDs - skipping creation")
                    continue
            
            # Create customer if doesn't exist
            if not customer:
                customer_id = str(uuid.uuid4())
                customer = {
                    "id": customer_id,
                    "name": customer_name,
                    "unitPrice": 0,
                    "historicalInvoices": []
                }
                if company_id:
                    customer["companyId"] = company_id
                await db.customers.insert_one(customer)
                created_count += 1
                logger.info(f"Created new customer: {customer_name} with company: {company_name_from_data if company_id else 'None'}")
            else:
                # Update company if it exists in the data
                if company_id:
                    await db.customers.update_one(
                        {"id": customer["id"]},
                        {"$set": {"companyId": company_id}}
                    )
                    logger.info(f"Updated customer '{customer_name}' with company: {company_name_from_data}")
            
            # Update historical invoices - keep manual entries, replace imported ones
            existing_history = customer.get("historicalInvoices", [])
            new_entries = historical_entries_by_customer[customer_name]
            
            # Mark new entries as imported
            for entry in new_entries:
                entry["source"] = "imported"
            
            # Extract unit price from last "Računovodstvo" entry
            unit_price_to_set = None
            for entry in reversed(new_entries):  # Start from most recent
                individual_rows = entry.get("individualRows", [])
                for row in reversed(individual_rows):  # Most recent rows first
                    description = row.get("description", "").lower()
                    if "računovodstvo" in description or "racunovodstvo" in description:
                        if row.get("unitPrice") is not None:
                            unit_price_to_set = row["unitPrice"]
                            logger.info(f"Found Računovodstvo unit price for {customer_name}: €{unit_price_to_set}")
                            break
                if unit_price_to_set is not None:
                    break
            
            # Merge by month: overwrite only periods present in the new file,
            # preserve all existing entries (manual AND imported) for other periods
            new_month_keys = {entry["month"] for entry in new_entries}
            preserved_entries = [e for e in existing_history if e.get("month") not in new_month_keys]
            merged_history = preserved_entries + new_entries
            logger.info(f"Merge '{customer_name}': {len(existing_history)} existing → kept {len(preserved_entries)} + {len(new_entries)} new = {len(merged_history)} total")
            
            # Prepare update data
            update_data = {"historicalInvoices": merged_history}
            if unit_price_to_set is not None:
                update_data["unitPrice"] = unit_price_to_set
                logger.info(f"Updating customer '{customer_name}' unit price to: €{unit_price_to_set}")
            
            # Update in database
            await db.customers.update_one(
                {"id": customer["id"]},
                {"$set": update_data}
            )
            updated_count += 1
            total_entries += len(new_entries)
        
        return {
            "message": f"Historical data uploaded and grouped by month",
            "customersUpdated": updated_count,
            "customersCreated": created_count,
            "monthlyEntriesCreated": total_entries
        }
        
    except Exception as e:
        logger.error(f"Error uploading historical data: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

# ============ TIME ENTRIES ============
@api_router.post("/time-entries/{entry_id}/move-customer")
async def move_time_entry_to_customer(
    entry_id: str, 
    new_customer_id: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """Move a time entry to a different customer"""
    logger.info(f"Move request: entry_id={entry_id}, new_customer_id={new_customer_id}")
    
    # Find the time entry
    entry = await db.timeEntries.find_one({"id": entry_id})
    if not entry:
        logger.error(f"Time entry not found: {entry_id}")
        raise HTTPException(status_code=404, detail="Time entry not found")
    
    old_customer_id = entry.get("customerId")
    logger.info(f"Found entry. Old customer: {old_customer_id}, New customer: {new_customer_id}")
    
    # Verify new customer exists
    new_customer = await db.customers.find_one({"id": new_customer_id})
    if not new_customer:
        logger.error(f"Target customer not found: {new_customer_id}")
        raise HTTPException(status_code=404, detail="Target customer not found")
    
    # Update the time entry
    result = await db.timeEntries.update_one(
        {"id": entry_id},
        {"$set": {"customerId": new_customer_id}}
    )
    logger.info(f"Time entry updated. Modified count: {result.modified_count}")
    
    # Find affected invoices and update them
    batch_id = entry.get("batchId")
    
    # Remove this entry from old customer's invoice lines in invoiceLines collection
    if old_customer_id:
        old_invoice = await db.invoices.find_one({
            "batchId": batch_id,
            "customerId": old_customer_id
        })
        
        if old_invoice:
            # Delete line from invoiceLines collection
            delete_result = await db.invoiceLines.delete_many({
                "invoiceId": old_invoice["id"],
                "timeEntryId": entry_id
            })
            
            # Recalculate total from remaining lines
            remaining_lines = await db.invoiceLines.find(
                {"invoiceId": old_invoice["id"]},
                {"_id": 0}
            ).to_list(1000)
            
            new_total = sum(line.get("amount", 0) for line in remaining_lines)
            
            await db.invoices.update_one(
                {"id": old_invoice["id"]},
                {"$set": {"total": new_total}}
            )
            logger.info(f"Removed {delete_result.deleted_count} line(s) from old invoice. New line count: {len(remaining_lines)}, New total: {new_total}")
    
    # Add to new customer's invoice or create if doesn't exist
    new_invoice = await db.invoices.find_one({
        "batchId": batch_id,
        "customerId": new_customer_id
    })
    
    # Get project info for the line description
    project = await db.projects.find_one({"id": entry.get("projectId")})
    project_name = project.get("name", "General") if project else "General"
    
    # Create new line item for this entry
    hours = entry.get("hours", 0)
    value = entry.get("value", 0)
    unit_price = (value / hours) if hours > 0 else 0
    
    new_line_id = str(uuid.uuid4())
    new_line_doc = {
        "id": new_line_id,
        "invoiceId": "",  # Will be set below
        "timeEntryId": entry_id,
        "description": f"{project_name} - {entry.get('employeeName', 'Unknown')} - {entry.get('notes', '')}",
        "quantity": hours,
        "unitPrice": unit_price,
        "amount": value,
        "taxCode": None
    }
    
    logger.info(f"Creating new line: hours={hours}, value={value}, unitPrice={unit_price}")
    
    if new_invoice:
        # Add line to existing invoice in invoiceLines collection
        new_line_doc["invoiceId"] = new_invoice["id"]
        await db.invoiceLines.insert_one(new_line_doc)
        
        # Recalculate total from all lines
        all_lines = await db.invoiceLines.find(
            {"invoiceId": new_invoice["id"]},
            {"_id": 0}
        ).to_list(1000)
        new_total = sum(line.get("amount", 0) for line in all_lines)
        
        await db.invoices.update_one(
            {"id": new_invoice["id"]},
            {"$set": {"total": new_total}}
        )
        logger.info(f"Added entry to existing invoice {new_invoice['id']}. New line count: {len(all_lines)}, New total: {new_total}")
    else:
        # Create new invoice for this customer
        batch = await db.importBatches.find_one({"id": batch_id})
        
        new_invoice_id = str(uuid.uuid4())
        await db.invoices.insert_one({
            "id": new_invoice_id,
            "batchId": batch_id,
            "customerId": new_customer_id,
            "customerName": new_customer.get("name", ""),
            "invoiceDate": batch.get("invoiceDate"),
            "dueDate": batch.get("dueDate"),
            "periodFrom": batch.get("periodFrom"),
            "periodTo": batch.get("periodTo"),
            "total": value,
            "status": "draft",
            "createdAt": datetime.now(timezone.utc).isoformat()
        })
        
        # Add line to invoiceLines collection
        new_line_doc["invoiceId"] = new_invoice_id
        await db.invoiceLines.insert_one(new_line_doc)
        
        logger.info(f"Created new invoice {new_invoice_id} for customer {new_customer.get('name')} with 1 line, total: {value}")
    

@api_router.post("/invoices/migrate-time-entry-ids")
async def migrate_time_entry_ids(current_user: User = Depends(get_current_user)):
    """Migration: Add timeEntryId to existing invoice lines"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get all invoice lines
    all_lines = await db.invoiceLines.find({}, {"_id": 0}).to_list(100000)
    
    updated_count = 0
    for line in all_lines:
        # Skip if already has timeEntryId
        if line.get("timeEntryId"):
            continue
        
        # Try to find matching time entry based on description, quantity, and invoice
        invoice = await db.invoices.find_one({"id": line["invoiceId"]})
        if not invoice:
            continue
        
        batch_id = invoice.get("batchId")
        customer_id = invoice.get("customerId")
        
        # Find time entry that matches this line
        # Match by hours (quantity), customerId, and batchId
        matching_entry = await db.timeEntries.find_one({
            "batchId": batch_id,
            "customerId": customer_id,
            "hours": line.get("quantity", 0)
        })
        
        if matching_entry:
            # Update the line with timeEntryId
            await db.invoiceLines.update_one(
                {"id": line["id"]},
                {"$set": {"timeEntryId": matching_entry["id"]}}
            )
            updated_count += 1
    
    logger.info(f"Migration complete: Updated {updated_count} invoice lines with timeEntryId")
    return {"message": f"Updated {updated_count} invoice lines", "total_lines": len(all_lines)}


    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "move_time_entry",
        "entity": "TimeEntry",
        "entityId": entry_id,
        "metadata": {
            "oldCustomerId": old_customer_id,
            "newCustomerId": new_customer_id
        },
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    logger.info(f"Move complete: Entry {entry_id} moved from {old_customer_id} to {new_customer_id}")
    
    return {
        "message": "Time entry moved successfully",
        "oldCustomerId": old_customer_id,
        "newCustomerId": new_customer_id
    }


@api_router.post("/invoices/compose")
async def compose_invoices(batchId: str, current_user: User = Depends(get_current_user)):
    batch = await db.importBatches.find_one({"id": batchId})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get only BILLABLE time entries (uninvoiced, ready - exclude forfait, internal, free, and already invoiced)
    entries = await db.timeEntries.find({
        "batchId": batchId, 
        "status": {"$in": ["uninvoiced", "ready"]}
    }).to_list(10000)
    
    # Group by customer
    customer_groups = defaultdict(list)
    for entry in entries:
        customer_groups[entry["customerId"]].append(entry)
    
    # Create invoices
    invoice_ids = []
    for customer_id, customer_entries in customer_groups.items():
        customer = await db.customers.find_one({"id": customer_id})
        
        invoice_id = str(uuid.uuid4())
        invoice_doc = {
            "id": invoice_id,
            "batchId": batchId,
            "customerId": customer_id,
            "customerName": customer["name"],
            "invoiceDate": batch["invoiceDate"],
            "periodFrom": batch["periodFrom"],
            "periodTo": batch["periodTo"],
            "dueDate": batch["dueDate"],
            "status": "draft",
            "total": sum(e["value"] for e in customer_entries),
            "createdAt": datetime.now(timezone.utc).isoformat()
        }
        await db.invoices.insert_one(invoice_doc)
        
        # Create invoice lines
        lines = []
        customer_unit_price = customer.get("unitPrice", 0)  # Get customer's default unit price
        
        for entry in customer_entries:
            line_id = str(uuid.uuid4())
            project = await db.projects.find_one({"id": entry["projectId"]})
            
            # Determine unit price: use customer's default if available and hours > 0, otherwise calculate
            if customer_unit_price > 0 and entry["hours"] > 0:
                unit_price = customer_unit_price
                amount = entry["hours"] * customer_unit_price
            else:
                # Fallback to calculated price from imported data
                unit_price = entry["value"] / entry["hours"] if entry["hours"] > 0 else 0
                amount = entry["value"]
            
            line_doc = {
                "id": line_id,
                "invoiceId": invoice_id,
                "timeEntryId": entry["id"],  # Add timeEntryId for move functionality
                "description": f"{project['name']} - {entry['employeeName']} - {entry['notes'] or ''}",
                "quantity": entry["hours"],
                "unitPrice": unit_price,
                "amount": amount,
                "taxCode": None
            }
            lines.append(line_doc)
        
        if lines:
            await db.invoiceLines.insert_many(lines)
            
            # Recalculate invoice total from actual line amounts
            line_total = sum(line["amount"] for line in lines)
            await db.invoices.update_one(
                {"id": invoice_id},
                {"$set": {"total": line_total}}
            )
        
        # Mark all entries in this invoice as "invoiced" (change status to 'invoiced')
        entry_ids_to_mark = [entry["id"] for entry in customer_entries]
        await db.timeEntries.update_many(
            {"id": {"$in": entry_ids_to_mark}},
            {"$set": {"status": "invoiced"}}
        )
        
        invoice_ids.append(invoice_id)
    
    # Update batch status
    await db.importBatches.update_one(
        {"id": batchId},
        {"$set": {"status": "composed"}}
    )
    
    return {"invoiceIds": invoice_ids}

@api_router.post("/invoices/compose-filtered")
async def compose_filtered_invoices(request: dict, current_user: User = Depends(get_current_user)):
    """
    Compose invoices only for specified time entry IDs (filtered rows).
    Accepts: { batchId, entryIds: [list of time entry IDs] }
    
    Enhanced with forfait linking logic:
    - Rows with status "uninvoiced" or "ready" are posted 1:1
    - Rows with status "forfait" (and src != "forfait_batch") are linked to forfait_batch rows
    - Validation: Each customer can have only 1 forfait_batch row with tariff "001 - Računovodstvo"
    """
    batch_id = request.get('batchId')
    entry_ids = request.get('entryIds', [])
    
    if not batch_id:
        raise HTTPException(status_code=400, detail="batchId is required")
    
    batch = await db.importBatches.find_one({"id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get ONLY the specified time entries that are billable (uninvoiced, ready)
    # EXCLUDE forfait_batch entries here (they are retrieved separately)
    entries = await db.timeEntries.find({
        "batchId": batch_id,
        "id": {"$in": entry_ids},
        "status": {"$in": ["uninvoiced", "ready"]},
        "entrySource": {"$ne": "forfait_batch"}  # CRITICAL: Exclude forfait_batch to prevent double-counting
    }).to_list(10000)
    
    # ALSO get forfait entries (status="forfait", src != "forfait_batch") for linking
    forfait_entries = await db.timeEntries.find({
        "batchId": batch_id,
        "id": {"$in": entry_ids},
        "status": "forfait",
        "entrySource": {"$ne": "forfait_batch"}  # Exclude forfait_batch entries
    }).to_list(10000)
    
    # Get all forfait_batch entries (src="forfait_batch") for validation
    forfait_batch_entries = await db.timeEntries.find({
        "batchId": batch_id,
        "id": {"$in": entry_ids},
        "entrySource": "forfait_batch"
    }).to_list(10000)
    
    if not entries and not forfait_batch_entries:
        raise HTTPException(status_code=400, detail="No matching time entries found")

    # --- Batch-load all customers and projects needed (2 queries) ---
    all_entry_docs = entries + forfait_entries + forfait_batch_entries
    needed_customer_ids = list({e["customerId"] for e in all_entry_docs if e.get("customerId")})
    needed_project_ids  = list({e["projectId"]  for e in all_entry_docs if e.get("projectId")})

    customers_bulk = await db.customers.find(
        {"id": {"$in": needed_customer_ids}}, {"_id": 0, "id": 1, "name": 1, "unitPrice": 1}
    ).to_list(None)
    customer_map = {c["id"]: c for c in customers_bulk}

    projects_bulk = await db.projects.find(
        {"id": {"$in": needed_project_ids}}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(None)
    project_map = {p["id"]: p for p in projects_bulk}

    # VALIDATION: Check forfait logic for each customer
    # Group forfait entries and forfait_batch entries by customer
    forfait_by_customer = defaultdict(list)
    for entry in forfait_entries:
        forfait_by_customer[entry["customerId"]].append(entry)
    
    forfait_batch_by_customer = defaultdict(list)
    for entry in forfait_batch_entries:
        forfait_batch_by_customer[entry["customerId"]].append(entry)
    
    # VALIDATION 1: For ALL customers with forfait_batch entries, ensure only 1 with tariff "001 - Računovodstvo"
    for customer_id, customer_fb_entries in forfait_batch_by_customer.items():
        customer = customer_map.get(customer_id)
        customer_name = customer.get("name", "Unknown") if customer else "Unknown"
        
        # Count forfait_batch entries with tariff "001 - Računovodstvo"
        forfait_batch_001 = []
        for fb_entry in customer_fb_entries:
            tariff_code = fb_entry.get("tariff", "")
            if "001" in tariff_code and "Računovodstvo" in tariff_code:
                forfait_batch_001.append(fb_entry)
        
        # Check: Maximum 1 forfait_batch with tariff 001 allowed per customer
        if len(forfait_batch_001) > 1:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot create invoice - Customer '{customer_name}' has multiple forfait batch entries with tariff 001 - Računovodstvo. Only 1 is allowed."
            )
    
    # VALIDATION 2: For customers with forfait entries, ensure they have exactly 1 forfait_batch with tariff "001 - Računovodstvo"
    for customer_id, customer_forfait_entries in forfait_by_customer.items():
        customer = customer_map.get(customer_id)
        customer_name = customer.get("name", "Unknown") if customer else "Unknown"
        
        # Find forfait_batch entries with tariff "001 - Računovodstvo"
        forfait_batch_001 = []
        if customer_id in forfait_batch_by_customer:
            for fb_entry in forfait_batch_by_customer[customer_id]:
                tariff_code = fb_entry.get("tariff", "")
                if "001" in tariff_code and "Računovodstvo" in tariff_code:
                    forfait_batch_001.append(fb_entry)
        
        # Check: Must have exactly 1 forfait_batch with tariff 001 to link forfait entries
        if len(forfait_batch_001) == 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot create invoice - Customer '{customer_name}' has forfait entries but no forfait batch entry with tariff 001 - Računovodstvo."
            )
    
    # Group by customer (include ONLY regular entries and forfait_batch entries)
    # Forfait entries are EXCLUDED from line items - they are only used for linking
    all_billable_entries = entries + forfait_batch_entries
    customer_groups = defaultdict(list)
    for entry in all_billable_entries:
        customer_groups[entry["customerId"]].append(entry)
    
    # Debug logging
    print(f"DEBUG: Total billable entries for invoice lines: {len(all_billable_entries)}")
    print(f"  - Regular entries (uninvoiced/ready): {len(entries)}")
    print(f"  - Forfait_batch entries: {len(forfait_batch_entries)}")
    print(f"  - Forfait entries (linked only, NOT line items): {len(forfait_entries)}")

    # Create invoices
    invoice_ids = []
    for customer_id, customer_entries in customer_groups.items():
        customer = customer_map.get(customer_id) or {}
        
        invoice_id = str(uuid.uuid4())
        invoice_doc = {
            "id": invoice_id,
            "batchId": batch_id,
            "customerId": customer_id,
            "customerName": customer["name"],
            "invoiceDate": batch["invoiceDate"],
            "periodFrom": batch["periodFrom"],
            "periodTo": batch["periodTo"],
            "dueDate": batch["dueDate"],
            "status": "draft",
            "total": sum(e["value"] for e in customer_entries),
            "createdAt": datetime.now(timezone.utc).isoformat()
        }
        await db.invoices.insert_one(invoice_doc)
        
        # Create invoice lines
        lines = []
        customer_unit_price = customer.get("unitPrice", 0)
        
        # Get forfait entries for this customer for linking
        customer_forfait_entries = forfait_by_customer.get(customer_id, [])
        
        # Debug logging for line creation
        print(f"DEBUG: Creating lines for customer {customer['name']}")
        print(f"  - Customer entries to process: {len(customer_entries)}")
        print(f"  - Forfait entries for linking: {len(customer_forfait_entries)}")

        for entry in customer_entries:
            line_id = str(uuid.uuid4())
            project = project_map.get(entry.get("projectId")) or {}
            
            # Check if this is a forfait_batch entry with tariff 001 - Računovodstvo
            is_forfait_batch_001 = False
            forfait_details_text = ""
            
            if entry.get("entrySource") == "forfait_batch":
                tariff_code = entry.get("tariff", "")
                if "001" in tariff_code and "Računovodstvo" in tariff_code:
                    is_forfait_batch_001 = True
                    
                    # Build forfait details text from linked forfait entries
                    # Format: dd.mm.yyyy | description | X.Xh (one per line)
                    forfait_details_lines = []
                    for f_entry in customer_forfait_entries:
                        # Convert date to EU format (dd.mm.yyyy)
                        entry_date = f_entry.get("date", "")
                        try:
                            if entry_date:
                                date_obj = datetime.fromisoformat(entry_date) if isinstance(entry_date, str) else entry_date
                                eu_date = date_obj.strftime("%d.%m.%Y")
                            else:
                                eu_date = ""
                        except:
                            eu_date = entry_date
                        
                        description = f_entry.get("notes", "") or f_entry.get("description", "")
                        hours = f_entry.get("hours", 0)
                        
                        forfait_details_lines.append(f"{eu_date} | {description} | {hours}h")
                    
                    forfait_details_text = "\n".join(forfait_details_lines)
            
            # Determine unit price
            if customer_unit_price > 0 and entry["hours"] > 0:
                unit_price = customer_unit_price
                amount = entry["hours"] * customer_unit_price
            else:
                unit_price = entry["value"] / entry["hours"] if entry["hours"] > 0 else 0
                amount = entry["value"]
            
            # Build description
            base_description = f"{project['name']} - {entry['employeeName']} - {entry['notes'] or ''}"
            
            # For forfait_batch with 001, prepend forfait details
            if is_forfait_batch_001 and forfait_details_text:
                description = f"{forfait_details_text}\n\n{base_description}"
            else:
                description = base_description
            
            line_doc = {
                "id": line_id,
                "invoiceId": invoice_id,
                "timeEntryId": entry["id"],
                "description": description,
                "quantity": entry["hours"],
                "unitPrice": unit_price,
                "amount": amount,
                "taxCode": None,
                "forfaitDetails": forfait_details_text if is_forfait_batch_001 else None  # Store separately for frontend rendering
            }
            lines.append(line_doc)
        
        if lines:
            await db.invoiceLines.insert_many(lines)
            
            # Recalculate invoice total from actual line amounts
            line_total = sum(line["amount"] for line in lines)
            await db.invoices.update_one(
                {"id": invoice_id},
                {"$set": {"total": line_total}}
            )
        
        
        # DO NOT change row statuses after posting - keep them as set by user
        # Rows should maintain their original status (uninvoiced, ready, internal, free, forfait)
        # No status update needed - statuses remain as user configured them
        
        invoice_ids.append(invoice_id)
    
    # Update batch status
    await db.importBatches.update_one(
        {"id": batch_id},
        {"$set": {"status": "composed"}}
    )
    
    # Calculate total entries processed (ONLY billable line items: regular + forfait_batch)
    # Forfait entries are NOT counted as they don't create separate line items
    total_processed = len(entries) + len(forfait_batch_entries)
    
    return {"invoiceIds": invoice_ids, "entriesProcessed": total_processed}


@api_router.get("/invoices/{invoice_id}", response_model=Dict[str, Any])
async def get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    lines = await db.invoiceLines.find({"invoiceId": invoice_id}, {"_id": 0}).to_list(1000)
    
    return {"invoice": invoice, "lines": lines}

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, update: InvoiceUpdate, current_user: User = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Update invoice header fields if provided
    header_updates = {}
    # Don't allow changing invoice number if already posted (set by e-računi)
    if update.number and invoice.get("status") != "posted":
        header_updates["number"] = update.number
    if update.invoiceDate:
        header_updates["invoiceDate"] = update.invoiceDate
    if update.dueDate:
        header_updates["dueDate"] = update.dueDate
    if update.periodFrom:
        header_updates["periodFrom"] = update.periodFrom
    if update.periodTo:
        header_updates["periodTo"] = update.periodTo
    
    if header_updates:
        await db.invoices.update_one({"id": invoice_id}, {"$set": header_updates})
    
    # Update lines and recalculate amounts
    await db.invoiceLines.delete_many({"invoiceId": invoice_id})
    
    total = 0
    for line_data in update.lines:
        line_id = line_data.get("id") or str(uuid.uuid4())
        amount = line_data["quantity"] * line_data["unitPrice"]
        total += amount
        
        line_doc = {
            "id": line_id,
            "invoiceId": invoice_id,
            "description": line_data["description"],
            "quantity": line_data["quantity"],
            "unitPrice": line_data["unitPrice"],
            "amount": amount,
            "taxCode": line_data.get("taxCode")
        }
        await db.invoiceLines.insert_one(line_doc)
    
    # Update total and mark as edited if status is imported
    update_fields = {"total": total}
    if invoice.get("status") == "imported":
        update_fields["status"] = "edited"
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_fields})
    

@api_router.put("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: str, 
    new_status: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """Update invoice status"""
    valid_statuses = ["imported", "edited", "draft", "issued", "posted", "deleted"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": new_status}}
    )
    
    return {"message": f"Invoice status updated to {new_status}"}

@api_router.post("/invoices/{invoice_id}/confirm-draft")
async def confirm_draft(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Confirm invoice as draft"""
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": "draft"}}
    )
    
    return {"message": "Invoice confirmed as draft"}

@api_router.post("/invoices/{invoice_id}/issue")
async def issue_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Issue invoice (mark as issued)"""
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin role required")
    
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": "issued"}}
    )
    
    return {"message": "Invoice issued successfully"}

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Completely delete invoice and its lines from database, and reset time entries to uninvoiced"""
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    batch_id = invoice.get("batchId")
    customer_id = invoice.get("customerId")
    
    # Count lines before deletion for audit
    lines_count = await db.invoiceLines.count_documents({"invoiceId": invoice_id})
    
    # Find all time entries that were invoiced for this customer in this batch
    # These are the entries that were used to create this invoice
    # We need to set their status back to "uninvoiced"
    if batch_id and customer_id:
        # Update all time entries for this batch+customer that are "invoiced" back to "uninvoiced"
        update_result = await db.timeEntries.update_many(
            {
                "batchId": batch_id,
                "customerId": customer_id,
                "status": "invoiced"
            },
            {"$set": {"status": "uninvoiced"}}
        )
        time_entries_updated = update_result.modified_count
    else:
        time_entries_updated = 0
    
    # Delete all invoice lines first
    await db.invoiceLines.delete_many({"invoiceId": invoice_id})
    
    # Delete the invoice itself
    await db.invoices.delete_one({"id": invoice_id})
    
    # Audit event
    await db.auditEvents.insert_one({
        "id": str(uuid.uuid4()),
        "actorId": current_user.email,
        "action": "delete_invoice",
        "entity": "Invoice",
        "entityId": invoice_id,
        "metadata": {
            "customerName": invoice.get("customerName", "Unknown"),
            "total": invoice.get("total", 0),
            "linesDeleted": lines_count,
            "timeEntriesResetToUninvoiced": time_entries_updated,
            "batchId": batch_id
        },
        "at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": "Invoice completely deleted",
        "linesDeleted": lines_count,
        "timeEntriesResetToUninvoiced": time_entries_updated
    }


    return {"message": "Invoice updated"}

@api_router.post("/invoices/{invoice_id}/post")
async def post_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin role required")
    
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get invoice lines
    lines = await db.invoiceLines.find({"invoiceId": invoice_id}, {"_id": 0}).to_list(1000)
    
    # Get customer details
    customer = await db.customers.find_one({"id": invoice.get("customerId")})
    
    if ERACUNI_MODE == "stub":
        external_number = f"ER-STUB-{int(datetime.now(timezone.utc).timestamp())}"
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "status": "posted",
                "number": external_number,  # Set invoice number from stub
                "externalNumber": external_number
            }}
        )
        
        # Audit event
        await db.auditEvents.insert_one({
            "id": str(uuid.uuid4()),
            "actorId": current_user.email,
            "action": "post_invoice",
            "entity": "Invoice",
            "entityId": invoice_id,
            "at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"externalNumber": external_number, "status": "posted"}
    
    # Real e-računi integration
    else:
        import httpx
        
        # Get e-računi settings
        user_settings = await db.aiSettings.find_one({"userId": current_user.email})
        if not user_settings or not user_settings.get("eracuniUsername") or not user_settings.get("eracuniToken"):
            raise HTTPException(status_code=400, detail="e-računi credentials not configured in Settings")
        
        endpoint = user_settings.get("eracuniEndpoint", "https://e-racuni.com/WebServicesSI/API")
        
        # Build e-računi API payload
        items = []
        for line in lines:
            # Use unitPrice directly, don't set to None if 0
            net_price = line.get("unitPrice", 0)
            items.append({
                "description": line.get("description", "Services"),
                "productCode": "000001",
                "quantity": line.get("quantity", 1),
                "unit": "h",
                "netPrice": net_price
            })
        
        payload = {
            "username": user_settings["eracuniUsername"],
            "secretKey": user_settings["eracuniSecretKey"],
            "token": user_settings["eracuniToken"],
            "method": "SalesInvoiceCreate",
            "parameters": {
                "SalesInvoice": {
                    "status": "IssuedInvoice",
                    "dateOfSupplyFrom": invoice.get("invoiceDate"),
                    "date": invoice.get("invoiceDate"),
                    "documentCurrency": "EUR",
                    "documentLanguage": "English",
                    "vatTransactionType": "0",
                    "type": "Gross",
                    "methodOfPayment": "BankTransfer",
                    "buyerName": customer.get("name", "") if customer else "",
                    "buyerStreet": "",
                    "buyerPostalCode": "",
                    "buyerCity": "",
                    "buyerCountry": "SI",
                    "buyerEMail": "",
                    "buyerPhone": "",
                    "buyerCode": "",
                    "buyerDocumentID": "",
                    "buyerTaxNumber": "",
                    "buyerVatRegistration": "None",
                    "Items": items,
                    "city": "Nova Gorica"
                }
            }
        }
        
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.post(
                    endpoint,
                    json=payload,
                    headers={"Content-Type": "application/json"}
                )
                
                logger.info(f"e-računi post - Status: {response.status_code}")
                
                if response.status_code == 200:
                    result = response.json()
                    
                    # Log full response for debugging
                    logger.info(f"e-računi response: {result}")
                    
                    # Check for errors in response
                    if result.get("response", {}).get("status") == "ok":
                        # Success! Extract document info
                        result_data = result.get("response", {}).get("result", {})
                        external_number = result_data.get("number")
                        document_id = result_data.get("documentID")
                        
                        await db.invoices.update_one(
                            {"id": invoice_id},
                            {"$set": {
                                "status": "posted",
                                "number": external_number,  # Set invoice number from e-računi
                                "externalNumber": external_number,
                                "documentID": document_id
                            }}
                        )
                        
                        # Audit event
                        await db.auditEvents.insert_one({
                            "id": str(uuid.uuid4()),
                            "actorId": current_user.email,
                            "action": "post_invoice",
                            "entity": "Invoice",
                            "entityId": invoice_id,
                            "metadata": {"externalNumber": external_number, "documentID": document_id},
                            "at": datetime.now(timezone.utc).isoformat()
                        })
                        
                        return {
                            "externalNumber": external_number, 
                            "documentID": document_id, 
                            "status": "posted",
                            "raw": result  # Include full response for debugging
                        }
                    else:
                        # API returned error in response body
                        error_details = result.get("response", {})
                        logger.error(f"e-računi API error in response: {error_details}")
                        raise HTTPException(
                            status_code=400, 
                            detail=f"e-računi API error: {error_details.get('error', error_details)}"
                        )
                else:
                    # Non-200 status code
                    response_body = response.text[:1000]
                    logger.error(f"e-računi API HTTP {response.status_code}: {response_body}")
                    raise HTTPException(
                        status_code=400, 
                        detail=f"e-računi API returned HTTP {response.status_code}. Response: {response_body}"
                    )
        
        except httpx.TimeoutException:
            raise HTTPException(status_code=400, detail="e-računi API timeout")
        except Exception as e:
            logger.error(f"e-računi posting error: {str(e)}")
            raise HTTPException(status_code=400, detail=f"Failed to post to e-računi: {str(e)}")

@api_router.get("/invoices")
async def list_invoices(current_user: User = Depends(get_current_user)):
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(1000)
    return invoices

# ============ AI ENDPOINTS ============
@api_router.get("/settings/ai")
async def get_ai_settings(current_user: User = Depends(get_current_user)):
    """Get AI settings for current user"""
    settings = await db.aiSettings.find_one({"userId": current_user.email})
    if not settings:
        # Return default settings
        default_settings = AISettings()
        return default_settings.model_dump()
    
    return {k: v for k, v in settings.items() if k != "_id" and k != "userId"}

@api_router.post("/settings/ai")
async def save_ai_settings(settings: AISettings, current_user: User = Depends(get_current_user)):
    """Save AI settings for current user"""
    settings_dict = settings.model_dump()
    settings_dict["userId"] = current_user.email
    settings_dict["updatedAt"] = datetime.now(timezone.utc).isoformat()
    
    await db.aiSettings.update_one(
        {"userId": current_user.email},
        {"$set": settings_dict},
        upsert=True
    )
    
    return {"message": "Settings saved successfully"}

@api_router.post("/settings/ai/test")
async def test_ai_connection(settings: AISettings, current_user: User = Depends(get_current_user)):
    """Test AI connection with provided settings"""
    # Get test prompt from settings (passed as extra field)
    test_prompt = getattr(settings, 'testPrompt', None) or "Hello, this is a connection test. Please respond with 'OK'."
    
    try:
        # LlmChat and UserMessage are now defined at top of file (litellm shim)
        # Resolve API key: custom/openai/anthropic use customApiKey, fallback to EMERGENT_LLM_KEY
        if settings.customApiKey and settings.aiProvider in ("custom", "openai", "anthropic"):
            api_key = settings.customApiKey
        elif EMERGENT_LLM_KEY:
            api_key = EMERGENT_LLM_KEY
        else:
            raise HTTPException(status_code=400, detail="API key required. Please configure AI settings first.")

        # Determine provider based on model name
        if "claude" in settings.customModel.lower():
            provider = "anthropic"
        elif "gemini" in settings.customModel.lower():
            provider = "google"
        else:
            provider = "openai"

        chat = LlmChat(
            api_key=api_key,
            session_id=f"test-{current_user.email}",
            system_message="You are a helpful AI assistant. IMPORTANT: Always respond in the SAME LANGUAGE as the user's prompt. If the prompt is in Slovenian, respond in Slovenian. If in English, respond in English. Match the language exactly."
        ).with_model(provider, settings.customModel)
        
        # Send test message
        message = UserMessage(text=test_prompt)
        response = await chat.send_message(message)
        
        return {
            "message": f"Connection successful! Model: {settings.customModel}",
            "response": response,
            "testPrompt": test_prompt
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")


# ============ E-RAČUNI ENDPOINTS ============
@api_router.post("/settings/eracuni/test")
async def test_eracuni_connection(
    username: str = Form(...),
    secretKey: str = Form(...),
    apiToken: str = Form(...),
    endpoint: str = Form(default="https://e-racuni.com/WebServices/API"),
    current_user: User = Depends(get_current_user)
):
    """Test e-računi API connection"""
    import httpx
    
    try:
        # Use configured endpoint or default
        api_url = endpoint or "https://e-racuni.com/WebServices/API"
        
        # Test with a simple PartnerList method (least intrusive)
        payload = {
            "username": username,
            "secretKey": secretKey,
            "token": apiToken,
            "method": "PartnerList",
            "parameters": {
                "page": 1,
                "limit": 1
            }
        }
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                api_url,
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            
            # Log the response for debugging
            logger.info(f"e-računi API test - Status: {response.status_code}, URL: {api_url}, Method: PartnerList")
            logger.info(f"Response body (first 200 chars): {response.text[:200]}")
            
            # Check if request was successful
            if response.status_code == 200:
                result = response.json()
                
                # Check if there's an error in the response
                if result.get("error"):
                    # Even if there's an error, connection was made
                    return {
                        "message": f"⚠️ API connection established but returned error: {result.get('error', 'Unknown')}. This may indicate invalid credentials or insufficient permissions.",
                        "fullResponse": result,
                        "statusCode": 200,
                        "warning": True
                    }
                
                return {
                    "message": "✅ e-računi connection successful! Credentials are valid and API responded correctly.",
                    "fullResponse": result,
                    "statusCode": 200
                }
            elif response.status_code == 404:
                # 404 might mean the method doesn't exist, but endpoint is reachable
                return {
                    "message": "⚠️ API endpoint is reachable but test method failed (404). This could mean: 1) Endpoint URL needs adjustment, 2) Test method 'PartnerList' not available, 3) Different API structure. For production use, configure the actual posting endpoint.",
                    "fullResponse": {"statusCode": 404, "note": "Endpoint reachable but test method not found", "suggestion": "Ready for production configuration"},
                    "statusCode": 404,
                    "warning": True
                }
            elif response.status_code == 401 or response.status_code == 403:
                # Authentication error
                return {
                    "message": f"❌ Authentication failed (HTTP {response.status_code}). Please verify your username, secret key, and token are correct.",
                    "fullResponse": {"statusCode": response.status_code, "body": response.text[:300]},
                    "statusCode": response.status_code
                }
            else:
                # Other errors
                error_body = response.text[:500] if response.text else "No response body"
                return {
                    "message": f"⚠️ Unexpected response (HTTP {response.status_code}). See debug details.",
                    "fullResponse": {"statusCode": response.status_code, "body": error_body},
                    "statusCode": response.status_code,
                    "warning": True
                }
    
    except httpx.TimeoutException:
        raise HTTPException(status_code=400, detail="Connection timeout - please check your network")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection test failed: {str(e)}")


@api_router.post("/ai/suggest")
async def ai_suggest(request: AIRequest, current_user: User = Depends(get_current_user)):
    # Get user's AI settings
    user_settings = await db.aiSettings.find_one({"userId": current_user.email})
    
    if not user_settings:
        user_settings = AISettings().model_dump()
    
    # Determine API key and model
    ai_provider = user_settings.get("aiProvider", "custom")
    if user_settings.get("customApiKey") and ai_provider in ("custom", "openai", "anthropic"):
        api_key = user_settings["customApiKey"]
        model = user_settings.get("customModel", "claude-3-5-sonnet-20241022")
    elif EMERGENT_LLM_KEY:
        api_key = EMERGENT_LLM_KEY
        model = "claude-3-5-sonnet-20241022"
    else:
        return {"suggestion": request.text, "message": "AI not configured. Please add an API key in Settings."}
    
    try:
        # Determine provider based on model
        if "claude" in model.lower():
            provider = "anthropic"
        elif "gemini" in model.lower():
            provider = "google"
        else:
            provider = "openai"
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"ai-{current_user.email}",
            system_message="You are an AI assistant for invoice processing."
        ).with_model(provider, model)
        
        # Get custom prompts
        prompts = {
            "grammar": user_settings.get("grammarPrompt", f"Fix grammar and spelling errors in this invoice text, return only the corrected text: {request.text}"),
            "fraud": user_settings.get("fraudPrompt", f"Analyze this invoice description for potential fraud indicators: {request.text}. Return a brief risk assessment."),
            "gdpr": user_settings.get("gdprPrompt", f"Identify and mask any personal data (names, emails, IDs) in this text: {request.text}. Return the masked version.")
        }
        
        # Format prompt with text
        base_prompt = prompts.get(request.feature, request.text)
        if "{text}" in base_prompt:
            prompt_text = base_prompt.replace("{text}", request.text)
        else:
            prompt_text = f"{base_prompt}\n\nText: {request.text}"
        
        message = UserMessage(text=prompt_text)
        response = await chat.send_message(message)
        
        return {"suggestion": response, "original": request.text}
    except Exception as e:
        return {"suggestion": request.text, "message": f"AI error: {str(e)}"}

# ============ ARTICLE CODES ENDPOINTS ============
@api_router.get("/articles")
async def get_articles(current_user: User = Depends(get_current_user)):
    """Get all article codes"""
    articles = await db.articles.find({}, {"_id": 0}).to_list(1000)
    return articles

@api_router.post("/articles")
async def create_article(article_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new article code"""
    code = article_data.get("code")
    
    if not code:
        raise HTTPException(status_code=400, detail="Article code is required")
    
    # Check if article already exists
    existing = await db.articles.find_one({"code": code})
    if existing:
        raise HTTPException(status_code=400, detail="Article code already exists")
    
    # Create article
    now = datetime.now(timezone.utc).isoformat()
    new_article = {
        "code": code,
        "description": article_data.get("description", ""),
        "unitMeasure": article_data.get("unitMeasure", "kos"),
        "priceWithoutVAT": float(article_data.get("priceWithoutVAT", 0)),
        "vatPercentage": float(article_data.get("vatPercentage", 22)),
        "tariffCode": article_data.get("tariffCode", ""),
        "created_at": now,
        "updated_at": now
    }
    
    await db.articles.insert_one(new_article)
    return {"message": "Article created successfully"}

@api_router.put("/articles/{article_code}")
async def update_article(article_code: str, article_data: dict, current_user: User = Depends(get_current_user)):
    """Update article code data"""
    # Only allow updating specific fields
    allowed_fields = ['description', 'unitMeasure', 'priceWithoutVAT', 'vatPercentage', 'tariffCode']
    update_data = {k: v for k, v in article_data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    result = await db.articles.update_one(
        {"code": article_code},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    
    return {"message": "Article updated successfully"}

# ============ EMPLOYEE COSTS ENDPOINTS ============
@api_router.get("/employee-costs")
async def get_employee_costs(archived: Optional[bool] = None, current_user: User = Depends(get_current_user)):
    """Get all employees with their cost data. Auto-extracts unique employee names from time entries and returns all manually created employees."""
    
    # Get unique employee names from time_entries collection
    employee_names_from_time_entries = await db.timeEntries.distinct("employeeName")
    
    # Auto-create employee entries for employees from time entries if they don't exist
    for name in employee_names_from_time_entries:
        if not name:  # Skip empty names
            continue
            
        # Check if employee exists in employee_costs collection
        existing = await db.employee_costs.find_one({"employee_name": name})
        
        if not existing:
            # Auto-create employee entry
            now = datetime.now(timezone.utc).isoformat()
            employee_data = {
                "employee_name": name,
                "cost": None,
                "archived": False,
                "created_at": now,
                "updated_at": now
            }
            await db.employee_costs.insert_one(employee_data.copy())
    
    # Now get ALL employees from employee_costs collection (including manually created ones)
    query = {}
    if archived is not None:
        query["archived"] = archived
    
    employees = await db.employee_costs.find(query, {"_id": 0}).to_list(1000)
    
    return employees

@api_router.post("/employee-costs")
async def update_employee_cost(employee_data: dict, current_user: User = Depends(get_current_user)):
    """Update employee cost"""
    employee_name = employee_data.get("employee_name")
    cost = employee_data.get("cost")
    
    if not employee_name:
        raise HTTPException(status_code=400, detail="Employee name is required")
    
    # Check if employee exists
    existing = await db.employee_costs.find_one({"employee_name": employee_name})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Update cost and updated_at timestamp
    now = datetime.now(timezone.utc).isoformat()
    result = await db.employee_costs.update_one(
        {"employee_name": employee_name},
        {"$set": {"cost": cost, "updated_at": now}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"message": "Employee cost updated successfully"}

@api_router.post("/employee-costs/create")
async def create_employee(employee_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new employee manually"""
    employee_name = employee_data.get("employee_name")
    
    if not employee_name or not employee_name.strip():
        raise HTTPException(status_code=400, detail="Employee name is required")
    
    # Check if employee already exists
    existing = await db.employee_costs.find_one({"employee_name": employee_name})
    if existing:
        raise HTTPException(status_code=400, detail="Employee already exists")
    
    # Create employee
    now = datetime.now(timezone.utc).isoformat()
    new_employee = {
        "employee_name": employee_name.strip(),
        "cost": employee_data.get("cost", 0),
        "archived": False,
        "created_at": now,
        "updated_at": now
    }
    
    await db.employee_costs.insert_one(new_employee)
    return {"message": "Employee created successfully"}

@api_router.put("/employee-costs/{employee_name}/archive")
async def archive_employee(employee_name: str, current_user: User = Depends(get_current_user)):
    """Archive an employee (soft delete)"""
    
    # Check if employee exists
    existing = await db.employee_costs.find_one({"employee_name": employee_name})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Archive employee
    now = datetime.now(timezone.utc).isoformat()
    result = await db.employee_costs.update_one(
        {"employee_name": employee_name},
        {"$set": {"archived": True, "updated_at": now}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"message": "Employee archived successfully"}

# ============ TARIFF CODES ENDPOINTS ============
@api_router.get("/tariffs")
async def get_tariffs(current_user: User = Depends(get_current_user)):
    """Get all tariff codes"""
    tariffs = await db.tariffs.find({}, {"_id": 0}).to_list(1000)
    return tariffs

@api_router.post("/tariffs")
async def create_tariff(tariff_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new tariff code"""
    code = tariff_data.get("code")
    description = tariff_data.get("description")
    
    if not code:
        raise HTTPException(status_code=400, detail="Tariff code is required")
    
    # Check if tariff already exists
    existing = await db.tariffs.find_one({"code": code})
    if existing:
        raise HTTPException(status_code=400, detail="Tariff code already exists")
    
    # Create tariff
    now = datetime.now(timezone.utc).isoformat()
    new_tariff = {
        "code": code,
        "description": description or "",
        "created_at": now,
        "updated_at": now
    }
    
    await db.tariffs.insert_one(new_tariff)
    return {"message": "Tariff created successfully"}

@api_router.put("/tariffs/{tariff_code}")
async def update_tariff(tariff_code: str, tariff_data: dict, current_user: User = Depends(get_current_user)):
    """Update tariff code data"""
    allowed_fields = ['description', 'value']
    update_data = {k: v for k, v in tariff_data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    now = datetime.now(timezone.utc).isoformat()
    update_data['updated_at'] = now
    
    result = await db.tariffs.update_one(
        {"code": tariff_code},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tariff not found")
    
    return {"message": "Tariff updated successfully"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@app.get("/health")
async def health_check():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8001))
    uvicorn.run("server:app", host="0.0.0.0", port=port)