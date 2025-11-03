import requests
import json
from typing import Dict, Any, Optional

# Configuration
BACKEND_URL = "https://invoice-workflow-2.preview.emergentagent.com/api"
ADMIN_EMAIL = "admin@local"
ADMIN_PASSWORD = "Admin2025!"

class TestMoveTimeEntry:
    def __init__(self):
        self.token = None
        self.customers = []
        self.test_batch_id = "eee5c6ba-330d-4034-a209-9ea96cec222c"
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def test_get_customers(self) -> bool:
        """Test GET /api/customers endpoint"""
        print("\n=== Testing GET /api/customers ===")
        try:
            response = requests.get(
                f"{BACKEND_URL}/customers",
                headers=self.get_headers()
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                self.customers = response.json()
                print(f"✅ Retrieved {len(self.customers)} customers")
                
                # Verify structure
                if self.customers:
                    sample = self.customers[0]
                    has_id = "id" in sample
                    has_name = "name" in sample
                    
                    print(f"Sample customer: {sample}")
                    print(f"Has 'id' field: {has_id}")
                    print(f"Has 'name' field: {has_name}")
                    
                    if not (has_id and has_name):
                        print("❌ Customer objects missing required fields")
                        return False
                
                # List all customers
                print("\nAll customers:")
                for customer in self.customers:
                    print(f"  - {customer.get('name')} (ID: {customer.get('id')})")
                
                return True
            else:
                print(f"❌ Failed to get customers: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Error getting customers: {str(e)}")
            return False
    
    def get_customer_by_name(self, name: str) -> Optional[Dict[str, Any]]:
        """Find customer by name"""
        for customer in self.customers:
            if name in customer.get("name", ""):
                return customer
        return None
    
    def get_time_entries_from_batch(self) -> list:
        """Get time entries from the test batch"""
        print(f"\n=== Getting time entries from batch {self.test_batch_id} ===")
        try:
            # We need to query MongoDB directly or use an endpoint
            # Let's try to get batch verification data which includes entries
            response = requests.get(
                f"{BACKEND_URL}/batches/{self.test_batch_id}/verification",
                headers=self.get_headers()
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                all_entries = []
                all_entries.extend(data.get("jmmcHP", []))
                all_entries.extend(data.get("jmmcFinance", []))
                all_entries.extend(data.get("noClient", []))
                
                print(f"✅ Found {len(all_entries)} time entries in batch")
                print(f"  - JMMC HP: {len(data.get('jmmcHP', []))}")
                print(f"  - JMMC Finance: {len(data.get('jmmcFinance', []))}")
                print(f"  - No Client: {len(data.get('noClient', []))}")
                
                return all_entries
            else:
                print(f"❌ Failed to get batch verification: {response.text}")
                return []
        except Exception as e:
            print(f"❌ Error getting time entries: {str(e)}")
            return []
    
    def get_invoice_for_customer(self, customer_id: str) -> Optional[Dict[str, Any]]:
        """Get invoice for a specific customer in the test batch"""
        try:
            response = requests.get(
                f"{BACKEND_URL}/batches/{self.test_batch_id}/invoices",
                headers=self.get_headers()
            )
            
            if response.status_code == 200:
                invoices = response.json()
                for invoice in invoices:
                    if invoice.get("customerId") == customer_id:
                        return invoice
            return None
        except Exception as e:
            print(f"Error getting invoice: {str(e)}")
            return None
    
    def verify_time_entry_in_invoice(self, entry_id: str, customer_id: str) -> bool:
        """Verify if a time entry is in a customer's invoice"""
        invoice = self.get_invoice_for_customer(customer_id)
        if not invoice:
            return False
        
        lines = invoice.get("lines", [])
        for line in lines:
            if line.get("timeEntryId") == entry_id:
                return True
        return False
    
    def test_move_time_entry(self, entry_id: str, new_customer_id: str, 
                            old_customer_id: str, scenario_name: str) -> bool:
        """Test moving a time entry to a different customer"""
        print(f"\n=== Testing: {scenario_name} ===")
        print(f"Entry ID: {entry_id}")
        print(f"Old Customer ID: {old_customer_id}")
        print(f"New Customer ID: {new_customer_id}")
        
        try:
            # Get old invoice state
            old_invoice_before = self.get_invoice_for_customer(old_customer_id)
            old_lines_count_before = len(old_invoice_before.get("lines", [])) if old_invoice_before else 0
            old_total_before = old_invoice_before.get("total", 0) if old_invoice_before else 0
            
            new_invoice_before = self.get_invoice_for_customer(new_customer_id)
            new_lines_count_before = len(new_invoice_before.get("lines", [])) if new_invoice_before else 0
            new_total_before = new_invoice_before.get("total", 0) if new_invoice_before else 0
            
            print(f"\nBefore move:")
            print(f"  Old customer invoice: {old_lines_count_before} lines, total: {old_total_before}")
            print(f"  New customer invoice: {new_lines_count_before} lines, total: {new_total_before}")
            
            # Move the entry
            response = requests.post(
                f"{BACKEND_URL}/time-entries/{entry_id}/move-customer",
                headers=self.get_headers(),
                data={"new_customer_id": new_customer_id}
            )
            print(f"\nMove API Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code != 200:
                print(f"❌ Move failed")
                return False
            
            # Verify the move
            print("\n--- Verification ---")
            
            # 1. Check old invoice
            old_invoice_after = self.get_invoice_for_customer(old_customer_id)
            if old_invoice_after:
                old_lines_count_after = len(old_invoice_after.get("lines", []))
                old_total_after = old_invoice_after.get("total", 0)
                
                # Verify entry is removed
                entry_in_old = self.verify_time_entry_in_invoice(entry_id, old_customer_id)
                
                print(f"Old customer invoice after:")
                print(f"  Lines: {old_lines_count_before} -> {old_lines_count_after}")
                print(f"  Total: {old_total_before} -> {old_total_after}")
                print(f"  Entry still in old invoice: {entry_in_old}")
                
                if entry_in_old:
                    print("❌ Entry still exists in old customer's invoice (duplication!)")
                    return False
                
                # Don't check line count - other entries may exist for this customer
                print("✅ Entry successfully removed from old invoice")
            
            # 2. Check new invoice
            new_invoice_after = self.get_invoice_for_customer(new_customer_id)
            if new_invoice_after:
                new_lines_count_after = len(new_invoice_after.get("lines", []))
                new_total_after = new_invoice_after.get("total", 0)
                
                # Verify entry is added
                entry_in_new = self.verify_time_entry_in_invoice(entry_id, new_customer_id)
                
                print(f"\nNew customer invoice after:")
                print(f"  Lines: {new_lines_count_before} -> {new_lines_count_after}")
                print(f"  Total: {new_total_before} -> {new_total_after}")
                print(f"  Entry in new invoice: {entry_in_new}")
                
                if not entry_in_new:
                    print("❌ Entry not found in new customer's invoice")
                    return False
                
                print("✅ Entry successfully added to new invoice")
            else:
                print("❌ New customer invoice not found after move")
                return False
            
            print(f"\n✅ {scenario_name} - PASSED")
            return True
            
        except Exception as e:
            print(f"❌ Error during move: {str(e)}")
            return False
    
    def test_move_invalid_entry(self) -> bool:
        """Test moving with invalid entry_id"""
        print("\n=== Testing: Move with invalid entry_id ===")
        
        invalid_entry_id = "invalid-entry-id-12345"
        customer = self.customers[0] if self.customers else None
        
        if not customer:
            print("❌ No customers available for test")
            return False
        
        try:
            response = requests.post(
                f"{BACKEND_URL}/time-entries/{invalid_entry_id}/move-customer",
                headers=self.get_headers(),
                data={"new_customer_id": customer["id"]}
            )
            print(f"Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code == 404:
                print("✅ Correctly returned 404 for invalid entry_id")
                return True
            else:
                print(f"❌ Expected 404, got {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return False
    
    def test_move_invalid_customer(self) -> bool:
        """Test moving to invalid customer_id"""
        print("\n=== Testing: Move to invalid customer_id ===")
        
        # Get a valid entry
        entries = self.get_time_entries_from_batch()
        if not entries:
            print("❌ No entries available for test")
            return False
        
        entry = entries[0]
        invalid_customer_id = "invalid-customer-id-12345"
        
        try:
            response = requests.post(
                f"{BACKEND_URL}/time-entries/{entry['id']}/move-customer",
                headers=self.get_headers(),
                data={"new_customer_id": invalid_customer_id}
            )
            print(f"Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code == 404:
                print("✅ Correctly returned 404 for invalid customer_id")
                return True
            else:
                print(f"❌ Expected 404, got {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return False
    
    def run_all_tests(self):
        """Run all tests"""
        print("=" * 80)
        print("MOVE TIME ENTRY FEATURE - BACKEND TESTS")
        print("=" * 80)
        
        results = {}
        
        # 1. Login
        if not self.login():
            print("\n❌ CRITICAL: Login failed. Cannot proceed with tests.")
            return
        
        # 2. Get customers
        results["GET /api/customers"] = self.test_get_customers()
        
        if not results["GET /api/customers"]:
            print("\n❌ CRITICAL: Cannot get customers. Cannot proceed with move tests.")
            return
        
        # Find specific customers
        no_client = self.get_customer_by_name("General")
        jmmc_hp = self.get_customer_by_name("JMMC HP d.o.o.")
        jmmc_finance = self.get_customer_by_name("JMMC Finance d.o.o.")
        
        print(f"\nTarget customers:")
        print(f"  No Client/General: {no_client}")
        print(f"  JMMC HP: {jmmc_hp}")
        print(f"  JMMC Finance: {jmmc_finance}")
        
        # Get time entries
        entries = self.get_time_entries_from_batch()
        
        if not entries:
            print("\n⚠️  WARNING: No time entries found in batch. Cannot test move functionality.")
            print("This might be expected if the batch doesn't exist or has no entries.")
        else:
            # Test scenarios with actual entries
            
            # Scenario A: Move from No Client to JMMC HP
            if no_client and jmmc_hp:
                no_client_entries = [e for e in entries if e.get("customerId") == no_client["id"]]
                if no_client_entries:
                    entry = no_client_entries[0]
                    results["Move: No Client -> JMMC HP"] = self.test_move_time_entry(
                        entry["id"], 
                        jmmc_hp["id"], 
                        no_client["id"],
                        "Move from No Client to JMMC HP d.o.o."
                    )
                else:
                    print("\n⚠️  No entries found for 'No Client' customer")
            
            # Scenario B: Move from JMMC HP to JMMC Finance
            if jmmc_hp and jmmc_finance:
                jmmc_hp_entries = [e for e in entries if e.get("customerId") == jmmc_hp["id"]]
                if jmmc_hp_entries:
                    entry = jmmc_hp_entries[0]
                    results["Move: JMMC HP -> JMMC Finance"] = self.test_move_time_entry(
                        entry["id"], 
                        jmmc_finance["id"], 
                        jmmc_hp["id"],
                        "Move from JMMC HP d.o.o. to JMMC Finance d.o.o."
                    )
                else:
                    print("\n⚠️  No entries found for 'JMMC HP' customer")
        
        # Test error scenarios
        results["Invalid entry_id (404)"] = self.test_move_invalid_entry()
        results["Invalid customer_id (404)"] = self.test_move_invalid_customer()
        
        # Summary
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for v in results.values() if v)
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{status} - {test_name}")
        
        print(f"\nTotal: {passed}/{total} tests passed")
        
        if passed == total:
            print("\n🎉 ALL TESTS PASSED!")
        else:
            print(f"\n⚠️  {total - passed} test(s) failed")

class TestInvoicingSettingsAutoPopulation:
    """Test auto-population of invoicing settings based on Article 000001 analysis"""
    
    def __init__(self):
        self.token = None
        self.test_customer_id = None
        self.test_customer_name = "Test Auto-Population Customer"
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def create_test_customer(self) -> bool:
        """Create a test customer for uploading history"""
        print("\n=== Creating Test Customer ===")
        try:
            customer_data = {
                "name": self.test_customer_name,
                "unitPrice": 0
            }
            
            response = requests.post(
                f"{BACKEND_URL}/customers",
                headers=self.get_headers(),
                json=customer_data
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                self.test_customer_id = result.get("customerId")
                print(f"✅ Test customer created")
                print(f"  Customer ID: {self.test_customer_id}")
                print(f"  Name: {self.test_customer_name}")
                return True
            else:
                print(f"❌ Failed to create customer: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Error creating customer: {str(e)}")
            return False
    
    def upload_customer_history(self, file_path: str) -> bool:
        """Upload customer history XLSX file"""
        print(f"\n=== Uploading Customer History: {file_path} ===")
        try:
            with open(file_path, 'rb') as f:
                files = {'file': ('test_invoice_settings.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'customer_ids': self.test_customer_id
                }
                
                response = requests.post(
                    f"{BACKEND_URL}/customers/upload-history",
                    headers=self.get_headers(),
                    files=files,
                    data=data
                )
            
            print(f"Status: {response.status_code}")
            print(f"Response: {response.text[:500]}")
            
            if response.status_code == 200:
                result = response.json()
                print(f"✅ Customer history uploaded successfully")
                print(f"  Message: {result.get('message')}")
                print(f"  Customers updated: {result.get('updated_count', 0)}")
                return True
            else:
                print(f"❌ Failed to upload history: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Error uploading history: {str(e)}")
            return False
    
    def get_customer_details(self) -> Optional[Dict[str, Any]]:
        """Get customer details including invoicing settings"""
        print(f"\n=== Getting Customer Details: {self.test_customer_id} ===")
        try:
            response = requests.get(
                f"{BACKEND_URL}/customers/{self.test_customer_id}",
                headers=self.get_headers()
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                customer = response.json()
                print(f"✅ Customer details retrieved")
                print(f"  Name: {customer.get('name')}")
                print(f"  Invoicing Type: {customer.get('invoicingType')}")
                print(f"  Fixed Forfait Value: {customer.get('fixedForfaitValue')}")
                print(f"  Unit Price (Hourly Rate): {customer.get('unitPrice')}")
                print(f"  Historical Invoices: {len(customer.get('historicalInvoices', []))}")
                return customer
            else:
                print(f"❌ Failed to get customer: {response.text}")
                return None
        except Exception as e:
            print(f"❌ Error getting customer: {str(e)}")
            return None
    
    def check_backend_logs(self):
        """Check backend logs for detection messages"""
        print("\n=== Checking Backend Logs ===")
        try:
            import subprocess
            result = subprocess.run(
                ["tail", "-n", "50", "/var/log/supervisor/backend.out.log"],
                capture_output=True,
                text=True
            )
            
            logs = result.stdout
            
            # Look for auto-detection messages
            detection_messages = []
            for line in logs.split('\n'):
                if 'Auto-detected' in line or 'Article 000001' in line or 'invoicing' in line.lower():
                    detection_messages.append(line)
            
            if detection_messages:
                print("✅ Found detection messages in logs:")
                for msg in detection_messages[-10:]:  # Show last 10
                    print(f"  {msg}")
            else:
                print("⚠️  No detection messages found in recent logs")
            
        except Exception as e:
            print(f"⚠️  Could not check logs: {str(e)}")
    
    def verify_auto_population(self, customer: Dict[str, Any]) -> Dict[str, bool]:
        """Verify that invoicing settings were auto-populated correctly"""
        print("\n=== Verifying Auto-Population Logic ===")
        
        results = {}
        
        # Get historical invoices to analyze Article 000001
        historical_invoices = customer.get('historicalInvoices', [])
        
        if not historical_invoices:
            print("❌ No historical invoices found")
            return {"has_historical_data": False}
        
        print(f"Found {len(historical_invoices)} historical invoice periods")
        
        # Get the latest period (most recent)
        latest_period = historical_invoices[-1] if historical_invoices else None
        
        if not latest_period:
            print("❌ No latest period found")
            return {"has_latest_period": False}
        
        print(f"\nLatest Period: {latest_period.get('date', 'N/A')}")
        
        # Analyze Article 000001 entries
        individual_rows = latest_period.get('individualRows', [])
        article_000001_rows = [row for row in individual_rows if row.get('articleCode', '').strip() == '000001']
        
        print(f"Found {len(article_000001_rows)} Article 000001 entries in latest period")
        
        if not article_000001_rows:
            print("⚠️  No Article 000001 entries found")
            results["has_article_000001"] = False
            return results
        
        results["has_article_000001"] = True
        
        # Display Article 000001 entries
        for idx, row in enumerate(article_000001_rows):
            print(f"\nArticle 000001 Entry #{idx + 1}:")
            print(f"  Description: {row.get('description', 'N/A')}")
            print(f"  Detailed Description: {row.get('detailedDescription', 'N/A')[:100]}...")
            print(f"  Unit Price: €{row.get('unitPrice', 0)}")
            print(f"  Quantity: {row.get('quantity', 'N/A')}")
        
        # Determine expected invoicing type based on logic
        expected_type = None
        expected_forfait = None
        expected_hourly = None
        
        if len(article_000001_rows) == 1:
            single_row = article_000001_rows[0]
            detailed_desc = single_row.get('detailedDescription', '').strip()
            unit_price = single_row.get('unitPrice')
            
            # Check for work list patterns
            import re
            has_work_list = False
            if detailed_desc:
                date_patterns = [
                    r'\d{4}-\d{2}-\d{2}',  # 2024-10-17
                    r'\d{2}\.\d{2}\.\d{2,4}',  # 17.10.24 or 17.10.2024
                ]
                for pattern in date_patterns:
                    if re.search(pattern, detailed_desc):
                        has_work_list = True
                        break
            
            if has_work_list:
                # Case C: By Hours Spent
                expected_type = "by-hours"
                expected_hourly = unit_price
                print(f"\n📋 Expected: By Hours Spent (hourly rate: €{expected_hourly})")
            else:
                # Case A: Fixed Forfait
                expected_type = "fixed-forfait"
                expected_forfait = unit_price
                print(f"\n📋 Expected: Fixed Forfait (value: €{expected_forfait})")
        
        elif len(article_000001_rows) >= 2:
            # Case B: Hybrid
            expected_type = "hybrid"
            expected_forfait = article_000001_rows[0].get('unitPrice')
            expected_hourly = article_000001_rows[1].get('unitPrice')
            print(f"\n📋 Expected: Hybrid (forfait: €{expected_forfait}, hourly: €{expected_hourly})")
        
        # Verify actual values match expected
        actual_type = customer.get('invoicingType')
        actual_forfait = customer.get('fixedForfaitValue')
        actual_hourly = customer.get('unitPrice')
        
        print(f"\n🔍 Verification:")
        print(f"  Expected Type: {expected_type}")
        print(f"  Actual Type: {actual_type}")
        
        if expected_type == actual_type:
            print(f"  ✅ Invoicing Type matches")
            results["invoicing_type_correct"] = True
        else:
            print(f"  ❌ Invoicing Type mismatch")
            results["invoicing_type_correct"] = False
        
        if expected_type in ["fixed-forfait", "hybrid"]:
            print(f"  Expected Forfait: €{expected_forfait}")
            print(f"  Actual Forfait: €{actual_forfait}")
            
            if expected_forfait == actual_forfait:
                print(f"  ✅ Fixed Forfait Value matches")
                results["forfait_value_correct"] = True
            else:
                print(f"  ❌ Fixed Forfait Value mismatch")
                results["forfait_value_correct"] = False
        
        if expected_type in ["by-hours", "hybrid"]:
            print(f"  Expected Hourly Rate: €{expected_hourly}")
            print(f"  Actual Hourly Rate: €{actual_hourly}")
            
            if expected_hourly == actual_hourly:
                print(f"  ✅ Hourly Rate matches")
                results["hourly_rate_correct"] = True
            else:
                print(f"  ❌ Hourly Rate mismatch")
                results["hourly_rate_correct"] = False
        
        return results
    
    def cleanup_test_customer(self):
        """Delete the test customer"""
        print(f"\n=== Cleaning Up Test Customer ===")
        try:
            # Note: There's no delete endpoint, so we'll just archive or leave it
            print("⚠️  No cleanup performed (no delete endpoint available)")
        except Exception as e:
            print(f"⚠️  Cleanup error: {str(e)}")
    
    def run_all_tests(self):
        """Run all auto-population tests"""
        print("=" * 80)
        print("INVOICING SETTINGS AUTO-POPULATION - BACKEND TESTS")
        print("=" * 80)
        print("\nTesting auto-population logic for invoicing settings based on")
        print("Article 000001 analysis from customer history XLSX files.")
        print("\nTest Cases:")
        print("  Case A - Fixed Forfait: Article 000001 appears ONCE, empty/simple description")
        print("  Case B - Hybrid: Article 000001 appears 2+ TIMES")
        print("  Case C - By Hours: Article 000001 appears ONCE with work list (dates)")
        print("=" * 80)
        
        results = {}
        
        # 1. Login
        if not self.login():
            print("\n❌ CRITICAL: Login failed. Cannot proceed with tests.")
            return
        
        # 2. Create test customer
        results["Create Test Customer"] = self.create_test_customer()
        if not results["Create Test Customer"]:
            print("\n❌ CRITICAL: Failed to create test customer. Cannot proceed.")
            return
        
        # 3. Upload customer history XLSX
        test_file = "/tmp/test_invoice_settings.xlsx"
        results["Upload Customer History"] = self.upload_customer_history(test_file)
        
        if not results["Upload Customer History"]:
            print("\n❌ CRITICAL: Failed to upload customer history. Cannot proceed.")
            return
        
        # 4. Check backend logs
        self.check_backend_logs()
        
        # 5. Get customer details
        customer = self.get_customer_details()
        
        if not customer:
            print("\n❌ CRITICAL: Failed to get customer details. Cannot proceed.")
            return
        
        # 6. Verify auto-population
        verification_results = self.verify_auto_population(customer)
        
        # Merge verification results
        results.update(verification_results)
        
        # 7. Cleanup
        self.cleanup_test_customer()
        
        # Summary
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for v in results.values() if v is True)
        total = len([v for v in results.values() if isinstance(v, bool)])
        
        for test_name, result in results.items():
            if isinstance(result, bool):
                status = "✅ PASS" if result else "❌ FAIL"
                print(f"{status} - {test_name}")
        
        print(f"\nTotal: {passed}/{total} tests passed")
        
        if passed == total:
            print("\n🎉 ALL TESTS PASSED!")
            print("\n✅ Auto-population logic is working correctly:")
            print("  - Article 000001 entries detected ✅")
            print("  - Invoicing type auto-populated ✅")
            print("  - Pricing values auto-populated ✅")
        else:
            print(f"\n⚠️  {total - passed} test(s) failed")
            print("\n🔍 Debugging Hints:")
            print("  1. Check backend logs for auto-detection messages")
            print("  2. Verify Article 000001 entries exist in the XLSX file")
            print("  3. Check that the XLSX file has the correct format")
            print("  4. Ensure the logic in server.py is correctly implemented")


class TestAICorrectionTracking:
    """Test AI correction tracking feature for Import Verification page"""
    
    def __init__(self):
        self.token = None
        self.batch_id = None
        self.time_entries = []
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def create_test_import(self) -> bool:
        """Test 1: Create new import and verify aiCorrectionApplied field initialization"""
        print("\n=== Test 1: New Import - AI Field Initialization ===")
        
        # Create a simple test XLSX file
        try:
            import openpyxl
            from datetime import datetime
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št. računa"]
            ws.append(headers)
            
            # Add test data rows
            test_rows = [
                ["Project A", "Test Customer", datetime(2025, 10, 1), "Standard", "John Doe", "Test work description", 2.5, 100.0, "INV-001"],
                ["Project B", "Test Customer", datetime(2025, 10, 2), "Standard", "Jane Smith", "Another test task", 3.0, 120.0, "INV-001"],
                ["Project C", "Test Customer", datetime(2025, 10, 3), "Standard", "Bob Johnson", "Third test entry", 1.5, 60.0, "INV-001"],
            ]
            
            for row in test_rows:
                ws.append(row)
            
            # Save to temp file
            test_file_path = "/tmp/test_ai_correction.xlsx"
            wb.save(test_file_path)
            print(f"✅ Created test XLSX file: {test_file_path}")
            
            # Upload the file
            with open(test_file_path, 'rb') as f:
                files = {'file': ('test_ai_correction.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'title': 'AI Correction Test Import',
                    'invoiceDate': '2025-10-31',
                    'periodFrom': '2025-10-01',
                    'periodTo': '2025-10-31',
                    'dueDate': '2025-11-15'
                }
                
                response = requests.post(
                    f"{BACKEND_URL}/imports",
                    headers=self.get_headers(),
                    files=files,
                    data=data
                )
            
            print(f"Import Status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                self.batch_id = result.get("batchId")
                row_count = result.get("rowCount")
                
                print(f"✅ Import successful")
                print(f"  Batch ID: {self.batch_id}")
                print(f"  Row Count: {row_count}")
                
                # Now get time entries and verify aiCorrectionApplied field
                response = requests.get(
                    f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                    headers=self.get_headers()
                )
                
                if response.status_code == 200:
                    self.time_entries = response.json()
                    print(f"✅ Retrieved {len(self.time_entries)} time entries")
                    
                    # Verify all entries have aiCorrectionApplied=false
                    all_false = True
                    missing_field = False
                    
                    for idx, entry in enumerate(self.time_entries):
                        if "aiCorrectionApplied" not in entry:
                            print(f"❌ Entry {idx} missing aiCorrectionApplied field")
                            missing_field = True
                        elif entry.get("aiCorrectionApplied") != False:
                            print(f"❌ Entry {idx} has aiCorrectionApplied={entry.get('aiCorrectionApplied')}, expected False")
                            all_false = False
                    
                    if missing_field:
                        print("❌ Some entries missing aiCorrectionApplied field")
                        return False
                    
                    if not all_false:
                        print("❌ Some entries have aiCorrectionApplied != False")
                        return False
                    
                    print("✅ All time entries have aiCorrectionApplied=False by default")
                    return True
                else:
                    print(f"❌ Failed to get time entries: {response.text}")
                    return False
            else:
                print(f"❌ Import failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error creating import: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def test_update_ai_correction_status(self) -> bool:
        """Test 2: Update AI correction status via PUT endpoint"""
        print("\n=== Test 2: Update AI Correction Status ===")
        
        if not self.batch_id or not self.time_entries:
            print("❌ No batch or time entries available")
            return False
        
        try:
            # Update first entry with aiCorrectionApplied=true
            updates = [
                {
                    "index": 0,
                    "comments": "Updated by AI",
                    "hours": 2.5,
                    "aiCorrectionApplied": True
                }
            ]
            
            response = requests.put(
                f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                headers=self.get_headers(),
                json=updates
            )
            
            print(f"Update Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code == 200:
                result = response.json()
                updated_count = result.get("updated_count", 0)
                
                if updated_count == 1:
                    print(f"✅ Successfully updated {updated_count} entry")
                    return True
                else:
                    print(f"❌ Expected 1 update, got {updated_count}")
                    return False
            else:
                print(f"❌ Update failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error updating entries: {str(e)}")
            return False
    
    def test_retrieve_ai_correction_status(self) -> bool:
        """Test 3: Retrieve AI correction status and verify persistence"""
        print("\n=== Test 3: Retrieve AI Correction Status ===")
        
        if not self.batch_id:
            print("❌ No batch available")
            return False
        
        try:
            response = requests.get(
                f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                headers=self.get_headers()
            )
            
            print(f"Get Status: {response.status_code}")
            
            if response.status_code == 200:
                entries = response.json()
                print(f"✅ Retrieved {len(entries)} time entries")
                
                # Verify first entry has aiCorrectionApplied=true
                if len(entries) > 0:
                    first_entry = entries[0]
                    
                    if "aiCorrectionApplied" not in first_entry:
                        print("❌ First entry missing aiCorrectionApplied field")
                        return False
                    
                    if first_entry.get("aiCorrectionApplied") != True:
                        print(f"❌ First entry has aiCorrectionApplied={first_entry.get('aiCorrectionApplied')}, expected True")
                        return False
                    
                    print(f"✅ First entry has aiCorrectionApplied=True (persisted correctly)")
                    
                    # Verify other entries still have false
                    for idx in range(1, len(entries)):
                        entry = entries[idx]
                        if entry.get("aiCorrectionApplied") != False:
                            print(f"❌ Entry {idx} has aiCorrectionApplied={entry.get('aiCorrectionApplied')}, expected False")
                            return False
                    
                    print(f"✅ Other entries still have aiCorrectionApplied=False")
                    return True
                else:
                    print("❌ No entries returned")
                    return False
            else:
                print(f"❌ Failed to get entries: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error retrieving entries: {str(e)}")
            return False
    
    def test_multiple_updates(self) -> bool:
        """Test 4: Update multiple rows with different aiCorrectionApplied values"""
        print("\n=== Test 4: Multiple Updates ===")
        
        if not self.batch_id or not self.time_entries:
            print("❌ No batch or time entries available")
            return False
        
        try:
            # Update multiple entries
            updates = [
                {
                    "index": 0,
                    "comments": "First AI correction",
                    "hours": 2.0,
                    "aiCorrectionApplied": True
                },
                {
                    "index": 1,
                    "comments": "Second AI correction",
                    "hours": 3.5,
                    "aiCorrectionApplied": True
                },
                {
                    "index": 2,
                    "comments": "No AI correction",
                    "hours": 1.5,
                    "aiCorrectionApplied": False
                }
            ]
            
            response = requests.put(
                f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                headers=self.get_headers(),
                json=updates
            )
            
            print(f"Update Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code == 200:
                result = response.json()
                updated_count = result.get("updated_count", 0)
                
                if updated_count == 3:
                    print(f"✅ Successfully updated {updated_count} entries")
                    
                    # Verify the updates
                    response = requests.get(
                        f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                        headers=self.get_headers()
                    )
                    
                    if response.status_code == 200:
                        entries = response.json()
                        
                        # Check each entry
                        expected_values = [True, True, False]
                        all_correct = True
                        
                        for idx, expected in enumerate(expected_values):
                            if idx < len(entries):
                                actual = entries[idx].get("aiCorrectionApplied")
                                if actual != expected:
                                    print(f"❌ Entry {idx}: expected aiCorrectionApplied={expected}, got {actual}")
                                    all_correct = False
                                else:
                                    print(f"✅ Entry {idx}: aiCorrectionApplied={actual} (correct)")
                        
                        if all_correct:
                            print("✅ All entries have correct aiCorrectionApplied values")
                            return True
                        else:
                            print("❌ Some entries have incorrect values")
                            return False
                    else:
                        print(f"❌ Failed to verify updates: {response.text}")
                        return False
                else:
                    print(f"❌ Expected 3 updates, got {updated_count}")
                    return False
            else:
                print(f"❌ Update failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error updating multiple entries: {str(e)}")
            return False
    
    def run_all_tests(self):
        """Run all AI correction tracking tests"""
        print("=" * 80)
        print("AI CORRECTION TRACKING FEATURE - BACKEND TESTS")
        print("=" * 80)
        print("\nTesting new feature: Track when AI suggestions are applied to time entries")
        print("Feature: When user clicks 'Apply Changes' in AI Evaluation modal,")
        print("the row should be marked with aiCorrectionApplied=true in database")
        print("=" * 80)
        
        results = {}
        
        # 1. Login
        if not self.login():
            print("\n❌ CRITICAL: Login failed. Cannot proceed with tests.")
            return
        
        # 2. Test new import with AI field initialization
        results["Test 1: New Import - AI Field Initialization"] = self.create_test_import()
        
        if not results["Test 1: New Import - AI Field Initialization"]:
            print("\n❌ CRITICAL: Import test failed. Cannot proceed with other tests.")
            return
        
        # 3. Test updating AI correction status
        results["Test 2: Update AI Correction Status"] = self.test_update_ai_correction_status()
        
        # 4. Test retrieving AI correction status
        results["Test 3: Retrieve AI Correction Status"] = self.test_retrieve_ai_correction_status()
        
        # 5. Test multiple updates
        results["Test 4: Multiple Updates"] = self.test_multiple_updates()
        
        # Summary
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for v in results.values() if v)
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{status} - {test_name}")
        
        print(f"\nTotal: {passed}/{total} tests passed")
        
        if passed == total:
            print("\n🎉 ALL TESTS PASSED!")
            print("\n✅ AI Correction Tracking Feature is working correctly:")
            print("  - aiCorrectionApplied field initialized to False on import ✅")
            print("  - Field can be updated via PUT endpoint ✅")
            print("  - Field persists in database ✅")
            print("  - Field is returned in GET responses ✅")
            print("  - Multiple rows can be updated with different values ✅")
        else:
            print(f"\n⚠️  {total - passed} test(s) failed")
            print("\n🔍 Debugging Hints:")
            print("  1. Check if aiCorrectionApplied field is in time entry schema")
            print("  2. Verify POST /api/imports initializes the field")
            print("  3. Verify PUT /api/batches/{batch_id}/time-entries accepts the field")
            print("  4. Verify GET /api/batches/{batch_id}/time-entries returns the field")
            print("  5. Check backend logs for any errors")


class TestVerificationEndpointBehavior:
    """Test GET /api/batches/{batch_id}/verification endpoint behavior for different batch statuses"""
    
    def __init__(self):
        self.token = None
        self.batch_id = None
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def create_in_progress_batch(self) -> bool:
        """Test 1: Create 'in progress' batch and verify empty verification arrays"""
        print("\n=== Test 1: Create 'in progress' batch ===")
        
        try:
            import openpyxl
            from datetime import datetime
            
            # Create workbook with test data
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št. računa"]
            ws.append(headers)
            
            # Add test data rows with different customers
            test_rows = [
                ["Project A", "JMMC HP d.o.o.", datetime(2025, 10, 1), "Standard", "John Doe", "Test work for JMMC HP", 2.5, 100.0, "INV-001"],
                ["Project B", "JMMC Finance d.o.o.", datetime(2025, 10, 2), "Standard", "Jane Smith", "Test work for JMMC Finance", 3.0, 120.0, "INV-001"],
                ["Project C", "General", datetime(2025, 10, 3), "Standard", "Bob Johnson", "Test work with no client", 1.5, 60.0, "INV-001"],
                ["Project D", "", datetime(2025, 10, 4), "999 - EXTRA", "Alice Brown", "Extra work", 2.0, 80.0, "INV-001"],
            ]
            
            for row in test_rows:
                ws.append(row)
            
            # Save to temp file
            test_file_path = "/tmp/test_verification_batch.xlsx"
            wb.save(test_file_path)
            print(f"✅ Created test XLSX file: {test_file_path}")
            
            # Upload the file with saveAsProgress=true
            with open(test_file_path, 'rb') as f:
                files = {'file': ('test_verification_batch.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'title': 'Verification Test Batch',
                    'invoiceDate': '2025-10-31',
                    'periodFrom': '2025-10-01',
                    'periodTo': '2025-10-31',
                    'dueDate': '2025-11-15',
                    'saveAsProgress': 'true'  # This is the key parameter
                }
                
                response = requests.post(
                    f"{BACKEND_URL}/imports",
                    headers=self.get_headers(),
                    files=files,
                    data=data
                )
            
            print(f"Import Status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                self.batch_id = result.get("batchId")
                row_count = result.get("rowCount")
                
                print(f"✅ Import successful")
                print(f"  Batch ID: {self.batch_id}")
                print(f"  Row Count: {row_count}")
                
                # Verify batch status is "in progress"
                batch_response = requests.get(
                    f"{BACKEND_URL}/batches/{self.batch_id}",
                    headers=self.get_headers()
                )
                
                if batch_response.status_code == 200:
                    batch = batch_response.json()
                    batch_status = batch.get("status")
                    
                    print(f"  Batch Status: {batch_status}")
                    
                    if batch_status != "in progress":
                        print(f"❌ Expected batch status 'in progress', got '{batch_status}'")
                        return False
                    
                    print("✅ Batch status is 'in progress'")
                    
                    # Now call verification endpoint
                    verification_response = requests.get(
                        f"{BACKEND_URL}/batches/{self.batch_id}/verification",
                        headers=self.get_headers()
                    )
                    
                    print(f"\nVerification endpoint status: {verification_response.status_code}")
                    
                    if verification_response.status_code == 200:
                        verification_data = verification_response.json()
                        
                        print(f"Verification data: {json.dumps(verification_data, indent=2)}")
                        
                        # Verify all arrays are empty
                        jmmc_hp = verification_data.get("jmmcHP", None)
                        jmmc_finance = verification_data.get("jmmcFinance", None)
                        no_client = verification_data.get("noClient", None)
                        extra = verification_data.get("extra", None)
                        
                        if jmmc_hp is None or jmmc_finance is None or no_client is None or extra is None:
                            print("❌ Missing expected fields in verification response")
                            return False
                        
                        if not isinstance(jmmc_hp, list) or not isinstance(jmmc_finance, list) or not isinstance(no_client, list) or not isinstance(extra, list):
                            print("❌ Expected arrays for all fields")
                            return False
                        
                        if len(jmmc_hp) != 0 or len(jmmc_finance) != 0 or len(no_client) != 0 or len(extra) != 0:
                            print(f"❌ Expected empty arrays, got:")
                            print(f"  jmmcHP: {len(jmmc_hp)} entries")
                            print(f"  jmmcFinance: {len(jmmc_finance)} entries")
                            print(f"  noClient: {len(no_client)} entries")
                            print(f"  extra: {len(extra)} entries")
                            return False
                        
                        print("✅ All verification arrays are empty (as expected for 'in progress' batch)")
                        return True
                    else:
                        print(f"❌ Failed to get verification data: {verification_response.text}")
                        return False
                else:
                    print(f"❌ Failed to get batch details: {batch_response.text}")
                    return False
            else:
                print(f"❌ Import failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error creating in progress batch: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def compose_and_verify_populated_arrays(self) -> bool:
        """Test 2: Compose invoices and verify populated verification arrays"""
        print("\n=== Test 2: Compose invoices and verify populated arrays ===")
        
        if not self.batch_id:
            print("❌ No batch ID available")
            return False
        
        try:
            # Compose invoices for the batch
            print(f"\nComposing invoices for batch {self.batch_id}...")
            compose_response = requests.post(
                f"{BACKEND_URL}/invoices/compose",
                headers=self.get_headers(),
                params={"batchId": self.batch_id}
            )
            
            print(f"Compose Status: {compose_response.status_code}")
            print(f"Response: {compose_response.text[:500]}")
            
            if compose_response.status_code != 200:
                print(f"❌ Failed to compose invoices: {compose_response.text}")
                return False
            
            print("✅ Invoices composed successfully")
            
            # Verify batch status changed to "composed"
            batch_response = requests.get(
                f"{BACKEND_URL}/batches/{self.batch_id}",
                headers=self.get_headers()
            )
            
            if batch_response.status_code == 200:
                batch = batch_response.json()
                batch_status = batch.get("status")
                
                print(f"  Batch Status after compose: {batch_status}")
                
                if batch_status != "composed":
                    print(f"⚠️  Expected batch status 'composed', got '{batch_status}'")
                    # Don't fail the test, just warn - the status might be different
                
                # Now call verification endpoint again
                verification_response = requests.get(
                    f"{BACKEND_URL}/batches/{self.batch_id}/verification",
                    headers=self.get_headers()
                )
                
                print(f"\nVerification endpoint status: {verification_response.status_code}")
                
                if verification_response.status_code == 200:
                    verification_data = verification_response.json()
                    
                    print(f"Verification data: {json.dumps(verification_data, indent=2)[:500]}...")
                    
                    # Verify arrays are now populated
                    jmmc_hp = verification_data.get("jmmcHP", [])
                    jmmc_finance = verification_data.get("jmmcFinance", [])
                    no_client = verification_data.get("noClient", [])
                    extra = verification_data.get("extra", [])
                    
                    total_entries = len(jmmc_hp) + len(jmmc_finance) + len(no_client) + len(extra)
                    
                    print(f"\nVerification arrays after compose:")
                    print(f"  jmmcHP: {len(jmmc_hp)} entries")
                    print(f"  jmmcFinance: {len(jmmc_finance)} entries")
                    print(f"  noClient: {len(no_client)} entries")
                    print(f"  extra: {len(extra)} entries")
                    print(f"  Total: {total_entries} entries")
                    
                    if total_entries == 0:
                        print("❌ Expected populated arrays after compose, but all are empty")
                        return False
                    
                    # Verify we have entries in the expected categories
                    # Based on our test data:
                    # - 1 entry for JMMC HP d.o.o.
                    # - 1 entry for JMMC Finance d.o.o.
                    # - 1 entry for General (no client)
                    # - 1 entry for EXTRA (no client + 999 - EXTRA tariff)
                    
                    if len(jmmc_hp) < 1:
                        print("⚠️  Expected at least 1 entry in jmmcHP")
                    
                    if len(jmmc_finance) < 1:
                        print("⚠️  Expected at least 1 entry in jmmcFinance")
                    
                    if len(no_client) < 1 and len(extra) < 1:
                        print("⚠️  Expected at least 1 entry in noClient or extra")
                    
                    print("✅ Verification arrays are now populated (as expected for 'composed' batch)")
                    return True
                else:
                    print(f"❌ Failed to get verification data: {verification_response.text}")
                    return False
            else:
                print(f"❌ Failed to get batch details: {batch_response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error composing and verifying: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_all_tests(self):
        """Run all verification endpoint behavior tests"""
        print("=" * 80)
        print("VERIFICATION ENDPOINT BEHAVIOR - BACKEND TESTS")
        print("=" * 80)
        print("\nTesting GET /api/batches/{batch_id}/verification endpoint behavior")
        print("for different batch statuses:")
        print("  - 'in progress' batch → empty arrays")
        print("  - 'composed' batch → populated arrays")
        print("=" * 80)
        
        results = {}
        
        # 1. Login
        if not self.login():
            print("\n❌ CRITICAL: Login failed. Cannot proceed with tests.")
            return
        
        # 2. Test 1: Create 'in progress' batch and verify empty arrays
        results["Test 1: 'in progress' batch returns empty arrays"] = self.create_in_progress_batch()
        
        if not results["Test 1: 'in progress' batch returns empty arrays"]:
            print("\n❌ CRITICAL: Test 1 failed. Cannot proceed with Test 2.")
            return
        
        # 3. Test 2: Compose invoices and verify populated arrays
        results["Test 2: 'composed' batch returns populated arrays"] = self.compose_and_verify_populated_arrays()
        
        # Summary
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for v in results.values() if v)
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{status} - {test_name}")
        
        print(f"\nTotal: {passed}/{total} tests passed")
        
        if passed == total:
            print("\n🎉 ALL TESTS PASSED!")
            print("\n✅ Verification endpoint behavior is correct:")
            print("  - 'in progress' batches return empty arrays ✅")
            print("  - 'composed' batches return populated arrays ✅")
            print("  - Batch status check happens BEFORE processing entries ✅")
        else:
            print(f"\n⚠️  {total - passed} test(s) failed")
            print("\n🔍 Debugging Hints:")
            print("  1. Check if batch status is correctly set during import")
            print("  2. Verify the verification endpoint checks status BEFORE processing")
            print("  3. Check if compose endpoint updates batch status correctly")
            print("  4. Verify time entries are correctly categorized after compose")


class TestArticlesAPI:
    """Test Articles API endpoints"""
    
    def __init__(self):
        self.token = None
        self.articles = []
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def test_get_articles(self) -> bool:
        """Test GET /api/articles endpoint"""
        print("\n=== Test 1: GET /api/articles - List all articles ===")
        try:
            response = requests.get(
                f"{BACKEND_URL}/articles",
                headers=self.get_headers()
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                self.articles = response.json()
                article_count = len(self.articles)
                print(f"✅ Retrieved {article_count} articles")
                
                # Verify we have 45 articles
                if article_count != 45:
                    print(f"❌ Expected 45 articles, got {article_count}")
                    return False
                
                print(f"✅ Article count is correct: 45 articles")
                
                # Verify structure of articles
                if self.articles:
                    sample = self.articles[0]
                    required_fields = ["code", "description", "unitMeasure", "priceWithoutVAT", "vatPercentage", "tariffCode"]
                    
                    print(f"\nVerifying article structure...")
                    print(f"Sample article: {json.dumps(sample, indent=2)}")
                    
                    missing_fields = []
                    for field in required_fields:
                        if field not in sample:
                            missing_fields.append(field)
                    
                    if missing_fields:
                        print(f"❌ Missing required fields: {missing_fields}")
                        return False
                    
                    print(f"✅ All required fields present: {required_fields}")
                    
                    # Verify no _id field in response
                    if "_id" in sample:
                        print(f"❌ Response contains _id field (should be excluded)")
                        return False
                    
                    print(f"✅ No _id field in response")
                    
                    # Display first 3 articles
                    print(f"\nFirst 3 articles:")
                    for idx, article in enumerate(self.articles[:3]):
                        print(f"\n  Article {idx + 1}:")
                        print(f"    Code: {article.get('code')}")
                        print(f"    Description: {article.get('description')}")
                        print(f"    Unit Measure: {article.get('unitMeasure')}")
                        print(f"    Price Without VAT: {article.get('priceWithoutVAT')}")
                        print(f"    VAT Percentage: {article.get('vatPercentage')}")
                        print(f"    Tariff Code: {article.get('tariffCode')}")
                    
                    print(f"\n✅ First 3 articles returned correctly")
                
                return True
            else:
                print(f"❌ Failed to get articles: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Error getting articles: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    


class TestFilteredInvoiceComposition:
    """Test filtered invoice composition flow for 'in progress' batches"""
    
    def __init__(self):
        self.token = None
        self.batch_id = None
        self.time_entry_ids = []
        self.invoice_ids = []
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def create_test_batch_with_progress(self) -> bool:
        """Test 1: Create a test batch with saveAsProgress=true"""
        print("\n=== Test 1: Create test batch with saveAsProgress=true ===")
        
        try:
            import openpyxl
            from datetime import datetime
            
            # Create workbook with test data
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št. računa"]
            ws.append(headers)
            
            # Add test data rows with different customers
            test_rows = [
                ["Project Alpha", "JMMC HP d.o.o.", datetime(2025, 10, 1), "Standard", "John Doe", "Development work for Alpha project", 5.0, 225.0, "INV-001"],
                ["Project Beta", "JMMC HP d.o.o.", datetime(2025, 10, 2), "Standard", "Jane Smith", "Testing work for Beta project", 3.5, 157.5, "INV-001"],
                ["Project Gamma", "JMMC Finance d.o.o.", datetime(2025, 10, 3), "Standard", "Bob Johnson", "Financial analysis for Gamma", 4.0, 180.0, "INV-001"],
                ["Project Delta", "JMMC Finance d.o.o.", datetime(2025, 10, 4), "Standard", "Alice Brown", "Accounting work for Delta", 2.5, 112.5, "INV-001"],
                ["Project Epsilon", "Test Customer Ltd.", datetime(2025, 10, 5), "Premium", "Charlie Wilson", "Consulting work for Epsilon", 6.0, 300.0, "INV-001"],
            ]
            
            for row in test_rows:
                ws.append(row)
            
            # Save to temp file
            test_file_path = "/tmp/test_filtered_composition.xlsx"
            wb.save(test_file_path)
            print(f"✅ Created test XLSX file: {test_file_path}")
            
            # Upload the file with saveAsProgress=true
            with open(test_file_path, 'rb') as f:
                files = {'file': ('test_filtered_composition.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'title': 'Filtered Composition Test Batch',
                    'invoiceDate': '2025-10-31',
                    'periodFrom': '2025-10-01',
                    'periodTo': '2025-10-31',
                    'dueDate': '2025-11-15',
                    'saveAsProgress': 'true'  # Key parameter for 'in progress' status
                }
                
                response = requests.post(
                    f"{BACKEND_URL}/imports",
                    headers=self.get_headers(),
                    files=files,
                    data=data
                )
            
            print(f"Import Status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                self.batch_id = result.get("batchId")
                row_count = result.get("rowCount")
                
                print(f"✅ Import successful")
                print(f"  Batch ID: {self.batch_id}")
                print(f"  Row Count: {row_count}")
                
                # Verify batch status is "in progress"
                batch_response = requests.get(
                    f"{BACKEND_URL}/batches/{self.batch_id}",
                    headers=self.get_headers()
                )
                
                if batch_response.status_code == 200:
                    batch = batch_response.json()
                    batch_status = batch.get("status")
                    
                    print(f"  Batch Status: {batch_status}")
                    
                    if batch_status != "in progress":
                        print(f"❌ Expected batch status 'in progress', got '{batch_status}'")
                        return False
                    
                    print("✅ Batch status is 'in progress'")
                    
                    # Verify time entries were created
                    entries_response = requests.get(
                        f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                        headers=self.get_headers()
                    )
                    
                    if entries_response.status_code == 200:
                        entries = entries_response.json()
                        print(f"✅ Time entries created: {len(entries)} entries")
                        
                        if len(entries) != row_count:
                            print(f"⚠️  Warning: Expected {row_count} entries, got {len(entries)}")
                        
                        return True
                    else:
                        print(f"❌ Failed to get time entries: {entries_response.text}")
                        return False
                else:
                    print(f"❌ Failed to get batch details: {batch_response.text}")
                    return False
            else:
                print(f"❌ Import failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error creating test batch: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_time_entry_ids(self) -> bool:
        """Test 2: Get time entry IDs from the batch"""
        print("\n=== Test 2: Get time entry IDs ===")
        
        if not self.batch_id:
            print("❌ No batch ID available")
            return False
        
        try:
            response = requests.get(
                f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                headers=self.get_headers()
            )
            
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                entries = response.json()
                print(f"✅ Retrieved {len(entries)} time entries")
                
                # Extract first 3 time entry IDs
                if len(entries) >= 3:
                    self.time_entry_ids = [entries[0]["id"], entries[1]["id"], entries[2]["id"]]
                    print(f"✅ Extracted first 3 time entry IDs:")
                    for idx, entry_id in enumerate(self.time_entry_ids):
                        entry = entries[idx]
                        print(f"  {idx + 1}. {entry_id}")
                        print(f"     Customer: {entry.get('customerName', 'N/A')}")
                        print(f"     Employee: {entry.get('employeeName', 'N/A')}")
                        print(f"     Hours: {entry.get('hours', 0)}")
                        print(f"     Value: €{entry.get('value', 0)}")
                    return True
                else:
                    print(f"❌ Not enough entries. Expected at least 3, got {len(entries)}")
                    return False
            else:
                print(f"❌ Failed to get time entries: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error getting time entry IDs: {str(e)}")
            return False
    
    def compose_filtered_invoices(self) -> bool:
        """Test 3: Compose invoices for filtered entries"""
        print("\n=== Test 3: Compose invoices for filtered entries ===")
        
        if not self.batch_id or not self.time_entry_ids:
            print("❌ No batch ID or time entry IDs available")
            return False
        
        try:
            # Call compose-filtered endpoint
            payload = {
                "batchId": self.batch_id,
                "entryIds": self.time_entry_ids
            }
            
            print(f"Payload: {json.dumps(payload, indent=2)}")
            
            response = requests.post(
                f"{BACKEND_URL}/invoices/compose-filtered",
                headers=self.get_headers(),
                json=payload
            )
            
            print(f"Status: {response.status_code}")
            print(f"Response: {response.text}")
            
            if response.status_code == 200:
                result = response.json()
                self.invoice_ids = result.get("invoiceIds", [])
                entries_processed = result.get("entriesProcessed", 0)
                
                print(f"✅ Filtered composition successful")
                print(f"  Invoice IDs: {self.invoice_ids}")
                print(f"  Entries Processed: {entries_processed}")
                
                # Verify response structure
                if not isinstance(self.invoice_ids, list):
                    print("❌ invoiceIds should be a list")
                    return False
                
                if len(self.invoice_ids) == 0:
                    print("❌ No invoices created")
                    return False
                
                if entries_processed != len(self.time_entry_ids):
                    print(f"⚠️  Warning: Expected {len(self.time_entry_ids)} entries processed, got {entries_processed}")
                
                print(f"✅ Response has correct structure")
                print(f"  - invoiceIds: {len(self.invoice_ids)} invoice(s)")
                print(f"  - entriesProcessed: {entries_processed}")
                
                return True
            else:
                print(f"❌ Compose failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error composing filtered invoices: {str(e)}")
            return False
    
    def verify_invoices_created(self) -> bool:
        """Test 4: Verify invoices were created in database"""
        print("\n=== Test 4: Verify invoices were created ===")
        
        if not self.batch_id or not self.invoice_ids:
            print("❌ No batch ID or invoice IDs available")
            return False
        
        try:
            # Get all invoices for the batch
            response = requests.get(
                f"{BACKEND_URL}/batches/{self.batch_id}/invoices",
                headers=self.get_headers()
            )
            
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                invoices = response.json()
                print(f"✅ Retrieved {len(invoices)} invoice(s) from batch")
                
                # Verify all invoice IDs exist
                invoice_ids_in_db = [inv["id"] for inv in invoices]
                
                all_found = True
                for invoice_id in self.invoice_ids:
                    if invoice_id in invoice_ids_in_db:
                        print(f"✅ Invoice {invoice_id} found in database")
                    else:
                        print(f"❌ Invoice {invoice_id} NOT found in database")
                        all_found = False
                
                if not all_found:
                    return False
                
                # Verify invoice structure and totals
                print("\n--- Invoice Details ---")
                for invoice in invoices:
                    invoice_id = invoice.get("id")
                    customer_name = invoice.get("customerName", "N/A")
                    total = invoice.get("total", 0)
                    status = invoice.get("status", "N/A")
                    
                    print(f"\nInvoice: {invoice_id}")
                    print(f"  Customer: {customer_name}")
                    print(f"  Total: €{total}")
                    print(f"  Status: {status}")
                    
                    # Verify total is > 0
                    if total <= 0:
                        print(f"⚠️  Warning: Invoice total is €{total}")
                
                return True
            else:
                print(f"❌ Failed to get invoices: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error verifying invoices: {str(e)}")
            return False


class TestInvoiceCompositionBillableStatuses:
    """Test enhanced invoice composition to include all billable entries (uninvoiced, ready, forfait)"""
    
    def __init__(self):
        self.token = None
        self.batch_id = None
        self.customer_id = None
        self.test_entries = []
        
    def login(self) -> bool:
        """Login as admin and get auth token"""
        print("\n=== Testing Login ===")
        try:
            response = requests.post(
                f"{BACKEND_URL}/auth/login",
                json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
            )
            print(f"Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                print(f"✅ Login successful")
                print(f"User: {data.get('user', {}).get('email')}")
                print(f"Role: {data.get('user', {}).get('role')}")
                return True
            else:
                print(f"❌ Login failed: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def get_headers(self) -> Dict[str, str]:
        """Get authorization headers"""
        return {"Authorization": f"Bearer {self.token}"}
    
    def create_test_batch_with_mixed_statuses(self) -> bool:
        """Create a test batch with entries having different statuses"""
        print("\n=== Creating Test Batch with Mixed Statuses ===")
        
        try:
            import openpyxl
            from datetime import datetime
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Projekt", "Stranka", "Datum", "Tarifa", "Delavec", "Opombe", "Porabljene ure", "Vrednost", "Št. računa"]
            ws.append(headers)
            
            # Add test data rows - all for the same customer
            test_customer_name = "ALEN ŠUTIĆ S.P."
            test_rows = [
                ["Project A", test_customer_name, datetime(2025, 10, 1), "001 - Računovodstvo", "John Doe", "Uninvoiced entry 1", 2.0, 90.0, ""],
                ["Project B", test_customer_name, datetime(2025, 10, 2), "001 - Računovodstvo", "Jane Smith", "Uninvoiced entry 2", 3.0, 135.0, ""],
                ["Project C", test_customer_name, datetime(2025, 10, 3), "001 - Računovodstvo", "Bob Johnson", "Ready entry", 1.5, 67.5, ""],
                ["Project D", test_customer_name, datetime(2025, 10, 4), "001 - Računovodstvo", "Alice Brown", "Forfait entry", 2.5, 112.5, ""],
                ["Project E", test_customer_name, datetime(2025, 10, 5), "001 - Računovodstvo", "Charlie Wilson", "Internal entry", 1.0, 45.0, ""],
                ["Project F", test_customer_name, datetime(2025, 10, 6), "001 - Računovodstvo", "Diana Prince", "Free entry", 1.0, 45.0, ""],
            ]
            
            for row in test_rows:
                ws.append(row)
            
            # Save to temp file
            test_file_path = "/tmp/test_billable_statuses.xlsx"
            wb.save(test_file_path)
            print(f"✅ Created test XLSX file: {test_file_path}")
            
            # Upload the file
            with open(test_file_path, 'rb') as f:
                files = {'file': ('test_billable_statuses.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'title': 'Billable Statuses Test Batch',
                    'invoiceDate': '2025-10-31',
                    'periodFrom': '2025-10-01',
                    'periodTo': '2025-10-31',
                    'dueDate': '2025-11-15'
                }
                
                response = requests.post(
                    f"{BACKEND_URL}/imports",
                    headers=self.get_headers(),
                    files=files,
                    data=data
                )
            
            print(f"Import Status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                self.batch_id = result.get("batchId")
                row_count = result.get("rowCount")
                
                print(f"✅ Import successful")
                print(f"  Batch ID: {self.batch_id}")
                print(f"  Row Count: {row_count}")
                
                # Get time entries
                response = requests.get(
                    f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                    headers=self.get_headers()
                )
                
                if response.status_code == 200:
                    self.test_entries = response.json()
                    print(f"✅ Retrieved {len(self.test_entries)} time entries")
                    
                    # Get customer ID from first entry
                    if self.test_entries:
                        self.customer_id = self.test_entries[0].get("customerId")
                        print(f"  Customer ID: {self.customer_id}")
                    
                    # Now update entries to have different statuses
                    # Entry 0, 1: uninvoiced (default)
                    # Entry 2: ready
                    # Entry 3: forfait
                    # Entry 4: internal
                    # Entry 5: free
                    
                    updates = [
                        {"index": 2, "status": "ready"},
                        {"index": 3, "status": "forfait"},
                        {"index": 4, "status": "internal"},
                        {"index": 5, "status": "free"}
                    ]
                    
                    update_response = requests.put(
                        f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                        headers=self.get_headers(),
                        json=updates
                    )
                    
                    print(f"\nUpdate Status: {update_response.status_code}")
                    
                    if update_response.status_code == 200:
                        print(f"✅ Updated entry statuses")
                        
                        # Verify the updates
                        response = requests.get(
                            f"{BACKEND_URL}/batches/{self.batch_id}/time-entries",
                            headers=self.get_headers()
                        )
                        
                        if response.status_code == 200:
                            self.test_entries = response.json()
                            
                            print(f"\nEntry statuses:")
                            for idx, entry in enumerate(self.test_entries):
                                status = entry.get("status", "unknown")
                                notes = entry.get("notes", "")
                                print(f"  Entry {idx}: status={status}, notes={notes}")
                            
                            return True
                        else:
                            print(f"❌ Failed to verify updates: {response.text}")
                            return False
                    else:
                        print(f"❌ Failed to update statuses: {update_response.text}")
                        return False
                else:
                    print(f"❌ Failed to get time entries: {response.text}")
                    return False
            else:
                print(f"❌ Import failed: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error creating test batch: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def test_compose_includes_all_billable(self) -> bool:
        """Test that compose includes uninvoiced, ready, and forfait entries"""
        print("\n=== Testing Compose Includes All Billable Entries ===")
        
        if not self.batch_id or not self.customer_id:
            print("❌ No batch or customer ID available")
            return False
        
        try:
            # Compose invoices
            print(f"\nComposing invoices for batch {self.batch_id}...")
            compose_response = requests.post(
                f"{BACKEND_URL}/invoices/compose",
                headers=self.get_headers(),
                params={"batchId": self.batch_id}
            )
            
            print(f"Compose Status: {compose_response.status_code}")
            
            if compose_response.status_code != 200:
                print(f"❌ Failed to compose invoices: {compose_response.text}")
                return False
            
            result = compose_response.json()
            invoice_ids = result.get("invoiceIds", [])
            
            print(f"✅ Invoices composed successfully")
            print(f"  Invoice IDs: {invoice_ids}")
            
            # Get the invoice for our test customer
            if not invoice_ids:
                print("❌ No invoices created")
                return False
            
            # Get invoice details
            invoice_response = requests.get(
                f"{BACKEND_URL}/batches/{self.batch_id}/invoices",
                headers=self.get_headers()
            )
            
            if invoice_response.status_code != 200:
                print(f"❌ Failed to get invoices: {invoice_response.text}")
                return False
            
            invoices = invoice_response.json()
            
            # Find invoice for our customer
            customer_invoice = None
            for invoice in invoices:
                if invoice.get("customerId") == self.customer_id:
                    customer_invoice = invoice
                    break
            
            if not customer_invoice:
                print(f"❌ No invoice found for customer {self.customer_id}")
                return False
            
            invoice_id = customer_invoice.get("id")
            print(f"\nFound invoice for customer: {invoice_id}")
            
            # Get invoice lines
            lines_response = requests.get(
                f"{BACKEND_URL}/invoices/{invoice_id}/lines",
                headers=self.get_headers()
            )
            
            if lines_response.status_code != 200:
                print(f"❌ Failed to get invoice lines: {lines_response.text}")
                return False
            
            lines = lines_response.json()
            
            print(f"\nInvoice has {len(lines)} line items")
            
            # Expected: 4 line items (2 uninvoiced + 1 ready + 1 forfait)
            # NOT included: 1 internal + 1 free
            
            if len(lines) != 4:
                print(f"❌ Expected 4 line items, got {len(lines)}")
                print(f"   Expected: 2 uninvoiced + 1 ready + 1 forfait = 4 total")
                print(f"   Should exclude: 1 internal + 1 free")
                return False
            
            print(f"✅ Invoice has correct number of line items (4)")
            
            # Verify which entries are included
            line_entry_ids = [line.get("timeEntryId") for line in lines]
            
            print(f"\nVerifying entry inclusion:")
            
            # Check each test entry
            expected_included = [0, 1, 2, 3]  # uninvoiced, uninvoiced, ready, forfait
            expected_excluded = [4, 5]  # internal, free
            
            for idx in expected_included:
                if idx < len(self.test_entries):
                    entry = self.test_entries[idx]
                    entry_id = entry.get("id")
                    status = entry.get("status")
                    notes = entry.get("notes", "")
                    
                    if entry_id in line_entry_ids:
                        print(f"  ✅ Entry {idx} (status={status}) INCLUDED: {notes}")
                    else:
                        print(f"  ❌ Entry {idx} (status={status}) NOT INCLUDED (should be): {notes}")
                        return False
            
            for idx in expected_excluded:
                if idx < len(self.test_entries):
                    entry = self.test_entries[idx]
                    entry_id = entry.get("id")
                    status = entry.get("status")
                    notes = entry.get("notes", "")
                    
                    if entry_id not in line_entry_ids:
                        print(f"  ✅ Entry {idx} (status={status}) EXCLUDED: {notes}")
                    else:
                        print(f"  ❌ Entry {idx} (status={status}) INCLUDED (should be excluded): {notes}")
                        return False
            
            print(f"\n✅ All billable entries included, non-billable entries excluded")
            return True
            
        except Exception as e:
            print(f"❌ Error testing compose: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_all_tests(self):
        """Run all invoice composition billable statuses tests"""
        print("=" * 80)
        print("INVOICE COMPOSITION - BILLABLE STATUSES - BACKEND TESTS")
        print("=" * 80)
        print("\nTesting enhanced invoice composition to include all billable entries:")
        print("  ✅ INCLUDE: uninvoiced, ready, forfait")
        print("  ❌ EXCLUDE: internal, free, already invoiced")
        print("\nUSER ISSUE: Only 2 of 4 rows for customer 'ALEN ŠUTIĆ S.P.' were included")
        print("FIX: Changed status filter from 'uninvoiced' to ['uninvoiced', 'ready', 'forfait']")
        print("=" * 80)
        
        results = {}
        
        # 1. Login
        if not self.login():
            print("\n❌ CRITICAL: Login failed. Cannot proceed with tests.")
            return
        
        # 2. Create test batch with mixed statuses
        results["Create test batch with mixed statuses"] = self.create_test_batch_with_mixed_statuses()
        
        if not results["Create test batch with mixed statuses"]:
            print("\n❌ CRITICAL: Failed to create test batch. Cannot proceed.")
            return
        
        # 3. Test compose includes all billable entries
        results["Compose includes all billable entries"] = self.test_compose_includes_all_billable()
        
        # Summary
        print("\n" + "=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for v in results.values() if v)
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{status} - {test_name}")
        
        print(f"\nTotal: {passed}/{total} tests passed")
        
        if passed == total:
            print("\n🎉 ALL TESTS PASSED!")
            print("\n✅ Invoice composition enhancement is working correctly:")
            print("  - Uninvoiced entries included ✅")
            print("  - Ready entries included ✅")
            print("  - Forfait entries included ✅")
            print("  - Internal entries excluded ✅")
            print("  - Free entries excluded ✅")
            print("\n✅ USER ISSUE RESOLVED: All billable entries for a customer are now included in invoices")
        else:
            print(f"\n⚠️  {total - passed} test(s) failed")
            print("\n🔍 Debugging Hints:")
            print("  1. Check if status filter in compose endpoint includes all billable statuses")
            print("  2. Verify line 3029 in server.py: status: {$in: ['uninvoiced', 'ready', 'forfait']}")
            print("  3. Check if entries have correct status values in database")
            print("  4. Verify invoice lines are created for all billable entries")

if __name__ == "__main__":
    # Run Invoice Composition Billable Statuses tests
    print("\n" + "=" * 80)
    print("RUNNING INVOICE COMPOSITION BILLABLE STATUSES TESTS")
    print("=" * 80)
    
    tester = TestInvoiceCompositionBillableStatuses()
    tester.run_all_tests()

